import pandas as pd
from fastapi import FastAPI, HTTPException, Request, Body, Query, Form, UploadFile, File, status
from fastapi.responses import JSONResponse, FileResponse, StreamingResponse
from typing import Dict, List, Optional, Tuple, Set, Any
from pathlib import Path
from pydantic import BaseModel, Json, Field, HttpUrl
import json
import logging
import shutil
import uuid
import math
import numpy as np
import openpyxl
from io import BytesIO
import re
from datetime import datetime, timedelta # Ensure datetime is imported
import os
from fastapi.middleware.cors import CORSMiddleware
import requests
import base64 # Import base64 for file decoding
from fastapi.staticfiles import StaticFiles # Import StaticFiles
import zipfile
import dateutil.parser
from dotenv import load_dotenv, find_dotenv
from collections import defaultdict 
import io
import httpx
import base64
import xml.etree.ElementTree as ET
from dotenv import dotenv_values
import time
import shutil
from pathlib import Path
import sys
import importlib.util


load_dotenv()
ORACLE_USERNAME = os.getenv("ORACLE_USERNAME")
ORACLE_PASSWORD = os.getenv("ORACLE_PASSWORD")
ORACLE_ENV = os.getenv("ORACLE_ENV") 

app = FastAPI(
    title="Excel Transformation API",
    description="API to transform and download Excel files.",
    version="1.0.0"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["X-Bundle-Filename", "Content-Disposition"]
)

UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
VALIDATION_RESULTS_DIR = UPLOAD_DIR / "validation_results"
VALIDATION_RESULTS_DIR.mkdir(parents=True, exist_ok=True)

COMPLETED_FOLDER = Path("validation/completed/Excel_Files")
COMPLETED_FOLDER.mkdir(parents=True, exist_ok=True) 
EXCEL_FILE_PATH = Path("Required_files/HDL_BO_Hierarchy_All_Objects_Charlie.xlsx")
TRANSFORMATION_ATTRIBUTES_FILE_PATH = Path("Required_files/Transformation - Common Attributes v3 2.xlsx")

# Configure logging
LOG_FILE_PATH = "server.log"    
logging.basicConfig(
    level=logging.DEBUG,  # <-- Only INFO and above will be logged
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE_PATH, mode='a', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


os.makedirs("static", exist_ok=True)
# "/home" for the static folder index.html and all those
app.mount("/home", StaticFiles(directory="static"), name="home")
@app.get("/home")
def read_root():
    return FileResponse("static/index.html")



pass_df = pd.DataFrame()
fail_df = pd.DataFrame()
def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [col.strip() for col in df.columns]
    df = df.dropna(subset=['Level-1']) 
    df = df.fillna("")
    return df

def extract_customer_instance_names_from_env(env_path: Path):
    combos = set()
    if env_path.exists():
        with env_path.open("r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, _ = line.split("=", 1)
                key = key.strip().strip('"').strip("'")
                m = re.match(r"([A-Za-z0-9]+)_([A-Za-z0-9]+)_", key)
                if m:
                    cust = m.group(1).capitalize()
                    inst = m.group(2).capitalize()
                    combos.add((cust, inst))
    return sorted(list(combos))

def read_and_normalize_excel(excel_path: Path):
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel not found: {excel_path}")
    xls = pd.ExcelFile(excel_path)
    df = None
    for sheet in xls.sheet_names:
        tmp = pd.read_excel(excel_path, sheet_name=sheet, dtype=str)
        tmp.columns = [str(c).strip() for c in tmp.columns]
        if tmp.shape[0] > 0 and tmp.dropna(how='all').shape[0] > 0:
            df = tmp
            break
    if df is None:
        df = pd.read_excel(excel_path, dtype=str)
        df.columns = [str(c).strip() for c in df.columns]
    df = df.fillna("")
    col_key_map = {}
    for col in df.columns:
        key = re.sub(r'[^a-z0-9]', '', str(col).lower())
        col_key_map[key] = col
    return df, col_key_map

def build_hierarchy_from_df(df: pd.DataFrame, col_map: dict, combos: list):
    """
    Build hierarchical tree from the cleaned dataframe + normalized col_map.

    - col_map is expected to be the normalized-key -> original-column-name map
      created by `read_and_normalize_excel`.
    - combos is list of (customer, instance) pairs (from .env).
    """
    import re
    # ---------- helpers ----------
    def normalize_key(s: str) -> str:
        if not s:
            return ""
        return re.sub(r'[^a-z0-9]', '', str(s).lower())

    def resolve_col(col_map_local: dict, *variants: str):
        """Try multiple human variants and return the original column name (or None)."""
        for v in variants:
            k = normalize_key(v)
            if k in col_map_local and col_map_local[k]:
                return col_map_local[k]
        return None

    def parse_mandatory(raw) -> bool:
        if raw is None:
            return False
        s = str(raw).strip().lower()
        if s in ("1", "1.0", "true", "t", "yes", "y", "required", "mandatory"):
            return True
        if s in ("0", "0.0", "false", "f", "no", "n", ""):
            return False
        # try numeric fallback
        try:
            return float(s) != 0.0
        except Exception:
            logging.warning(f"Unexpected Mandatory_Objects value: {raw!r}; defaulting to False")
            return False

    # ---------- resolve columns robustly ----------
    level_cols = {
        i: resolve_col(col_map,
                      f"Level-{i}", f"Level {i}", f"level{i}", f"level_{i}")
        for i in range(3, 11)
    }
    file_col = resolve_col(col_map, "File Name", "File", "Filename", "file")
    template_col = resolve_col(col_map, "Template Name", "Template", "templatename")
    mandatory_col = resolve_col(col_map, "Mandatory Objects", "Mandatory_Objects", "Mandatory")
    required_helper_col = resolve_col(col_map,
                                     "Required - Helper Text",
                                     "Required Helper Text",
                                     "RequiredHelperText",
                                     "Required_Helper_Text",
                                     "Required Helper",
                                     "Required")
    supported_helper_col = resolve_col(col_map,
                                      "Supported Action - Helper Text",
                                      "Supported Action Helper Text",
                                      "SupportedActionHelperText",
                                      "Supported_Helper",
                                      "Supported Action",
                                      "SupportedAction")

    logging.debug("Resolved column mapping in build_hierarchy_from_df: "
                  f"file={file_col}, template={template_col}, mandatory={mandatory_col}, "
                  f"required_helper={required_helper_col}, supported_helper={supported_helper_col}, "
                  f"level_cols={level_cols}")

    # ---------- collect variants ----------
    variants = []
    for _, row in df.iterrows():
        # Collect levels 3..10 values (use original column names if resolved)
        levels = []
        for i in range(3, 11):
            colname = level_cols.get(i)
            val = ""
            if colname:
                # row.get works on Series; fallback to empty string
                raw = row.get(colname, "")
                val = "" if pd.isna(raw) else str(raw).strip()
                if val.lower() == "nan":
                    val = ""
            levels.append(val)

        if not any(levels):
            continue

        file_val = ""
        if file_col:
            raw = row.get(file_col, "")
            file_val = "" if pd.isna(raw) else str(raw).strip()

        template_val = ""
        if template_col:
            raw = row.get(template_col, "")
            template_val = "" if pd.isna(raw) else str(raw).strip()

        mandatory_val = False
        if mandatory_col:
            mandatory_val = parse_mandatory(row.get(mandatory_col, ""))

        required_helper_val = ""
        if required_helper_col:
            raw = row.get(required_helper_col, "")
            required_helper_val = "" if pd.isna(raw) else str(raw).strip()

        supported_helper_val = ""
        if supported_helper_col:
            raw = row.get(supported_helper_col, "")
            supported_helper_val = "" if pd.isna(raw) else str(raw).strip()

        variants.append({
            "levels": levels,
            "file": file_val,
            "dat_template": template_val,
            "Mandatory_Objects": mandatory_val,
            "Required - Helper Text": required_helper_val,
            "Supported Action - Helper Text": supported_helper_val,
        })

    # ---------- tree builder with safe stamping ----------
    folder_roots = {}

    def ensure_all_fields(node: dict, extra_fields: dict, set_file_fields: bool = False):
        """Ensure the node has the expected structural and helper keys without clobbering good data."""
        keys = [
            "file", "dat_template",
            "level_1", "level_2", "level_3", "level_4",
            "level_5", "level_6", "level_7", "level_8", "level_9", "level_10",
            "Required - Helper Text", "Supported Action - Helper Text",
        ]
        for k in keys:
            # prefer explicit values from extra_fields (only set when non-empty),
            # but ensure the key exists on the node (default to empty string or False for Mandatory_Objects).
            incoming = extra_fields.get(k) if extra_fields is not None else None

            if k in ("file", "dat_template"):
                if set_file_fields and incoming:
                    node[k] = incoming
                elif k not in node:
                    node[k] = ""
            else:
                # text fields / levels / helper texts
                if incoming is not None and incoming != "":
                    node[k] = incoming
                else:
                    # if missing, ensure key exists but do not overwrite existing non-empty values
                    if k not in node:
                        node[k] = ""
        # Handle Mandatory_Objects separately
        if "Mandatory_Objects" not in node:
            node["Mandatory_Objects"] = False
        return node

    def get_or_create_node(parent_collection, node_name, is_root=False, extra_fields=None, set_file_fields=False):
        # parent_collection is either a dict (for roots) or list (for children)
        if is_root:
            if node_name not in parent_collection:
                parent_collection[node_name] = {"name": node_name, "children": []}
            if extra_fields:
                ensure_all_fields(parent_collection[node_name], extra_fields, set_file_fields)
            return parent_collection[node_name]
        else:
            for child in parent_collection:
                if child.get("name") == node_name:
                    if extra_fields:
                        ensure_all_fields(child, extra_fields, set_file_fields)
                    return child
            new_node = {"name": node_name, "children": []}
            if extra_fields:
                ensure_all_fields(new_node, extra_fields, set_file_fields)
            parent_collection.append(new_node)
            return new_node

    # ---------- assemble tree ----------
    for cust, inst in combos:
        for v in variants:
            try:
                last_nonempty_idx = max(idx for idx, nm in enumerate(v["levels"]) if nm)
            except ValueError:
                continue

            extra = {
                "file": v["file"],
                "dat_template": v["dat_template"],
                "level_1": cust,
                "level_2": inst,
                "level_3": v["levels"][0],
                "level_4": v["levels"][1],
                "level_5": v["levels"][2],
                "level_6": v["levels"][3],
                "level_7": v["levels"][4],
                "level_8": v["levels"][5],
                "level_9": v["levels"][6],
                "level_10": v["levels"][7],
                "Required - Helper Text": v["Required - Helper Text"],
                "Supported Action - Helper Text": v["Supported Action - Helper Text"],
                # We will handle Mandatory_Objects separately at the leaf
            }

            root = get_or_create_node(folder_roots, cust, is_root=True, extra_fields=extra)
            inst_node = get_or_create_node(root["children"], inst, extra_fields=extra)
            current = inst_node

            for i, name in enumerate(v["levels"]):
                if not name:
                    continue
                is_leaf = (i == last_nonempty_idx)
                current = get_or_create_node(
                    current["children"], name,
                    extra_fields=extra,
                    set_file_fields=is_leaf
                )
                if is_leaf:
                    # Explicitly set Mandatory_Objects only on the leaf node
                    if "Mandatory_Objects" in v:
                        current["Mandatory_Objects"] = bool(v["Mandatory_Objects"])
                    # Helper texts: only set if non-empty (do not overwrite existing helper text with empty)
                    req_ht = v.get("Required - Helper Text", "")
                    if req_ht:
                        current["Required - Helper Text"] = req_ht
                    sup_ht = v.get("Supported Action - Helper Text", "")
                    if sup_ht:
                        current["Supported Action - Helper Text"] = sup_ht

    # Return list of root nodes (preserves the structure expected by frontend)
    return list(folder_roots.values())

@app.get("/")
def root():
    """
    Root endpoint providing a welcome message.
    """
    return {"message": "Hit /api/utils/menu-items to get the 8-level hierarchy tree."}


USER_EXCEL_FILE_PATH = Path("Required_files/Users.xlsx")
def load_user_data(file_path: Path):
    try:
        # Assuming it's an Excel file that pandas can read, or a CSV named .xlsx
        # If it's truly an .xlsx file, ensure openpyxl is installed (`pip install openpyxl`)
        # If it's actually a CSV that happens to be named .xlsx, pandas will often handle it.
        # If it's a CSV, you might want to use pd.read_csv instead depending on actual file format
        df = pd.read_excel(file_path) # Changed to read_excel based on user's original filename
        
        users = {}
        for index, row in df.iterrows():
            if 'UserName' in row and 'Password' in row and 'UserType' in row:
                users[row['UserName']] = {
                    "password": row['Password'],
                    "usertype": row['UserType']
                }
            else:
                logging.warning(f"Row {index} in {file_path} is missing expected columns (UserName, Password, UserType).")
        return users
    except FileNotFoundError:
        logging.error(f"Error: User file not found at {file_path}")
        return {}
    except Exception as e:
        logging.error(f"Error loading user data from {file_path}: {e}")
        return {}

# Load user data when the application starts
USER_DB = load_user_data(USER_EXCEL_FILE_PATH)

class UserLogin(BaseModel):
    username: str
    password: str

@app.post("/api/utils/login-access")
async def login_access(user_login: UserLogin):
    """
    Login page code will fetch the details from the Users.xlsx file and check and send it back to the frontend.
    """
    username = user_login.username
    password = user_login.password

    if username in USER_DB and USER_DB[username]["password"] == password:
        user_type = USER_DB[username]["usertype"]
        return JSONResponse(
            content={
                "message": "Login successful",
                "username": username,
                "user_type": user_type
            },
            status_code=status.HTTP_200_OK
        )
    else:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid username or password"
        )

ENV_PATH = Path(".env")
@app.get("/api/utils/menu-items")
def get_hierarchy_api():
    # Validate Excel path
    if not EXCEL_FILE_PATH.exists():
        raise HTTPException(status_code=500, detail=f"Excel file not found at {EXCEL_FILE_PATH}")

    try:
        # Read Excel and get cleaned DataFrame + column key map
        df_cleaned, col_map = read_and_normalize_excel(EXCEL_FILE_PATH)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read Excel file: {str(e)}")

    # Extract customer-instance combos from .env
    combos = extract_customer_instance_names_from_env(ENV_PATH)
    if not combos:
        raise HTTPException(status_code=500, detail="No customer-instance combos found in .env")

    # Build hierarchy
    try:
        hierarchy_data = build_hierarchy_from_df(df_cleaned, col_map, combos)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to build hierarchy: {str(e)}")

    return {"hierarchy": hierarchy_data}
def get_columns_from_dat(file_bytes: bytes) -> Tuple[List[str], List[str], List[str]]:
    try:
        content = file_bytes.decode("utf-8-sig").splitlines()
        if not content:
            raise ValueError("The provided .dat file is empty.")

        header_line = content[0]
        all_columns = [col.strip() for col in header_line.split("|")]

        skipped_columns_list = []
        non_skipped_columns = []
        last_source_system_id_index = -1

        # Reverted to a simple substring check for "SourceSystemId"
        # This will correctly identify both "SourceSystemId" and "PersonId(SourceSystemId)"
        for i, col_name in enumerate(all_columns):
            if "(SourceSystemId)" in col_name:
                last_source_system_id_index = i
            if "SourceSystemId" in col_name:
                last_source_system_id_index = i  

        if last_source_system_id_index != -1:
            skipped_columns_list = all_columns[:last_source_system_id_index + 1]
            non_skipped_columns = all_columns[last_source_system_id_index + 1:]
        else:
            skipped_columns_list = []
            non_skipped_columns = all_columns

        return all_columns, non_skipped_columns, skipped_columns_list
    except Exception as e:
        logging.error(f"Failed to parse .dat file: {e}", exc_info=True)
        raise ValueError(f"Failed to parse .dat file: {str(e)}. Please ensure it's a valid pipe-separated file with a header row.")

@app.post("/api/hdl/upload-dat")
async def upload_dat_file(datFile: UploadFile = File(...)):
    if not datFile.filename.endswith(".dat"):
        raise HTTPException(status_code=400, detail="Invalid file format. Only .dat files are accepted.")

    dat_path = UPLOAD_DIR / "uploaded_dat.dat"

    try:
        dat_bytes = await datFile.read()
        all_columns, non_skipped_columns, skipped_columns = get_columns_from_dat(dat_bytes)

        with open(dat_path, "wb") as f:
            f.write(dat_bytes)

        # Ensure the response keys are consistent with the variable names
        return JSONResponse(content={
            "message": "DAT file uploaded successfully.",
            "datFileName": datFile.filename,
            "all_columns": all_columns,
            "non_skipped_columns": non_skipped_columns, # Renamed for clarity and consistency
            "skipped_columns": skipped_columns
        })

    except ValueError as ve:
        raise HTTPException(status_code=400, detail=str(ve))
    except Exception as e:
        logging.error(f"DAT file upload failed: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"DAT file upload failed: An unexpected server error occurred. Please try again.")

@app.post("/api/hdl/upload-excel")
async def upload_excel_file(excelFile: UploadFile = File(...)):
    """
    Uploads an .xlsx file and saves it.
    The file is saved as 'uploaded_excel.xlsx' in the UPLOAD_DIR.
    """
    if not excelFile.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Invalid file format. Upload an .xlsx file.")

    excel_path = UPLOAD_DIR / "uploaded_excel.xlsx"

    try:
        with open(excel_path, "wb") as f:
            shutil.copyfileobj(excelFile.file, f)

        return JSONResponse(content={
            "message": "Excel file uploaded successfully.",
            "excelFileName": excelFile.filename,
        })

    except Exception as e:
        logging.error(f"Excel file upload failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Excel file upload failed: {str(e)}")

@app.post("/api/hdl/upload")
async def upload_hdl_files(datFile: UploadFile = File(...), excelFile: UploadFile = File(...)):
    """
    Uploads DAT and Excel files, saves them, and returns .dat column names.
    This endpoint expects both files simultaneously. Use /api/hdl/upload-dat
    and /api/hdl/upload-excel for separate uploads.
    """
    if not datFile.filename.endswith(".dat") or not excelFile.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Invalid file format. Upload a .dat and .xlsx file.")

    dat_path = UPLOAD_DIR / "uploaded_file.dat"
    excel_path = UPLOAD_DIR / "uploaded_file.xlsx"

    try:
        dat_bytes = await datFile.read()
        columns = get_columns_from_dat(dat_bytes)
        with open(dat_path, "wb") as f:
            f.write(dat_bytes)

        # Save Excel file
        with open(excel_path, "wb") as f:
            shutil.copyfileobj(excelFile.file, f)

        return JSONResponse(content={
            "message": "Files uploaded successfully.",
            "datFileName": datFile.filename,
            "excelFileName": excelFile.filename,
            "columns_from_dat": columns
        })

    except Exception as e:
        logging.error(f"Combined upload failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Combined upload failed: {str(e)}")

@app.post("/api/hdl/{component}/save")
async def save_hdl_component(component: str, request: Request):
    """
    Save HDL component data sent from the frontend into a JSON file.
    Stores it in a temp folder under 'uploads/user/{component}/timestamped.json'.
    """
    try:
        payload = await request.json()

        # Create directories
        user_dir = UPLOAD_DIR / "user" / component
        user_dir.mkdir(parents=True, exist_ok=True)

        # Generate timestamped file name
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = user_dir / f"{component}_{timestamp}.json"

        # Save payload as JSON
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2)

        return {"message": f"Component '{component}' saved successfully.", "file": str(file_path)}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error saving component: {str(e)}")


def fetch_key_values(bo: str, attributes: list):
    """
    Fetches key values (Yes/No as boolean) for given component (bo) and attributes
    from the Key Values Mapping.xlsx file (single global file).
    """
    try:
        file_path = Path(__file__).parent / "Required_files" / "Key Values Mapping.xlsx"

        if not file_path.exists():
            logger.error(f"{file_path.name} not found in Required_files directory.")
            return {}

        df = pd.read_excel(file_path)

        # Normalize headers
        df.columns = df.columns.str.strip()
        column_map = {
            "BO HDL File Name": "GLOBAL_BO",
            "Component Name": "PARENT_BO",
            "HDL Attribute Name": "HDL_ATTRIBUTE_NAME",
            "Key_Values": "KEY_VALUES"
        }
        df = df.rename(columns={k: v for k, v in column_map.items() if k in df.columns})

        # Normalize columns
        for col in ["GLOBAL_BO", "PARENT_BO", "HDL_ATTRIBUTE_NAME", "KEY_VALUES"]:
            if col in df.columns:
                df[col] = df[col].fillna("").astype(str).str.strip()

        # Filter for this BO + attributes
        filtered = df[
            (df["PARENT_BO"].str.lower() == bo.lower()) &
            (df["HDL_ATTRIBUTE_NAME"].isin(attributes))
        ]

        # Build mapping: attr -> boolean
        key_values_map = {}
        for attr in attributes:
            row = filtered.loc[filtered["HDL_ATTRIBUTE_NAME"] == attr, "KEY_VALUES"]
            if not row.empty and row.iloc[0].lower() == "yes":
                key_values_map[attr] = True
            else:
                key_values_map[attr] = False

        return key_values_map

    except Exception as e:
        logger.error(f"Key values fetch failed: {str(e)}", exc_info=True)
        return {}


@app.post("/api/hdl/mandatory/batch")
def get_required_batch(
    bo: str = Body(..., alias="componentName"),
    attributes: list = Body(..., embed=True),
    customerName: str = Body(None, alias="customerName"),
    instanceName: str = Body(None, alias="instanceName"),
):
    """
    Retrieves mandatory attributes, helper text, and key values
    for a given component and a list of attributes.
    """
    try:
        if not customerName or not instanceName:
            return JSONResponse(
                status_code=400,
                content={"error": "customerName and instanceName are required."}
            )

        # Mandatory fields Excel
        file_path = Path(__file__).parent / "Required_files" / f"{customerName}_{instanceName}_MandatoryFields.xlsx"
        if not file_path.exists():
            return JSONResponse(
                status_code=404,
                content={"error": f"{file_path.name} not found."}
            )

        mandate = pd.read_excel(file_path)
        mandate.columns = mandate.columns.str.strip()
        column_map = {
            "BO HDL File Name": "GLOBAL_BO",
            "Component Name": "PARENT_BO",
            "HDL Attribute Name": "HDL_ATTRIBUTE_NAME",
            "Required": "REQUIRED",
            "Helper_Text": "HELPER_TEXT",
            "Data Type": "DATA_TYPE"
        }
        mandate = mandate.rename(columns={k: v for k, v in column_map.items() if k in mandate.columns})
        for col in ["GLOBAL_BO", "PARENT_BO", "HDL_ATTRIBUTE_NAME"]:
            if col in mandate.columns:
                mandate[col] = mandate[col].fillna("").astype(str).str.strip()
        if "HELPER_TEXT" in mandate.columns:
            mandate["HELPER_TEXT"] = mandate["HELPER_TEXT"].fillna("").astype(str).str.strip()
        else:
            mandate["HELPER_TEXT"] = ""

        if "DATA_TYPE" in mandate.columns:
            mandate["DATA_TYPE"] = mandate["DATA_TYPE"].fillna("").astype(str).str.strip()
        else:
            mandate["DATA_TYPE"] = ""
        if "REQUIRED" in mandate.columns:
            mandate["REQUIRED"] = mandate["REQUIRED"].fillna("").astype(str).str.strip()
        else:
            mandate["REQUIRED"] = "No"
        # Filter rows
        filtered = mandate[
            (mandate["PARENT_BO"].str.lower() == bo.lower()) &
            (mandate["HDL_ATTRIBUTE_NAME"].isin(attributes))
        ]

        # Fetch key values for this BO + attributes
        key_values_map = fetch_key_values(bo, attributes)

        # Build response dict
        mandatory_dict = {}
        for attr in attributes:
            row = filtered[filtered["HDL_ATTRIBUTE_NAME"] == attr]
            mandatory_dict[attr] = {
                "mandatory": row["REQUIRED"].iloc[0].strip().lower() == "yes" if not row.empty else False,
                "helper_text": row["HELPER_TEXT"].iloc[0] if not row.empty else "",
                "data_type": row["DATA_TYPE"].iloc[0] if not row.empty and "DATA_TYPE" in row.columns else "",
                "key_values": key_values_map.get(attr, [])
            }

        return {"mandatory": mandatory_dict}

    except Exception as e:
        logger.error(f"Mandatory fetch failed: {str(e)}", exc_info=True)
        return JSONResponse(
            status_code=500,
            content={"error": f"Mandatory fetch failed: {str(e)}"}
        )

@app.post("/api/hdl/lookup/batch")
def robust_lookup(
    bo: str = Body(..., alias="componentName"),
    global_bo: str = Body(None, alias="globalComponentName"),
    attributes: List[str] = Body(..., embed=True, alias="Attributes"),
    transaction: Optional[bool] = Body(False, alias="transaction"),
    customerName: str = Body(None, alias="customerName"),
    instanceName: str = Body(None, alias="instanceName"),
):
    """
    Retrieves lookup data for a given component and a list of attributes,
    with strict or fallback matching based on provided parameters.
    If transaction is True, always append results for only attribute name as well.
    """
    try:
        
        # Changed to read Excel since the provided file is an Excel sheet
        lookup_df = pd.read_excel(Path(__file__).parent/ "Required_files"/f"{customerName}_{instanceName}_LookupData.xlsx")

        logger.info(f"Lookup fetching started with global_bo='{global_bo}', bo='{bo}', attributes={attributes}, transaction={transaction}")

        # Normalize relevant columns for comparison
        for col in ["BO_NAME", "COMP_NAME", "HDL_Attribute_Name", "CODE_Name"]:
            if col in lookup_df.columns:
                # Explicitly fill NaN with empty string before other string operations
                lookup_df[col] = lookup_df[col].fillna('').astype(str).str.strip().str.lower()
            else:
                # Ensure the column exists, even if empty, to prevent KeyError
                lookup_df[col] = "" # Initialize with empty string for consistent comparison

        bo_name = global_bo.strip().lower() if global_bo else ""
        comp_name = bo.strip().lower() if bo else ""
        normalized_attrs = [attr.strip().lower() for attr in attributes]

        # --- Initial Filtering Logic ---
        # Strictest: all three columns
        if bo_name and comp_name:
            filtered_df = lookup_df[
                (lookup_df["BO_NAME"] == bo_name) &
                (lookup_df["COMP_NAME"] == comp_name) &
                (lookup_df["HDL_Attribute_Name"].isin(normalized_attrs))
            ]
        # Fallback: COMP_NAME + HDL_Attribute_Name
        elif comp_name:
            filtered_df = lookup_df[
                (lookup_df["COMP_NAME"] == comp_name) &
                (lookup_df["HDL_Attribute_Name"].isin(normalized_attrs))
            ]
        # Fallback: BO_NAME + HDL_Attribute_Name
        elif bo_name:
            filtered_df = lookup_df[
                (lookup_df["BO_NAME"] == bo_name) &
                (lookup_df["HDL_Attribute_Name"].isin(normalized_attrs))
            ]
        # Fallback: only HDL_Attribute_Name (when no BO or COMP name is provided)
        else:
            filtered_df = lookup_df[
                lookup_df["HDL_Attribute_Name"].isin(normalized_attrs)
            ]
        
        # --- Transaction Logic ---
        # Ensure transaction is a boolean
        if not isinstance(transaction, bool):
            raise ValueError("Transaction parameter must be a boolean value.")

        # If transaction is True, always append results where BO_NAME and COMP_NAME are empty
        if transaction:
            # Filter for transaction-specific lookups (where BO_NAME and COMP_NAME are empty)
            transaction_df = lookup_df[
                (lookup_df["HDL_Attribute_Name"].isin(normalized_attrs)) &
                (lookup_df["BO_NAME"] == "") &
                (lookup_df["COMP_NAME"] == "")
            ]
            logger.debug(f"Transaction DF before concat for attributes {normalized_attrs}: \n{transaction_df}") # Debug log
            
            # Concatenate only if transaction_df is not empty to avoid unnecessary operations
            if not transaction_df.empty:
                # Use pd.concat to combine the dataframes and drop duplicates
                # This ensures that if a row from transaction_df is already in filtered_df, it's not duplicated.
                filtered_df = pd.concat([filtered_df, transaction_df]).drop_duplicates().reset_index(drop=True)
        
        logger.debug(f"Final filtered DF before processing: \n{filtered_df}") # Debug log

        if filtered_df.empty:
            logger.info("No lookups found for the given criteria.")
            return {"lookups": {}, "default_code_names": {}}

        lookups = {}
        default_code_names = {}
        # Create a mapping from normalized attribute names back to their original case
        norm_to_orig = {attr.strip().lower(): attr for attr in attributes}

        for norm_attr in normalized_attrs:
            # Filter for rows corresponding to the current normalized attribute
            attr_rows = filtered_df[filtered_df["HDL_Attribute_Name"] == norm_attr]
            
            # Convert relevant columns to a list of dictionaries, filling NaN with empty string
            lookup_list = attr_rows[["CODE_Name", "Value", "Meaning", "Enabled_Flag", "Effective_Date"]].fillna("").to_dict(orient="records")
            
            if lookup_list:
                # Use the original attribute name for the key in the response
                lookups[norm_to_orig[norm_attr]] = lookup_list
                # Set the default_code_name from the first entry in the lookup list
                default_code_names[norm_to_orig[norm_attr]] = lookup_list[0]["CODE_Name"]

        logger.info("Lookup fetching completed successfully.")
        return {"lookups": lookups, "default_code_names": default_code_names}

    except FileNotFoundError:
        logger.error(f"{customerName}_{instanceName}_LookupData.xlsx not found. Please ensure it's in the 'Required_files' directory.", exc_info=True)
        return JSONResponse(status_code=404, content={"error": f"{customerName}_{instanceName}_LookupData.xlsx not found."})
    except ValueError as ve:
        logger.error(f"Validation error: {str(ve)}", exc_info=True)
        return JSONResponse(status_code=400, content={"error": f"Validation error: {str(ve)}"})
    except Exception as e:
        logger.error(f"Lookup failed: {str(e)}", exc_info=True)
        return JSONResponse(status_code=500, content={"error": f"Lookup failed: {str(e)}"})


class ColumnRequest(BaseModel):
    file_id: str
    columns: list[str]

@app.post("/api/get-excel-columns")
async def get_excel_columns(file: UploadFile = File(...), user_id: str = Form(...)):
    try:
        file_id = str(uuid.uuid4())
        file_path = os.path.join(UPLOAD_DIR, f"{file_id}.xlsx")
        with open(file_path, "wb") as f:
            content = await file.read()
            f.write(content)

        logger.info(f"Saved file to: {file_path}")
        df = pd.read_excel(file_path, engine="openpyxl")
        logger.info(f"Excel columns: {df.columns.tolist()}")
        return {
            "columns": df.columns.tolist(),
            "file_id": file_id
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing Excel file: {str(e)}")

@app.post("/api/hdl/apply-transformation-and-download", summary="Apply transformation and download Excel")
async def apply_transformation_and_download(
    attribute_column: str = Form(..., description="Name of the column to use for transformation (e.g., 'Suffix')"),
    raw_excel_file: UploadFile = File(..., description="Raw Excel file for transformation"),
):
    logger.info(f"Received request for file: {raw_excel_file.filename} with attribute column: {attribute_column}")

    try:
        contents = await raw_excel_file.read()
        logger.debug(f"Read {len(contents)} bytes from uploaded file.")
        excel_file_bytes = BytesIO(contents)

        try:
            df = pd.read_excel(excel_file_bytes, engine='openpyxl')
            logger.info(f"Successfully read Excel into DataFrame. Shape: {df.shape}")
            if df.empty:
                logger.warning("Uploaded Excel file is empty.")
                raise HTTPException(status_code=400, detail="Uploaded Excel file is empty or contains no data.")
        except pd.errors.EmptyDataError:
            logger.error("Pandas could not read data from the Excel file. It might be empty or malformed.")
            raise HTTPException(status_code=400, detail="Could not read data from Excel. File might be empty or corrupted.")
        except Exception as e:
            logger.exception("Error reading Excel file with pandas.")
            raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {e}")

        df_transformed = df.copy() 
        logger.debug("Starting data transformation.")

        for col in df_transformed.columns:
            if df_transformed[col].dtype == 'object':
                if pd.isna(df_transformed[col]).any():
                    df_transformed[col] = df_transformed[col].fillna('')
                    logger.debug(f"Replaced NaN with empty string in column: {col}")
            elif pd.api.types.is_numeric_dtype(df_transformed[col]):
                if pd.isna(df_transformed[col]).any() or np.isinf(df_transformed[col]).any():
                    df_transformed[col] = df_transformed[col].replace({np.nan: 0, np.inf: 1e308, -np.inf: -1e308})
                    logger.debug(f"Handled NaN/Inf in numeric column: {col}")

        if attribute_column in df_transformed.columns:
            def transform_value(value):
                clean_value = str(value).strip().lower() 
                if clean_value == 'senior':
                    return 'Sr'
                elif clean_value == 'junior':
                    return 'Jr'
                elif clean_value == '': 
                    return '' 
                else:
                    return value 
            
            # Apply the transformation function directly to the specified column
            df_transformed[attribute_column] = df_transformed[attribute_column].apply(transform_value)
            logger.info(f"Applied specific transformation directly to column: '{attribute_column}'.")
        else:
            logger.warning(f"Attribute column '{attribute_column}' not found in the Excel file. No in-place transformation performed.")
            df_transformed['Info'] = f"Attribute column '{attribute_column}' not found for in-place transformation."

        # --- End of specific transformation logic ---

        output_excel_file = BytesIO()
        try:
            df_transformed.to_excel(output_excel_file, index=False, engine='openpyxl')
            output_excel_file.seek(0)
            file_size = output_excel_file.getbuffer().nbytes
            logger.info(f"Successfully saved transformed DataFrame to BytesIO. File size: {file_size} bytes.")
            if file_size == 0:
                logger.error("Generated Excel file is empty (0 bytes).")
                raise HTTPException(status_code=500, detail="Generated Excel file is empty. Transformation might have failed.")
        except Exception as e:
            logger.exception("Error saving DataFrame to Excel BytesIO.")
            raise HTTPException(status_code=500, detail=f"Failed to generate Excel file: {e}")

        original_filename = raw_excel_file.filename
        base_filename = original_filename.split('/')[-1].split('\\')[-1]
        filename = f"transformed_{base_filename}"
        
        headers = {
            "Content-Disposition": f"attachment; filename=\"{filename}\"",
            "Content-Length": str(file_size)
        }
        
        logger.info(f"Sending transformed Excel file: {filename}")
        return StreamingResponse(
            output_excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    except HTTPException as e:
        raise e
    except Exception as e:
        logger.exception("An unhandled error occurred during file transformation.")
        raise HTTPException(status_code=500, detail=f"An unexpected server error occurred: {e}. Please try again or contact support.")


@app.get("/api/transform/get-mapping")
async def get_attribute_mapping(attribute: str = Query(..., description="The attribute for which to retrieve the mapping.")):
    """
    Retrieves Oracle Value as source and Customer Value as target for a given attribute from 'Transformation - Common Attributes v1.xlsx'.
    This endpoint is designed to provide data for React Flow to visualize mappings.
    Null values (NaN, None) will be returned as empty strings.
    If all retrieved mappings have empty source_value and target_value, an empty JSON object {} is returned.

    """
    try:
        if not TRANSFORMATION_ATTRIBUTES_FILE_PATH.exists():
            logging.error(f"Transformation mapping file not found at {TRANSFORMATION_ATTRIBUTES_FILE_PATH}")
            raise HTTPException(status_code=500, detail=f"Transformation mapping file not found at {TRANSFORMATION_ATTRIBUTES_FILE_PATH}")

        df = pd.read_excel(TRANSFORMATION_ATTRIBUTES_FILE_PATH, engine="openpyxl")
        df.columns = [col.strip() for col in df.columns] # Clean column names

        # Expected columns for filtering and output
        filter_col = 'Attributes for Transformation'
        source_col = 'Customer Value'
        target_col = 'Oracle Value'

        # Validate that the necessary columns exist in the Excel file
        required_cols = [filter_col, source_col, target_col]
        if not all(col in df.columns for col in required_cols):
            raise HTTPException(
                status_code=500,
                detail=f"Required columns '{filter_col}', '{source_col}', or '{target_col}' not found in "
                       f"'{TRANSFORMATION_ATTRIBUTES_FILE_PATH}'. Please ensure the Excel file has these columns."
            )

        # Filter for the specific attribute (case-insensitive and trimmed)
        filtered_df = df[df[filter_col].astype(str).str.strip().str.lower() == attribute.strip().lower()]

        # Prepare the list of mappings for the frontend
        mappings = []
        for index, row in filtered_df.iterrows():
            # Get source and target values
            source_val = row[source_col]
            target_val = row[target_col]

            # Convert NaN or None to empty string
            processed_source_val = ""
            if source_val is not None and (not (isinstance(source_val, float) and math.isnan(source_val))):
                processed_source_val = str(source_val)

            processed_target_val = ""
            if target_val is not None and (not (isinstance(target_val, float) and math.isnan(target_val))):
                processed_target_val = str(target_val)
            
            mappings.append({
                "source_value": processed_source_val,
                "target_value": processed_target_val
            })
        
        # Check if the mappings list is empty or if all entries have empty source and target values
        if not mappings or all(item["source_value"] == "" and item["target_value"] == "" for item in mappings):
            return JSONResponse(content={}, status_code=200)
        
        return JSONResponse(content=mappings, status_code=200)

    except Exception as e:
        logging.error(f"Error fetching attribute mapping for '{attribute}': {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error fetching attribute mapping: {str(e)}")


# Define the path to your transformation attributes Excel file
TRANSFORMATION_ATTRIBUTES_FILE_PATH = Path("./Required_files/Transformation - Common Attributes v3 2.xlsx")

# Ensure the Required_files directory exists
TRANSFORMATION_ATTRIBUTES_FILE_PATH.parent.mkdir(parents=True, exist_ok=True)


# --- Pydantic Model for Request Body ---
class BulkTransformationPayload(BaseModel):
    componentName: str
    transformations: Dict[str, str] # Dictionary for 'Attribute for transformation' -> 'Transformation map name'
    # NEW FIELD: Dictionary for 'Customer Value' -> 'Oracle Value' replacements
    customer_oracle_replacements: Optional[Dict[str, str]] = {}

# --- API Endpoint ---
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %levelname)s - %message)s')

# --- Updated Payload Model ---
class BulkTransformationPayload(BaseModel):
    """
    Payload for bulk data transformation mappings and triggering customer value transformations.
    """
    componentName: str
    transformations: Dict[str, str] # e.g., {"Attribute A": "Map1", "Attribute B": "Map2"}
    # New field: List of attribute names for which Customer Value should be replaced by Oracle Value
    attributes_for_customer_value_transformation: Optional[List[str]] = None

@app.post("/api/hdl/transform-customer-excel")
async def transform_customer_excel(
    raw_excel_file: UploadFile = File(..., description="Customer Excel file to be transformed")
):
    """
    Receives a customer's Excel file, applies transformations based on the 
    'Transformation - Common Attributes v3 2.xlsx' file, and returns the
    modified Excel file for download.
    """
    try:
        # 1. Load the transformation rules from the master Excel file on the server
        # It's good practice to ensure this file exists before trying to read it
        if not TRANSFORMATION_ATTRIBUTES_FILE_PATH.exists():
            logging.error(f"Transformation rules file not found at {TRANSFORMATION_ATTRIBUTES_FILE_PATH}")
            raise HTTPException(
                status_code=500, 
                detail=f"Transformation rules file not found at {TRANSFORMATION_ATTRIBUTES_FILE_PATH}. Please ensure it's in the 'Required_files' directory."
            )
        
        rules_df = pd.read_excel(TRANSFORMATION_ATTRIBUTES_FILE_PATH, engine='openpyxl')
        rules_df.columns = [col.strip() for col in rules_df.columns]
        
        # Ensure required columns exist in the rules file
        required_rule_cols = ['Attributes for Transformation', 'Customer Value', 'Oracle Value']
        if not all(col in rules_df.columns for col in required_rule_cols):
            raise HTTPException(status_code=500, detail="Transformation rules file is missing required columns. Expected: 'Attributes for Transformation', 'Customer Value', 'Oracle Value'.")

        # 2. Load the uploaded customer Excel file into a pandas DataFrame
        # FIX: Read the content into BytesIO first
        contents = await raw_excel_file.read()
        excel_in_memory = BytesIO(contents)
        
        # First read without header to check number of rows
        temp_df = pd.read_excel(excel_in_memory, engine='openpyxl', header=None)
        if len(temp_df) < 2:  # Check if file has at least 2 rows (header + data)
            raise HTTPException(
                status_code=400,
                detail="Excel file must contain at least 2 rows: one header row and at least one data row."
            )
        
        # Reset file pointer and read with proper header
        excel_in_memory.seek(0)
        customer_df = pd.read_excel(excel_in_memory, engine='openpyxl', header=1)
        # Ensure column names are strings before stripping
        customer_df.columns = [str(col).strip() for col in customer_df.columns]
        logging.info(f"Successfully read customer Excel file. Shape: {customer_df.shape}")

        # 3. Apply transformations column by column
        for column_to_transform in customer_df.columns:
            # Find the transformation rules for the current column
            specific_rules = rules_df[rules_df['Attributes for Transformation'] == column_to_transform]
            
            if not specific_rules.empty:
                # Create a mapping dictionary: { 'Customer Value': 'Oracle Value' }
                # Drop rows where 'Customer Value' is empty to avoid incorrect mapping
                specific_rules = specific_rules.dropna(subset=['Customer Value'])
                
                # Convert both keys and values to string before creating Series/dict
                # This handles cases where Excel might interpret values as numbers
                value_map = pd.Series(
                    specific_rules['Oracle Value'].astype(str).values,
                    index=specific_rules['Customer Value'].astype(str)
                ).to_dict()

                if value_map:
                    # Apply the mapping to the column in the customer's DataFrame
                    # Ensure values in customer_df column are also strings for consistent replacement
                    customer_df[column_to_transform] = customer_df[column_to_transform].astype(str).replace(value_map)
                    logging.info(f"Applied transformation to column: '{column_to_transform}' using map: {value_map}")
                else:
                    logging.info(f"No valid value mappings found for column '{column_to_transform}' in the transformation rules.")
            else:
                logging.debug(f"No specific transformation rules found for column: '{column_to_transform}'. Skipping.")


        # 4. Save the transformed DataFrame to an in-memory Excel file
        output_excel_file = BytesIO()
        try:
            customer_df.to_excel(output_excel_file, index=False, engine='openpyxl')
            output_excel_file.seek(0) # Rewind the buffer to the beginning
            file_size = output_excel_file.getbuffer().nbytes
            logging.info(f"Successfully created transformed Excel file in memory. Size: {file_size} bytes.")
            if file_size == 0:
                logging.warning("Generated transformed Excel file is empty (0 bytes).")
                raise HTTPException(status_code=500, detail="Transformed Excel file is empty. Transformation might have resulted in no data.")
        except Exception as save_error:
            logging.error(f"Error saving transformed DataFrame to in-memory Excel: {save_error}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"Failed to generate transformed Excel file: {save_error}")


        # 5. Return the in-memory file as a downloadable response
        original_filename = raw_excel_file.filename
        # Sanitize filename to prevent directory traversal or other issues
        safe_original_filename = Path(original_filename).name
        output_filename = f"transformed_{safe_original_filename}"
        
        headers = {
            "Content-Disposition": f"attachment; filename=\"{output_filename}\"",
            "Content-Length": str(file_size) # Set Content-Length header
        }
        
        logging.info(f"Sending transformed Excel file: {output_filename}")
        return StreamingResponse(
            output_excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    except HTTPException as http_exc:
        # Re-raise HTTPExceptions as they contain specific error details and status codes
        raise http_exc
    except Exception as e:
        # Catch any other unexpected errors and log them
        logging.error(f"An unexpected error occurred during transform_customer_excel: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"An unexpected server error occurred during file transformation: {str(e)}")


@app.post("/api/hdl/nlp/validate")
async def proxy_nlp_validate(request: Request):
    """
    Proxies the request to the Flask NLP validation service at /validate.
    Accepts multipart/form-data with 'csv_file' and 'validation_file'.
    """
    try:
        # Read the incoming form data
        form = await request.form()
        files = {}
        for key in form:
            file = form[key]
            if hasattr(file, 'filename'):
                files[key] = (file.filename, await file.read(), file.content_type)
        # Proxy to Flask app (assume running on localhost:9000)
        flask_url = "http://localhost:9000/validate"
        response = requests.post(flask_url, files=files)
        return JSONResponse(status_code=response.status_code, content=response.json())
    except Exception as e:
        logger.error(f"Error proxying to NLP validation service: {e}")
        raise HTTPException(status_code=500, detail=f"Proxy error: {e}")

@app.post("/api/hdl/save_code")
async def save_code(request: Request):
    """
    Receives Python code, attribute, rules, and conditions from the chatbot component, saves the code as a .py file named after the attribute, and updates an Excel sheet with the rules and conditions.
    """
    try:
        import openpyxl
        payload = await request.json()
        code = payload.get("code", "")
        componentName = payload.get("component_name", "Default")
        attribute = payload.get("attribute", "")
        rules = payload.get("rules", [])
        conditions = payload.get("conditions", [])
        customerName = payload.get("customerName","")
        instanceName = payload.get("instanceName", "")
        # Sanitize attribute for filename
        safe_attribute = "_".join(attribute.strip().split()) or "untitled"
        save_dir = UPLOAD_DIR / "saved_code"
        save_dir.mkdir(parents=True, exist_ok=True)
        filename = f"{customerName}_{instanceName}_{componentName}.py"
        file_path = save_dir / filename
        # Save only the code as a .py file
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(code)
        # --- Update Excel sheet ---
        excel_path = Path("Required_files/Available_NLP.xlsx")
        if excel_path.exists():
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            found = False
            # Ensure header has 'File Name' column
            header = [cell.value for cell in ws[1]]
            if 'File Name' not in header:
                ws.cell(row=1, column=len(header)+1, value='File Name')
                header.append('File Name')
            file_name_col = header.index('File Name') + 1
            for row in ws.iter_rows(min_row=2):
                cell_attr = str(row[0].value).strip() if row[0].value else ""
                if cell_attr.lower() == attribute.strip().lower():
                    row[1].value = ", ".join(rules) if rules else ""
                    row[2].value = ", ". join(conditions) if conditions else ""
                    row[file_name_col-1].value = filename
                    found = True
                    break
            if not found:
                new_row = [attribute, ", ".join(rules), ", ". join(conditions)]
                # Pad with empty cells if needed
                while len(new_row) < file_name_col-1:
                    new_row.append("")
                new_row.append(filename)
                ws.append(new_row)
            wb.save(excel_path)
        return {"success": True, "message": "Python code and Excel updated successfully.", "file": str(file_path)}
    except Exception as e:
        logger.error(f"Error saving code from chatbot: {e}")
        return {"success": False, "error": str(e)}


@app.post("/api/hdl/nlr/batch")
def get_nlr_rules_batch(
    attributes: List[str] = Body(..., embed=True, alias="attributes")
):
    """
    Retrieves NLR rules for a list of attributes from the Excel file 'Required_files/Available_NLP.xlsx'.
    Returns a dictionary mapping each attribute to its rules and conditions, and a 'has_rules' boolean for each attribute.
    """
    try:
        import openpyxl
        excel_path = Path("Required_files/Available_NLP.xlsx")
        if not excel_path.exists():
            return JSONResponse(status_code=404, content={"error": "NLR rules file not found."})

        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        # Build a dict of all available rules
        rules_dict = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            attribute = str(row[0]).strip() if row[0] else ""
            if attribute:
                rules = str(row[1]).strip() if row[1] else ""
                conditions = str(row[2]).strip() if row[2] else ""
                rules_list = rules.split(", ") if rules else []
                conditions_list = conditions.split(", ") if conditions else []
                rules_dict[attribute] = {
                    "rules": rules_list,
                    "conditions": conditions_list
                }

        # Filter for requested attributes only and add has_rules
        filtered = {}
        for attr in attributes:
            entry = rules_dict.get(attr, {"rules": [], "conditions": []})
            filtered[attr] = {
                **entry,
                "has_rules": bool(entry["rules"] and any(r.strip() for r in entry["rules"]))
            }

        return {
            "nlr_rules": filtered,
            "available_attributes": list(rules_dict.keys())
        }

    except Exception as e:
        return JSONResponse(status_code=500, content={"error": f"NLR rules batch lookup failed: {str(e)}"})

@app.get("/api/hdl/get_rules")
def get_nlr_rules_for_attribute(attribute: str = Query(..., description="The attribute (HDL) to fetch NLR rules for.")):
    """
    Returns NLR rules and conditions for a single attribute from 'Required_files/Available_NLP.xlsx'.
    """
    try:
        import openpyxl
        excel_path = Path("Required_files/Available_NLP.xlsx")
        if not excel_path.exists():
            return JSONResponse(status_code=404, content={"error": "NLR rules file not found."})

        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            attr = str(row[0]).strip() if row[0] else ""
            if attr.lower() == attribute.strip().lower():
                rules = str(row[1]).strip() if row[1] else ""
                conditions = str(row[2]).strip() if row[2] else ""
                rules_list = rules.split(", ") if rules else []
                conditions_list = conditions.split(", ") if conditions else []
                return {
                    "attribute": attr,
                    "rules": rules_list,
                    "conditions": conditions_list
                }
        # If not found
        return {
            "attribute": attribute,
            "rules": [],
            "conditions": []
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": f"NLR rules lookup failed: {str(e)}"})
    

# DIR for bulk excel upload (redefined for clarity, or can use UPLOAD_DIR)
DIR = Path("uploads/Excel_Files")
DIR.mkdir(parents=True, exist_ok=True) # Ensure this directory exists at startup

def populate_actual_termination_date_from_resignation(file_path: Path, header_row_num: int = 2, save_as_new_file=False, hireActions: List = [], globalTransfers: List = [], termAction: List = []):
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        headers = [cell.value for cell in ws[header_row_num]]
        logger.info(f"[INFO] Headers: {headers}")
        combined_hire_glb = hireActions + globalTransfers
        # Ensure ActualTerminationDate exists
        if "ActualTerminationDate" not in headers:
            headers.append("ActualTerminationDate")
            ws.cell(row=header_row_num, column=len(headers)).value = "ActualTerminationDate"

        try:
            person_idx = headers.index("PersonNumber")
            action_idx = headers.index("ActionCode")
            eff_start_idx = headers.index("EffectiveStartDate")
            actual_term_idx = headers.index("ActualTerminationDate")
        except ValueError as ve:
            logger.info("[ERROR] Required headers not found:", ve)
            return

        # Map of PersonNumber -> all rows (sorted)
        person_rows = {}

        for i, row in enumerate(ws.iter_rows(min_row=header_row_num+1), start=header_row_num+1):
            person = str(row[person_idx].value).strip() if row[person_idx].value else None
            if not person:
                continue
            if person not in person_rows:
                person_rows[person] = []
            person_rows[person].append((i, row))

        for person, rows in person_rows.items():
            # Sort rows by EffectiveStartDate
            sorted_rows = sorted(rows, key=lambda r: r[1][eff_start_idx].value)

            open_cycle_indices = []
            cycle_start_idx = None

            for idx, row in sorted_rows:
                action = str(row[action_idx].value).strip().upper() if row[action_idx].value else ""
                date_val = row[eff_start_idx].value

                if action in combined_hire_glb:
                    cycle_start_idx = idx
                    open_cycle_indices = [idx]  # start fresh
                elif action in termAction:
                    if cycle_start_idx:
                        resignation_date = date_val
                        termination_date = resignation_date - timedelta(days=1)

                        for sub_idx, sub_row in sorted_rows:
                            if cycle_start_idx <= sub_idx < idx:
                                ws.cell(row=sub_idx, column=actual_term_idx + 1, value=termination_date)
                        cycle_start_idx = None  # reset for next cycle
                elif cycle_start_idx:
                    open_cycle_indices.append(idx)

        # Save
        if save_as_new_file:
            out_file = file_path.parent / f"{file_path.stem}.xlsx"
            wb.save(out_file)
            logger.info(f"[DONE] ✅ Saved as: {out_file.name}")
            return out_file
        else:
            wb.save(file_path)
            logger.info("[DONE] ✅ Overwritten the original file.")
            return file_path

    except Exception as e:
        logger.info(f"[ERROR] ❌ Something went wrong: {e}")
        return None
    
def parse_excel_date(val):
    if isinstance(val, datetime):
        return val
    try:
        return datetime.strptime(str(val).split(' ')[0], '%Y-%m-%d')
    except Exception:
        return None


def validate_termination_date(file_path: Path, header_row_num: int = 2, save_as_new_file: bool = True, TermActions: List[str] = [], HireActions: List[str] = []):
    if TermActions is None:
        logger.info("[INFO] No Termination Actions provided")
        return
    if HireActions is None:
        logger.info("[INFO] No Hire Actions provided")
        return

    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        logger.info(f"[VALIDATION] 🚦 Checking termination logic on {file_path.name}")

        headers = [cell.value for cell in ws[header_row_num]]
        person_idx = headers.index("PersonNumber")
        action_idx = headers.index("ActionCode")
        eff_start_idx = headers.index("EffectiveStartDate")

        start_row = header_row_num + 1
        rows_to_delete = set()
        person_rows = {}

        for i, row in enumerate(ws.iter_rows(min_row=start_row), start=start_row):
            person = str(row[person_idx].value).strip() if row[person_idx].value else None
            if not person:
                continue
            if person not in person_rows:
                person_rows[person] = []
            person_rows[person].append((i, row))

        for person, rows in person_rows.items():
            hire_dates = []
            termination_rows = []

            for idx, row in rows:
                action = str(row[action_idx].value).strip().upper() if row[action_idx].value else ""
                date_val = parse_excel_date(row[eff_start_idx].value)
                if action in HireActions and date_val:
                    hire_dates.append(date_val)
                if action in TermActions and date_val:
                    termination_rows.append((idx, date_val))

            if termination_rows:
                if not hire_dates:
                    logger.info(f"[INVALID] ❌ {person} has termination but no hire record. Removing all rows.")
                    rows_to_delete.update(idx for idx, _ in rows)
                    continue

                earliest_hire = min(hire_dates)
                for term_idx, term_date in termination_rows:
                    if term_date < earliest_hire:
                        logger.info(f"[INVALID] ❌ {person} termination date {term_date} < hire date {earliest_hire}. Removing all rows.")
                        rows_to_delete.update(idx for idx, _ in rows)
                        break


        if rows_to_delete:
            logger.info(f"[CLEANUP] 🧹 Total Rows Marked for Deletion: {len(rows_to_delete)}")
        else:
            logger.info("[CLEANUP] ✅ No invalid termination records found.")

        deleted_count = 0
        for row_num in sorted(rows_to_delete, reverse=True):
            if row_num > header_row_num:
                ws.delete_rows(row_num)
                deleted_count += 1

        logger.info(f"[DELETED] 🗑️ Total Rows Deleted: {deleted_count}")

        if save_as_new_file:
            out_path = file_path.parent / f"{file_path.stem}_validated.xlsx"
            wb.save(out_path)
            logger.info(f"[DONE] ✅ Saved as {out_path.name}")
            return out_path
        else:
            wb.save(file_path)
            logger.info(f"[DONE] ✅ Overwritten {file_path.name}")
            return file_path

    except Exception as e:
        logger.error(f"[ERROR] ❌ Validation failed: {e}")
        return None

def validate_workrelationship_sheet(file_path: Path, header_row_num: int = 2, save_as_new_file: bool = True, hireActions: List = [], globalTransfers: List = []):
    """
    Filters the Excel file to keep:
    ✅ ONLY HIRE / REHIRE / GLB_TRANSFER (aka GLOBALTRANSFER)
    ✅ The earliest HIRE
    ✅ The latest REHIRE or GT in each cycle
    ❌ All other actions (PROMOTION, ASG_CHANGE, RESIGNATION, etc) are discarded
    """

    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        headers = [cell.value for cell in ws[header_row_num]]
        logger.info(f"[INFO] Headers found in row {header_row_num}: {headers}")

        try:
            person_idx = headers.index("PersonNumber")
            action_idx = headers.index("ActionCode")
            date_idx = headers.index("EffectiveStartDate")
        except ValueError:
            logger.info(f"[WARN] Required headers missing in {file_path.name}")
            return

        start_row = header_row_num + 1
        person_rows = {}

        for i, row in enumerate(ws.iter_rows(min_row=start_row), start=start_row):
            person = str(row[person_idx].value).strip() if row[person_idx].value else None
            if not person:
                continue
            if person not in person_rows:
                person_rows[person] = []
            person_rows[person].append((i, row))

        allowed_starts = hireActions + globalTransfers
        keep_rows = set()

        for person, rows in person_rows.items():
            sorted_rows = sorted(rows, key=lambda r: r[1][date_idx].value)
            current_start = None
            latest_action_row = None

            for idx, row in sorted_rows:
                action = str(row[action_idx].value).strip().upper() if row[action_idx].value else ""

                if action not in allowed_starts:
                    continue

                # New cycle detected
                if action in hireActions or globalTransfers:
                    if current_start is not None:
                        keep_rows.add(current_start)
                        if latest_action_row and latest_action_row != current_start:
                            keep_rows.add(latest_action_row)
                    current_start = idx
                    latest_action_row = idx
                else:
                    # Should never hit this due to earlier if guard
                    continue

            # End of loop – commit the final cycle
            if current_start:
                keep_rows.add(current_start)
                if latest_action_row and latest_action_row != current_start:
                    keep_rows.add(latest_action_row)

        logger.info(f"[INFO] Keeping strictly HIRE/REHIRE/GT only rows: {sorted(keep_rows)}")

        # Delete non-matching rows bottom-up for all that person numbers
        for row_num in range(ws.max_row, header_row_num, -1):
            if row_num not in keep_rows:
                ws.delete_rows(row_num)

        # Save
        if save_as_new_file:
            output_path = file_path.parent / f"{file_path.stem}.xlsx"
            wb.save(output_path)
            logger.info(f"[DONE] ✅ Saved as: {output_path.name}")
            return output_path
        else:
            wb.save(file_path)
            logger.info(f"[DONE] ✅ Overwritten: {file_path.name}")
            return file_path

    except Exception as e:
        logger.info(f"[ERROR] ❌ Failed while strictly filtering WorkRelationship: {e}")

def Assignment_type_Code(file_path: Path, assignment_status_rules: list, header_row_num: int = 2, save_as_new_file: bool = True):
    """
    Validates AssignmentStatusTypeCode using regex to extract first word.
    Logs all first words found. Matches based on assignment_status_rules.
    """
    try:
        logger.info(f"[INFO] Processing Assignment_type_Code for file: {file_path.name}")
        logger.info(f"[INFO] Assignment status rules received: {assignment_status_rules}")

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        headers = [cell.value for cell in ws[header_row_num]]
        logger.info(f"[INFO] Headers found at row {header_row_num}: {headers}")

        try:
            person_idx = headers.index("PersonNumber")
            action_idx = headers.index("ActionCode")
            assignment_type_idx = headers.index("AssignmentStatusTypeCode")
        except ValueError as ve:
            logger.error(f"[ERROR] Required columns missing in {file_path.name}: {ve}")
            return None

        start_row = header_row_num + 1
        valid_rows = []
        first_words_logged = set()

        for i, row in enumerate(ws.iter_rows(min_row=start_row), start=start_row):
            person = str(row[person_idx].value).strip() if row[person_idx].value else None
            if not person:
                continue

            action = str(row[action_idx].value).strip().upper() if row[action_idx].value else ""
            assignment_status_full = str(row[assignment_type_idx].value).strip().upper() if row[assignment_type_idx].value else ""

            # Extract first word using regex (match till space, dash, or non-word char)
            match = re.match(r'^([A-Z0-9]+)', assignment_status_full)
            assignment_status_first_word = match.group(1) if match else ""

            first_words_logged.add(assignment_status_first_word)

            matched = False
            for rule in assignment_status_rules:
                rule_key = rule.get("key", "").strip().lower()
                rule_value = rule.get("value", "").upper()
                rule_result = rule.get("result", "").upper()

                if rule_key == "else":
                    # ELSE rule when nothing matched before
                    expected_status = rule_result
                    if assignment_status_first_word == expected_status:
                        valid_rows.append([cell.value for cell in row])
                    else:
                        logger.info(f"[REMOVED][ELSE] Row {i}: Person={person}, Action={action}, AssignmentType={assignment_status_full} — Expected '{expected_status}' as first word")
                    matched = True
                    break
                else:
                    action_list = [val.strip().upper() for val in rule_value.split(",")]
                    if action in action_list:
                        expected_status = rule_result
                        if assignment_status_first_word == expected_status:
                            valid_rows.append([cell.value for cell in row])
                        else:
                            logger.info(f"[REMOVED] Row {i}: Person={person}, Action={action}, AssignmentType={assignment_status_full} — Expected '{expected_status}' as first word")
                        matched = True
                        break

            if not matched:
                logger.info(f"[REMOVED] Row {i}: Person={person}, Action={action} didn't match any rule and no ELSE provided.")

        logger.info(f"[INFO] Unique first words of AssignmentStatusTypeCode found: {sorted(first_words_logged)}")

        # Clear old rows & save valid ones
        ws.delete_rows(start_row, ws.max_row - header_row_num)

        for idx, row_data in enumerate(valid_rows, start=start_row):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=idx, column=col_idx, value=value)

        out_path = file_path.parent / f"{file_path.stem}.xlsx" if save_as_new_file else file_path
        wb.save(out_path)
        logger.info(f"[DONE] ✅ Assignment Type Code validated and saved as: {out_path.name}")
        return out_path

    except Exception as e:
        logger.error(f"[ERROR] ❌ Exception during Assignment_type_Code: {e}")
        return None

def validate_LegalEmployer_change(
    file_content: BytesIO,
    original_filename: str,
    hire_action_codes: Optional[str] = None,
    termination_action_codes: Optional[str] = None,
    allowed_le_change_action_codes: Optional[str] = None
):
    logger.info(f"Validating file: {original_filename}")
    inconsistent_records = []
    person_numbers_with_errors = set()

    if not original_filename.lower().endswith(('.xls', '.xlsx', '.csv')):
        raise HTTPException(status_code=400, detail="Invalid file type. Only Excel (.xls, .xlsx) or CSV files are allowed.")

    try:
        file_content.seek(0)
        if original_filename.lower().endswith(('.xls', '.xlsx')):
            df = pd.read_excel(file_content, header=1)
        else:
            df = pd.read_csv(file_content, header=1)

        df.columns = df.columns.str.strip().str.replace(' ', '').str.upper()
        required_columns = ['PERSONNUMBER', 'ACTIONCODE', 'LEGALEMPLOYERNAME', 'EFFECTIVESTARTDATE']
        for col in required_columns:
            if col not in df.columns:
                raise HTTPException(status_code=400, detail=f"Missing required column: {col}")

        original_df = df.copy()
        df.dropna(subset=required_columns, inplace=True)

        dropped_rows = original_df[~original_df.index.isin(df.index)]
        for _, row in dropped_rows.iterrows():
            pn = str(row.get('PERSONNUMBER', 'N/A')).strip()
            missing_cols = [col for col in required_columns if pd.isna(row.get(col)) or str(row.get(col)).strip() == '']
            inconsistent_records.append({
                'PersonNumber': pn,
                'Scenario': 'Missing Required Data',
                'Status': f"Missing columns: {', '.join(missing_cols)}"
            })
            if pn != 'N/A':
                person_numbers_with_errors.add(pn)

        if df.empty:
            return _make_response(inconsistent_records, person_numbers_with_errors, df)

        df['ACTIONCODE'] = df['ACTIONCODE'].astype(str).str.strip().str.upper()
        df['LEGALEMPLOYERNAME'] = df['LEGALEMPLOYERNAME'].astype(str).str.strip()
        df['PERSONNUMBER'] = df['PERSONNUMBER'].astype(str).str.strip()
        df['EFFECTIVESTARTDATE'] = pd.to_datetime(df['EFFECTIVESTARTDATE'], errors='coerce')

        seen_inconsistencies = set()

        def add_inconsistency(rec):
            rec_key = tuple(sorted(rec.items()))
            if rec_key not in seen_inconsistencies:
                inconsistent_records.append(rec)
                person_numbers_with_errors.add(rec['PersonNumber'])
                seen_inconsistencies.add(rec_key)

        # Invalid dates
        invalid_dates = df[df['EFFECTIVESTARTDATE'].isnull()]
        for _, row in invalid_dates.iterrows():
            add_inconsistency({
                'PersonNumber': row['PERSONNUMBER'],
                'EffectiveStartDate': 'Invalid Date',
                'ActionCode': row['ACTIONCODE'],
                'LegalEmployerName': row['LEGALEMPLOYERNAME'],
                'Scenario': 'Invalid Effective Start Date',
                'Status': 'Date could not be parsed'
            })
        df = df[df['EFFECTIVESTARTDATE'].notnull()]

        # Missing LE
        missing_le = df[df['LEGALEMPLOYERNAME'] == '']
        for _, row in missing_le.iterrows():
            add_inconsistency({
                'PersonNumber': row['PERSONNUMBER'],
                'EffectiveStartDate': row['EFFECTIVESTARTDATE'].strftime('%Y-%m-%d'),
                'ActionCode': row['ACTIONCODE'],
                'LegalEmployerName': '',
                'Scenario': 'Missing Legal Employer Name',
                'Status': 'Legal Employer is empty'
            })
        df = df[df['LEGALEMPLOYERNAME'] != '']

        hire_actions = [code.strip().upper() for code in hire_action_codes.split(',')] if hire_action_codes else []
        term_actions = [code.strip().upper() for code in termination_action_codes.split(',')] if termination_action_codes else []
        allowed_le_change_actions = [code.strip().upper() for code in allowed_le_change_action_codes.split(',')] if allowed_le_change_action_codes else []

        for pn, group in df.groupby('PERSONNUMBER'):
            group = group.sort_values(by='EFFECTIVESTARTDATE').reset_index(drop=True)
            current_le = None
            last_was_termination = False

            for _, row in group.iterrows():
                action = row['ACTIONCODE']
                le = row['LEGALEMPLOYERNAME']
                eff_date = row['EFFECTIVESTARTDATE']

                if action in hire_actions:
                    if current_le and not last_was_termination and current_le != le:
                        add_inconsistency({
                            'PersonNumber': pn,
                            'EffectiveStartDate': eff_date.strftime('%Y-%m-%d'),
                            'ActionCode': action,
                            'LegalEmployerName': le,
                            'PreviousLegalEmployer': current_le,
                            'Scenario': 'Hire without prior Termination with LE change',
                            'Status': f"LE changed to '{le}' without termination from '{current_le}'"
                        })
                    current_le = le
                    last_was_termination = False

                elif action in term_actions:
                    if not current_le:
                        add_inconsistency({
                            'PersonNumber': pn,
                            'EffectiveStartDate': eff_date.strftime('%Y-%m-%d'),
                            'ActionCode': action,
                            'LegalEmployerName': le,
                            'Scenario': 'Termination without prior Hire',
                            'Status': 'Termination found without Hire'
                        })
                    if current_le and current_le != le:
                        add_inconsistency({
                            'PersonNumber': pn,
                            'EffectiveStartDate': eff_date.strftime('%Y-%m-%d'),
                            'ActionCode': action,
                            'LegalEmployerName': le,
                            'PreviousLegalEmployer': current_le,
                            'Scenario': 'Termination with mismatched LE',
                            'Status': f"Termination LE '{le}' does not match '{current_le}'"
                        })
                    current_le = None
                    last_was_termination = True

                else:
                    if not current_le:
                        add_inconsistency({
                            'PersonNumber': pn,
                            'EffectiveStartDate': eff_date.strftime('%Y-%m-%d'),
                            'ActionCode': action,
                            'LegalEmployerName': le,
                            'Scenario': 'Action without Hire',
                            'Status': f"Action '{action}' outside of employment period"
                        })
                    elif current_le != le:
                        if action in allowed_le_change_actions:
                            current_le = le
                        else:
                            add_inconsistency({
                                'PersonNumber': pn,
                                'EffectiveStartDate': eff_date.strftime('%Y-%m-%d'),
                                'ActionCode': action,
                                'LegalEmployerName': le,
                                'PreviousLegalEmployer': current_le,
                                'Scenario': 'LE changed mid-employment without valid action',
                                'Status': f"LE changed to '{le}' with action '{action}'"
                            })
                            current_le = le
                    last_was_termination = False

        if person_numbers_with_errors:
            df = df[~df['PERSONNUMBER'].isin(person_numbers_with_errors)]
            logger.info(f"[INFO] Removed inconsistent PersonNumbers: {person_numbers_with_errors}")

        for pn, group in df.groupby('PERSONNUMBER'):
            hires = group[group['ACTIONCODE'].isin(hire_actions)].sort_values(by='EFFECTIVESTARTDATE')
            terms = group[group['ACTIONCODE'].isin(term_actions)].sort_values(by='EFFECTIVESTARTDATE')

            if not hires.empty and not terms.empty:
                if hires.iloc[0]['LEGALEMPLOYERNAME'] != terms.iloc[0]['LEGALEMPLOYERNAME']:
                    add_inconsistency({
                        'PersonNumber': pn,
                        'Scenario': 'First Hire vs First Termination LE mismatch',
                        'FirstHireLE': hires.iloc[0]['LEGALEMPLOYERNAME'],
                        'FirstTerminationLE': terms.iloc[0]['LEGALEMPLOYERNAME'],
                        'Status': 'Mismatch between hire and termination LE'
                    })
            elif hires.empty and not terms.empty:
                add_inconsistency({
                    'PersonNumber': pn,
                    'Scenario': 'Termination without any Hire',
                    'Status': 'Termination action found but no hire action'
                })

        return _make_response(inconsistent_records, person_numbers_with_errors, df)

    except Exception as e:
        logger.exception(f"Error while validating file {original_filename}: {e}")
        raise HTTPException(status_code=500, detail=f"Internal Error: {str(e)}")

def _make_response(inconsistent_records, person_numbers_with_errors, df):
    def safe_cast_errors(errors: List[dict]):
        for record in errors:
            for key, value in record.items():
                if isinstance(value, (np.integer, np.int64)):
                    record[key] = int(value)
                elif isinstance(value, (np.floating, np.float64)):
                    record[key] = float(value)
                elif isinstance(value, pd.Timestamp):
                    record[key] = value.strftime('%Y-%m-%d')
        return errors

    def df_to_dict(df: pd.DataFrame):
        df_copy = df.copy()
        for col in df_copy.columns:
            if pd.api.types.is_datetime64_any_dtype(df_copy[col]):
                df_copy[col] = df_copy[col].dt.strftime('%Y-%m-%d')
            elif pd.api.types.is_numeric_dtype(df_copy[col]):
                df_copy[col] = df_copy[col].astype(float)
        return df_copy.to_dict(orient='records')

    return {
        "inconsistencies": safe_cast_errors(inconsistent_records),
        "person_numbers_with_errors": list(person_numbers_with_errors),
        "cleaned_data": df_to_dict(df.reset_index(drop=True))
    }




@app.post("/api/hdl/bulk-excel-upload")
async def bulk_excel_upload(
    parent_name: str = Form(...),
    excelFile: UploadFile = File(...),
    Mandatory_Objects: str = Form(...),
    assignment_status_rules: str = Form(...),
    TermActions: str = Form(...),
    HireActions: str = Form(...),
    glbTransfers: str = Form(...),
    all_mandatory_objects: str = Form(...),
    all_non_mandatory_objects: str = Form(...),
    # Add customerName and InstanceName as form parameters
    customerName: str = Form(...),
    InstanceName: str = Form(...)
):
    if not excelFile.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Invalid file format. Please upload an .xlsx file.")

    # Construct the new parent_folder path: uploads/customerName/InstanceName/parent_name
    customer_folder = DIR / customerName
    instance_folder = customer_folder / InstanceName
    parent_folder = instance_folder / parent_name

    # Create the nested directories if they don't exist
    parent_folder.mkdir(parents=True, exist_ok=True)
    logger.info(f"Created directory structure: {parent_folder}")

    try:
        # ✅ Parse JSON strings into lists
        term_actions_list = json.loads(TermActions)
        hire_actions_list = json.loads(HireActions)
        glb_transfer_list = json.loads(glbTransfers)
        assignment_status_rules_list = json.loads(assignment_status_rules)
        mandatory_objects_list = json.loads(all_mandatory_objects)
        all_non_mandatory_objects_list = json.loads(all_non_mandatory_objects) # Added this as it's passed from frontend

        logger.info(f"✅ Parsed Actions: Term={term_actions_list}, Hire={hire_actions_list}, GT={glb_transfer_list}")

        contents = await excelFile.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        saved_files = []
        mandatory_person_numbers = set()

        # === 1️⃣ First Pass — Collect PersonNumbers from Mandatory Sheets ===
        for sheet_name in wb.sheetnames:
            if sheet_name.strip() in mandatory_objects_list:
                ws = wb[sheet_name]
                # Assuming headers are in the 3rd row (index 2 for 0-indexed list)
                headers = [cell.value for cell in ws[3]]
                if not headers:
                    continue
                try:
                    person_idx = headers.index("PersonNumber")
                except ValueError:
                    continue

                for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
                    if person_idx is not None and person_idx < len(row) and row[person_idx]:
                        mandatory_person_numbers.add(str(row[person_idx]).strip())

        # === 2️⃣ Second Pass — Filter Rows and Save Each Sheet ===
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Assuming headers are in the 3rd row (index 2 for 0-indexed list)
            headers = [cell.value for cell in ws[3]]
            if not headers:
                continue

            new_wb = openpyxl.Workbook()
            new_ws = new_wb.active
            new_ws.title = sheet_name
            # Append the original headers
            new_ws.append(headers)

            try:
                person_idx = headers.index("PersonNumber")
            except ValueError:
                person_idx = None

            is_mandatory = sheet_name.strip() in mandatory_objects_list
            rows_added = 0

            for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
                if all(cell is None or (isinstance(cell, str) and cell.strip() == "") for cell in row):
                    continue # Skip entirely empty rows

                if person_idx is not None:
                    if person_idx < len(row) and row[person_idx]:
                        person_number = str(row[person_idx]).strip()
                        # If it's a non-mandatory sheet, and the person number is NOT in the mandatory set, skip this row
                        if not is_mandatory and person_number not in mandatory_person_numbers:
                            continue
                    elif is_mandatory:
                        # If it's a mandatory sheet and PersonNumber is missing for a row, log a warning and skip
                        logger.warning(f"⚠️ Mandatory sheet '{sheet_name}' row {row_idx} missing PersonNumber. Skipping row.")
                        continue

                new_ws.append(row)
                rows_added += 1

            if rows_added > 0:
                safe_sheet_name = "_".join(sheet_name.strip().split())
                out_path = parent_folder / f"{safe_sheet_name}.xlsx"
                new_wb.save(out_path)
                saved_files.append({
                    "sheet": sheet_name,
                    # Ensure the relative path is correct from the base upload directory
                    "file": str(out_path.relative_to(DIR)),
                    "parent": parent_name,
                    "child": sheet_name
                })
                logger.info(f"[✅] Saved sheet '{sheet_name}' to '{out_path.name}'")

        # === 3️⃣ Post-Save Validations (assuming these functions are defined elsewhere) ===
        errors = []

        # Placeholder for validation functions (you need to define these or import them)
        def populate_actual_termination_date_from_resignation(file_path, hireActions, globalTransfers, termAction):
            logger.info(f"Running populate_actual_termination_date_from_resignation for {file_path}")
            # Implement your logic here
            pass

        def validate_workrelationship_sheet(file_path, hireActions, globalTransfers):
            logger.info(f"Running validate_workrelationship_sheet for {file_path}")
            # Implement your logic here
            pass

        def Assignment_type_Code(file_path, assignment_status_rules, save_as_new_file):
            logger.info(f"Running Assignment_type_Code for {file_path}")
            # Implement your logic here
            pass

        def validate_termination_date(file_path, save_as_new_file, TermActions, HireActions):
            logger.info(f"Running validate_termination_date for {file_path}")
            # Implement your logic here
            pass

        def validate_LegalEmployer_change(file_content, original_filename, hire_action_codes, termination_action_codes, allowed_le_change_action_codes):
            logger.info(f"Running validate_LegalEmployer_change for {original_filename}")
            # Implement your logic here, return a list of errors if any
            return []


        for file_meta in saved_files:
            sheet_name_lower = file_meta["sheet"].strip().lower()
            # Ensure the path is absolute for internal operations
            current_file_path = DIR / file_meta["file"]

            if sheet_name_lower == "workrelationship":
                populate_actual_termination_date_from_resignation(
                    current_file_path, hireActions=hire_actions_list,
                    globalTransfers=glb_transfer_list, termAction=term_actions_list
                )

                validate_workrelationship_sheet(
                    current_file_path, hireActions=hire_actions_list,
                    globalTransfers=glb_transfer_list
                )

                assignment_file = parent_folder / "Assignment.xlsx"
                if assignment_file.exists():
                    Assignment_type_Code(
                        assignment_file, assignment_status_rules=assignment_status_rules_list, save_as_new_file=False
                    )
                    validate_termination_date(
                        assignment_file, save_as_new_file=False, TermActions=term_actions_list, HireActions=hire_actions_list
                    )
                else:
                    logger.warning("⚠️ Assignment.xlsx not found — skipping validations.")

                workterms_file = parent_folder / "WorkTerms.xlsx"
                if workterms_file.exists():
                    with open(workterms_file, "rb") as f:
                        workterms_content = BytesIO(f.read())

                    le_validation_results = validate_LegalEmployer_change(
                        file_content=workterms_content,
                        original_filename="WorkTerms.xlsx",
                        hire_action_codes=",".join(hire_actions_list),
                        termination_action_codes=",".join(term_actions_list),
                        allowed_le_change_action_codes=",".join(glb_transfer_list)
                    )

                    if le_validation_results:
                        errors.extend(le_validation_results)
                        logger.warning(f"⚠️ Legal Employer validation found inconsistencies: {le_validation_results}")
                    else:
                        logger.info("✅ Legal Employer validation passed.")
                else:
                    logger.warning("⚠️ WorkTerms.xlsx not found — skipping LE validation.")

                break  # Only validating WorkRelationship cycle

        return {"parent": parent_name, "files": saved_files, "errors": errors}

    except HTTPException as http_exc:
        raise http_exc
    except Exception as e:
        logger.critical(f"❌ Unhandled error during bulk upload for {excelFile.filename}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Internal server error during file processing: {str(e)}")




# Mount the static directory for Dat_Files. Ensure the directory exists.
DAT_FILES_DIR = Path("Required_files/Dat_Files")
DAT_FILES_DIR.mkdir(parents=True, exist_ok=True) # Ensure this directory exists at startup
app.mount("/static", StaticFiles(directory=DAT_FILES_DIR), name="static_dat_files")

@app.post("/api/hdl/data-transformation")
async def get_data_transformation_mapping(Attributes: List[str] = Body(..., embed=True)):
    """
    Retrieves existing data transformation mappings for a list of attributes from the Excel file.
    """
    try:
        # Define the path to the mapping file (using the Excel file name)
        mapping_file_path = TRANSFORMATION_ATTRIBUTES_FILE_PATH
        
        # Load the Excel file into a DataFrame
        df = pd.read_excel(mapping_file_path, engine="openpyxl")

        # Clean column names by stripping whitespace
        df.columns = [col.strip() for col in df.columns]

        # Filter the DataFrame to include only the correct column names
        mapping_df = df[['Attributes for Transformation', 'Transformation Map Name']].copy()

        # Create a dictionary from the filtered DataFrame
        existing_mappings = mapping_df.set_index('Attributes for Transformation')['Transformation Map Name'].fillna('').to_dict()
        existing_mappings = {k: v.strip() for k, v in existing_mappings.items()}
        filtered_mappings = {attr: existing_mappings.get(attr, "") for attr in Attributes}        
        return {"mapping": filtered_mappings}

    except FileNotFoundError:
        logging.error(f"Transformation file not found at {mapping_file_path}")
        raise HTTPException(status_code=500, detail=f"Transformation file not found at {mapping_file_path}")
    except Exception as e:
        logging.error(f"Error retrieving data transformation mapping: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error retrieving data transformation mapping: {str(e)}")

class TransformationPayload(BaseModel):
    attribute: str
    newMapping: str # Changed from Json[dict] to str for simple string mapping
    componentName: str # Added componentName

class ApplyTransformationPayload(BaseModel):
    mapping_json_file: str
    raw_excel_file: str
    attribute_column: str
    
logger = logging.getLogger(__name__)

# Helper function to apply transformations to a DataFrame
def apply_data_transformation(df: pd.DataFrame, transformations: Dict[str, str]) -> pd.DataFrame:
    """
    Applies a set of transformation rules to the DataFrame.
    `transformations` is a dict where key is attribute name and value is the transformation string.
    """
    df_transformed = df.copy()
    for col in df_transformed.columns:
        if df_transformed[col].dtype == 'object':
            if pd.isna(df[col]).any(): # Check original df for NaN
                df_transformed[col] = df_transformed[col].fillna('')
                logger.debug(f"Replaced NaN with empty string in column: {col}")
            elif pd.api.types.is_numeric_dtype(df_transformed[col]):
                if pd.isna(df[col]).any() or np.isinf(df[col]).any(): # Check original df for NaN/Inf
                    df_transformed[col] = df_transformed[col].replace({np.nan: 0, np.inf: 1e308, -np.inf: -1e308})
                    logger.debug(f"Handled NaN/Inf in numeric column: {col}")

    for attr, transformation_str in transformations.items():
        if attr in df_transformed.columns:
            if attr.lower() == 'suffix':
                def transform_value(value):
                    clean_value = str(value).strip().lower()
                    if clean_value == 'senior':
                        return 'Sr'
                    elif clean_value == 'junior':
                        return 'Jr'
                    elif clean_value == '':
                        return ''
                    else:
                        return value 
                df_transformed[attr] = df_transformed[attr].apply(transform_value)
                logger.info(f"Applied specific 'Suffix' transformation to column: '{attr}'.")
            else:
                logger.info(f"Transformation rule for '{attr}' is '{transformation_str}'. Generic application not implemented for this type.")


    return df_transformed



class ExcelRequest(BaseModel):
    customerName: str
    InstanceName: str
    parent: str  # Global BO
    filename: str  # Component Name (Excel file name)

@app.post("/excel")
async def get_excel_file_post(request_data: ExcelRequest):
    """
    Returns the requested Excel file from the uploads/Excel_Files directory using POST request with payload.
    """
    logger.info(f"Received Excel POST request with {request_data.customerName}/{request_data.InstanceName}/{request_data.parent}/{request_data.filename}")
    
    file_path = DIR / request_data.customerName / request_data.InstanceName / request_data.parent / request_data.filename

    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found.")
    
    return FileResponse(
        file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=request_data.filename
    )

# --- NEW VALIDATION ENDPOINT AND HELPER FUNCTIONS ---

class AttributeValidationData(BaseModel):
    Attributes: str
    required: bool
    LookUp_data: Optional[str] = None
    CodeName: Optional[str] = None
    Data_Transformation: Optional[str] = None
    # --- NEW: Add the includeInDatFileGeneration field to the Pydantic model ---
    includeInDatFileGeneration: bool = True # Default to True as per frontend logic

def required_field_validations(
    df: pd.DataFrame, required_columns: List[str]
) -> pd.DataFrame:
    """
    Performs validation to ensure required fields are present and not empty in the DataFrame.
    Adds a 'Reason for Failed' column to the DataFrame, detailing validation failures.

    This function operates on a copy of the DataFrame to prevent modifying the original.
    It identifies two types of failures for 'required' fields:
    1.  Missing required columns in the DataFrame itself.
    2.  Empty (or whitespace-only) values within existing required columns.

    Args:
        df (pd.DataFrame): The input DataFrame to validate.
        required_columns (List[str]): A list of column names that are mandatory.

    Returns:
        pd.DataFrame: A new DataFrame with an added 'Reason for Failed' column,
                      detailing validation issues for each row. Rows without issues
                      will have an empty string in this column.
    """
    if not isinstance(df, pd.DataFrame):
        logger.error("Input 'df' must be a pandas DataFrame.")
        raise TypeError("Input 'df' must be a pandas DataFrame.")
    if not isinstance(required_columns, list) or not all(isinstance(col, str) for col in required_columns):
        logger.error("Input 'required_columns' must be a list of strings.")
        raise TypeError("Input 'required_columns' must be a list of strings.")

    if df.empty and required_columns:
        logger.warning("Input DataFrame is empty, but required columns were specified. No validation performed on rows.")
        # Create a 'Reason for Failed' column if it doesn't exist, as per function contract
        if "Reason for Failed" not in df.columns:
            df["Reason for Failed"] = ""
        return df # Return empty DataFrame, as there are no rows to validate

    df_copy = df.copy()

    # Initialize a list of lists to store reasons for each row
    # This is more robust than string concatenation inside a loop
    all_row_reasons = [[] for _ in range(len(df_copy))]

    # Check for missing required columns in the DataFrame schema
    actual_columns = set(df_copy.columns)
    missing_cols_in_excel = [col for col in required_columns if col not in actual_columns]

    if missing_cols_in_excel:
        missing_cols_str = ', '.join(missing_cols_in_excel)
        logger.warning(f"Missing required columns in Excel file: {missing_cols_str}. Marking all rows as failed for these columns.")
        # Mark all rows as failed for missing columns
        for i in range(len(df_copy)):
            all_row_reasons[i].append(f"Missing required column(s): {missing_cols_str}")
        # Add the missing columns to the DataFrame so subsequent steps don't fail,
        # but they will contain NaN or empty strings, correctly failing validation.
        for col in missing_cols_in_excel:
            df_copy[col] = '' # Or pd.NA, depends on desired behavior for subsequent steps

    # Validate required values (non-empty) for existing required columns
    for col in required_columns:
        if col in df_copy.columns: # Only check if the column actually exists in the DataFrame
            # Identify rows where the required column is empty or contains only whitespace
            # Convert to string to handle mixed types and NaN safely
            empty_mask = df_copy[col].astype(str).str.strip() == ""
            
            # Append reason for each affected row
            for i in df_copy.index[empty_mask]:
                all_row_reasons[i].append(f"'{col}' is required and cannot be empty")
            logger.info(f"Validated required field '{col}'. Found {empty_mask.sum()} empty values.")
        # else: already logged in missing_cols_in_excel block or handled above.

    # Consolidate all reasons for each row into the 'Reason for Failed' column
    df_copy["Reason for Failed"] = ["; ".join(filter(None, reasons)).strip(" ;") for reasons in all_row_reasons]
    
    # Replace empty strings (from rows with no failures) with a proper empty string, if needed.
    # The filter(None, ...) and strip(";") already handles this for the most part.
    df_copy["Reason for Failed"] = df_copy["Reason for Failed"].replace("^$", "", regex=True)

    return df_copy

def lookup_validations(df: pd.DataFrame, all_lookups: Dict[str, List[Dict[str, str]]]) -> pd.DataFrame:
    """
    Performs lookup validations on the DataFrame.
    If a column has lookup values defined but a row's data for that column is empty,
    it will be ignored and not added to the failed list for lookup validation.
    """
    df_copy = df.copy()
    if "Reason for Failed" not in df_copy.columns:
        df_copy["Reason for Failed"] = ""
    else:
        df_copy["Reason for Failed"] = df_copy["Reason for Failed"].astype(str)

    for attribute, lookup_list in all_lookups.items():
        if attribute in df_copy.columns:
            # FIX: Use getattr to safely access 'Value' from Pydantic LookupItem objects
            valid_values = set(
                str(getattr(item, 'Value', '')).strip().lower()
                for item in lookup_list
                if getattr(item, 'Value', '') is not None and str(getattr(item, 'Value', '')).strip() != ''
            )
            
            # If no valid values are defined for a lookup, skip validation for this attribute entirely
            if not valid_values:
                logger.warning(f"No valid lookup values found for attribute '{attribute}'. Skipping lookup validation for this column.")
                continue

            is_empty_in_row_data = df_copy[attribute].astype(str).str.strip() == ""
            is_valid_lookup = df_copy[attribute].astype(str).str.strip().str.lower().isin(valid_values)
            
            # An invalid lookup occurs if the value is NOT empty AND is NOT a valid lookup value
            invalid_mask = (~is_empty_in_row_data) & (~is_valid_lookup)
            df_copy.loc[invalid_mask, "Reason for Failed"] += f"; Invalid lookup value for {attribute}"
            logger.info(f"Validated lookup for '{attribute}'. Found {invalid_mask.sum()} invalid values (excluding empty row data).")
        else:
            logger.warning(f"Lookup attribute '{attribute}' not found in DataFrame columns. Skipping lookup validation for this column.")
    
    # Clean up the "Reason for Failed" column
    df_copy["Reason for Failed"] = df_copy["Reason for Failed"].str.strip(" ;").replace("^$", "", regex=True)
    return df_copy


def transformation_for_validation(excel_file_io: BytesIO):
    """
    Applies data transformations based on the transformation file (Excel).
    Reads the Excel file from the BytesIO stream and applies transformations
    according to the mapping in TRANSFORMATION_ATTRIBUTES_FILE_PATH for all attributes present in the file.
    """
    # Load the mapping from the Excel file (same as apply_transformation_and_download)
    mapping_file_path = TRANSFORMATION_ATTRIBUTES_FILE_PATH
    mapping_df = pd.read_excel(mapping_file_path, engine="openpyxl")
    mapping_df.columns = [col.strip() for col in mapping_df.columns]
    mapping_dict = mapping_df[['Attributes for Transformation', 'Transformation Map Name']].set_index('Attributes for Transformation')['Transformation Map Name'].fillna('').to_dict()
    mapping_dict = {k: v.strip() for k, v in mapping_dict.items()}

    df = pd.read_excel(excel_file_io, engine='openpyxl')
    df.columns = [col.strip() for col in df.columns]
    df = df.fillna("")
    for col in df.columns:
        if df[col].dtype == 'object':
            if pd.isna(df[col]).any():
                df[col] = df[col].fillna('')
        elif pd.api.types.is_numeric_dtype(df[col]):
            if pd.isna(df[col]).any() or pd.isinf(df[col]).any():
                df[col] = df[col].replace({np.nan: 0, np.inf: 1e308, -np.inf: -1e308})
    # Apply transformations based on the mapping for each attribute in the DataFrame
    for attr in df.columns:
        transformation_str = mapping_dict.get(attr, "")
        if transformation_str:
            def transform_value(value):
                clean_value = str(value).strip().lower()
                # Example: handle Suffix transformation, can be extended for other rules
                if transformation_str.lower().startswith('suffix'):
                    if clean_value == 'senior':
                        return 'Sr'
                    elif clean_value == 'junior':
                        return 'Jr'
                    elif clean_value == '':
                        return ''
                return value
            df[attr] = df[attr].apply(transform_value)
            logger.info(f"Applied transformation '{transformation_str}' to column: '{attr}'.")
    if "Reason for Failed" not in df.columns:
        df["Reason for Failed"] = ""
    else:
        df["Reason for Failed"] = df["Reason for Failed"].astype(str)
    return df


def get_generic_filename(prefix: str, identifier: str, extension: str) -> str:
    """Generates a unique filename."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{identifier}_{timestamp}.{extension}"

# Helper function to convert base64 to DataFrame
def base64_to_dataframe(base64_string: str) -> pd.DataFrame:
    """Decodes a base64 string to a pandas DataFrame."""
    try:
        decoded_bytes = base64.b64decode(base64_string)
        # Try reading as CSV first, then Excel
        try:
            df = pd.read_csv(io.StringIO(decoded_bytes.decode('utf-8')))
        except Exception:
            df = pd.read_excel(io.BytesIO(decoded_bytes))
        return df
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error decoding base64 or reading file: {e}")


# Pydantic model for individual attribute configuration
class AttributeConfig(BaseModel):
    Attributes: str
    required: bool
    keyValues: bool
    LookUp_data: str
    CodeName: str
    Data_Transformation: str
    data_type: str
    includeInDatFileGeneration: bool

# Pydantic model for lookup data structure
class LookupItem(BaseModel):
    CODE_Name: str
    Value: str
    Meaning: str
    Enabled_Flag: str
    Effective_Date: str

# Pydantic model for the entire validation payload
class ValidatePayload(BaseModel):
    pyFileName: Optional[str] = None
    componentName: str
    attributes: List[AttributeConfig]
    allLookups: Dict[str, List[LookupItem]]
    allMapping: Dict[str, Any] # Assuming mapping can be flexible
    excelFile: str # Base64 encoded Excel file content
    globalBoName: Optional[str] = None
    sourceKeys: Optional[Dict[str, str]] = None # Changed from List[str] to Dict[str, str]
    datColumnOrder: Optional[List[str]] = None
    hireActions: Optional[List[str]] = Field(default_factory=list) # Added with default
    rehireActions: Optional[List[str]] = Field(default_factory=list) # Added with default
    terminationActions: Optional[List[str]] = Field(default_factory=list) # Added with default
    globalTransferActions: Optional[List[str]] = Field(default_factory=list) # Added with default
    customerName: Optional[str] = None # Added customerName
    InstanceName: Optional[str] = None # Added InstanceName

# Helper function to convert base64 to DataFrame
def base64_to_dataframe(base64_string: str) -> pd.DataFrame:
    """Decodes a base64 string to a pandas DataFrame."""
    try:
        decoded_bytes = base64.b64decode(base64_string)
        # Try reading as CSV first, then Excel
        try:
            df = pd.read_csv(io.StringIO(decoded_bytes.decode('utf-8')))
        except Exception:
            df = pd.read_excel(io.BytesIO(decoded_bytes))
        return df
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error decoding base64 or reading file: {e}")


def get_hdl_setup_validate_fetch(customer_name: str, instance_name: str) -> dict:
    try:
        setup_dir = Path("User/setup_files")

        filename = f"{customer_name.replace(' ', '_')}_{instance_name.replace(' ', '_')}_setup.json"
        filepath = setup_dir / filename
        logger.info(f"Looking for setup file at: {filepath}")
        if not filepath.exists():
            logger.info(f"Looking for setup file at: {filepath}")
            raise HTTPException(status_code=404, detail=f"Setup file not found. Searched at: {filepath}")

        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)

        return data  # <-- just return dict, not JSONResponse

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching setup: {e}")


DATA_TYPE_MAPPING = {
    "VARCHAR": "string",
    "CHAR": "string",
    "TEXT": "string",
    "STRING": "string",
    "NUMBER": "float",  # or "integer" if you know it's always int
    "INTEGER": "integer",
    "INT": "integer",
    "FLOAT": "float",
    "DECIMAL": "float",
    "DATE": "date",
    "DATETIME": "date",
    "BOOLEAN": "boolean",
    "BOOL": "boolean",
}

def validate_data_types(df: pd.DataFrame, attributes: List[AttributeConfig]) -> pd.DataFrame:
    """
    Validate each column's values against its expected data_type.
    Updates 'Reason for Failed' column with errors.
    Supports mapping of SQL/HCM types to Python types.
    """
    for attr in attributes:
        col = attr.Attributes
        dtype_raw = attr.data_type.upper()
        expected_type = DATA_TYPE_MAPPING.get(dtype_raw)

        if not expected_type:
            # Unknown type, skip or log
            for idx in df.index:
                existing_reason = df.at[idx, "Reason for Failed"]
                df.at[idx, "Reason for Failed"] = f"{existing_reason}; Unknown data_type '{attr.data_type}' specified for column." if existing_reason else f"Unknown data_type '{attr.data_type}' specified for column."
            continue

        if col not in df.columns:
            continue  # Skip missing columns

        for idx, val in df[col].items():
            if val == "" or pd.isna(val):
                continue  # Empty values handled by required field validation
            reason = ""
            try:
                if expected_type == "string":
                    str(val)
                elif expected_type == "integer":
                    if not str(val).isdigit():
                        reason = f"Expected integer but got '{val}'"
                elif expected_type == "float":
                    float(val)
                elif expected_type == "boolean":
                    if str(val).strip().upper() not in ["TRUE", "FALSE", "1", "0"]:
                        reason = f"Expected boolean (TRUE/FALSE/1/0) but got '{val}'"
                elif expected_type == "date":
                    pd.to_datetime(val, errors='raise')
            except Exception:
                reason = f"Value '{val}' does not match expected data_type '{expected_type}'"

            if reason:
                existing_reason = df.at[idx, "Reason for Failed"]
                df.at[idx, "Reason for Failed"] = f"{existing_reason}; {reason}" if existing_reason else reason

    return df

def apply_workrelationship_rules(df: pd.DataFrame,
                                hire_actions: list,
                                rehire_actions: list,
                                gt_actions: list,
                                term_actions: list) -> pd.DataFrame:
    """
    HCM-style WorkRelationship rules:
    - Tracks multiple hires/rehire sequences after terminations.
    - ActualTerminationDate = next termination - 1 for the current hire/rehire/GT sequence.
    - Returns only hire/rehire/GT rows.
    """
    if df.empty:
        logger.info("Input DataFrame is empty. Returning as-is.")
        return df

    # Ensure necessary columns exist
    for col in ["PersonNumber", "ActionCode", "EffectiveStartDate"]:
        if col not in df.columns:
            logger.warning(f"Column '{col}' missing. Skipping WR rules.")
            return df

    # Normalize columns
    df["ActionCode"] = df["ActionCode"].astype(str).str.strip().str.upper()
    df["EffectiveStartDate"] = pd.to_datetime(df["EffectiveStartDate"], errors="coerce")

    # Ensure ActualTerminationDate exists
    if "ActualTerminationDate" not in df.columns:
        df["ActualTerminationDate"] = None

    final_rows = []

    for person_number, group in df.groupby("PersonNumber"):
        # Sort chronologically
        group = group.sort_values("EffectiveStartDate").copy()
        last_term_date = None
        next_term_dates = group[group["ActionCode"].isin(term_actions)]["EffectiveStartDate"].tolist()
        term_idx = 0  # pointer for next termination

        for idx, row in group.iterrows():
            action = row["ActionCode"]

            # If termination, skip it (we’ll assign its date to prior hires)
            if action in term_actions:
                continue

            # Determine the next termination date for this hire/rehire/GT
            next_term_date = None
            while term_idx < len(next_term_dates) and next_term_dates[term_idx] <= row["EffectiveStartDate"]:
                term_idx += 1  # skip past terminations before or on this row

            if term_idx < len(next_term_dates):
                next_term_date = next_term_dates[term_idx] - pd.Timedelta(days=1)

            # Assign ActualTerminationDate
            if next_term_date:
                group.at[idx, "ActualTerminationDate"] = next_term_date.strftime("%Y/%m/%d")

        # Append only hire/rehire/GT rows
        final_rows.append(group[group["ActionCode"].isin(hire_actions + rehire_actions + gt_actions)])

    df_final = pd.concat(final_rows, ignore_index=True)

    # Ensure proper string formatting for ActualTerminationDate
    df_final["ActualTerminationDate"] = df_final["ActualTerminationDate"].apply(safe_format_date)


    return df_final

def safe_format_date(val):
    if pd.isna(val) or val is None:
        return None
    try:
        dt = pd.to_datetime(val, errors="coerce")
        if pd.isna(dt):
            return None
        return dt.strftime("%Y/%m/%d")
    except Exception:
        return None


def load_validation_module(file_path: Path):
    spec = importlib.util.spec_from_file_location(file_path.stem, file_path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


@app.post("/api/hdl/validate-data")
async def validate_data(payload: ValidatePayload):
    """
    Validates the uploaded Excel file against the provided attributes, lookups, and mappings.
    Returns a JSON response with validation results.
    Calls each function to check for its status and return the results.
    Fetches excel file from the static directory using the componentName and globalBoName.
    """
    component_name = payload.componentName
    global_bo_name = payload.globalBoName
    attributes_to_validate = payload.attributes
    all_lookups = payload.allLookups
    hire_actions = [act.upper().strip() for act in (payload.hireActions or [])]
    term_actions = [act.upper().strip() for act in (payload.terminationActions or [])]
    gt_actions = [act.upper().strip() for act in (payload.globalTransferActions or [])]
    rehire_actions = [act.upper().strip() for act in (payload.rehireActions or [])]
    customerName = payload.customerName or "default_customer"
    instanceName = payload.InstanceName or ""
    # ----------------------------------------------------------
    all_mapping = payload.allMapping
    excel_base64 = payload.excelFile
    dat_column_order = payload.datColumnOrder or []

    # -----------------------------------------------------------
    # use the setup data endpoint function to get these values for hire_actions, term_actions, gt_actions
    # Ensure these are lists even if None is passed
    
    
    # Parse the JSON content to extract the lists
    # Fetch setup data if any actions are missing
    if not hire_actions or not term_actions or not gt_actions or not rehire_actions:
        logger.info("Fetching missing action lists from setup data.")
        try:
            setup_data = get_hdl_setup_validate_fetch(customer_name=customerName, instance_name=instanceName)
            hire_actions = setup_data.get("hireActions", hire_actions)
            term_actions = setup_data.get("termActions", term_actions)
            gt_actions = setup_data.get("globalTransferActions", gt_actions)
            rehire_actions = setup_data.get("rehireActions", rehire_actions)
            logger.info(f"Fetched actions - Hire: {hire_actions}, Termination: {term_actions}, GT: {gt_actions}, Rehire: {rehire_actions}")
        except HTTPException as e:
            logger.error(f"Setup fetch error: {e.detail}")
            raise
        except Exception as e:
            logger.error(f"Unexpected error fetching setup: {e}")
            raise HTTPException(status_code=500, detail=f"Error fetching setup data: {str(e)}")

    # ----------------------------------------------------------
    if not excel_base64:
        raise HTTPException(status_code=400, detail="Excel file content is missing.")

    try:
        # Decode the base64 Excel file
        excel_bytes = base64.b64decode(excel_base64)
        excel_file_io = BytesIO(excel_bytes)
        excel_filename = f"{component_name}.xlsx"
        store_excel = DIR / (global_bo_name or component_name) / excel_filename
        # Ensure the directory exists
        store_excel.parent.mkdir(parents=True, exist_ok=True)
        # Save the Excel file to the determined path
        with open(store_excel, "wb") as f:
            f.write(excel_bytes)
        file_path_in_static = Path("uploads/Excel_Files") / (global_bo_name or component_name) / excel_filename
        # Load Excel file into DataFrame
        df = pd.read_excel(excel_file_io, engine='openpyxl')
        df.columns = [str(col).strip() for col in df.columns]
        df.columns = [col.strip() for col in df.columns] # Clean column names
        df = df.fillna("") # Fill NaN values with empty string for consistent validation

        # Get the original columns from the uploaded Excel (before any additions like sourceKeys)
        original_excel_columns = df.columns.tolist()

        # Initialize 'Reason for Failed' column
        if "Reason for Failed" not in df.columns:
            df["Reason for Failed"] = ""
        else:
            df["Reason for Failed"] = df["Reason for Failed"].astype(str)

        all_errors = []

        # Remove only leading and trailing whitespaces from all string values
        df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)


        # 1. Perform Data Transformation/Mapping Validations (NOW FIRST)
        logger.info("Skipping data transformation step as requested.")
        # df = transformation_for_validation(excel_file_io) 
        # logger.info("Finished data transformation.")

        # 2. Perform Required Field Validations (after transformation)
        logger.info("Starting required field validations.")
        required_cols_from_payload = [attr.Attributes for attr in attributes_to_validate if attr.required]
        df = required_field_validations(df, required_cols_from_payload)
        logger.info("Finished required field validations.")

        # 2a. Perform Data Type Validation
        logger.info("Starting datatype validations.")
        df = validate_data_types(df, attributes_to_validate)
        logger.info("Finished datatype validations.")

        # 2b. Perform Key Values Uniqueness Validation
        logger.info("Starting key values uniqueness validation.")

        

        # Collect all attributes flagged as keyValues=True
        key_value_cols_from_payload = [attr.Attributes for attr in attributes_to_validate if attr.keyValues]

        if key_value_cols_from_payload:
            missing_cols = [col for col in key_value_cols_from_payload if col not in df.columns]
            if missing_cols:
                logger.warning(f"Key value validation skipped for missing columns: {', '.join(missing_cols)}")
            else:
                # Build composite key if multiple key columns exist
                df["_key_combo"] = df[key_value_cols_from_payload].astype(str).agg("|".join, axis=1)

                # Detect duplicates in that composite key
                dup_mask = df["_key_combo"].duplicated(keep=False)  # mark all duplicates, not just later ones
                if dup_mask.any():
                    for col in key_value_cols_from_payload:
                        df.loc[dup_mask, "Reason for Failed"] = df.loc[dup_mask, "Reason for Failed"].astype(str) + \
                            f"; Duplicate detected in key combination ({', '.join(key_value_cols_from_payload)})"

                # Drop helper column after validation
                df = df.drop(columns=["_key_combo"])

        logger.info("Finished key values uniqueness validation.")


        # --- VALIDATE START DATE BEFORE END DATE ---
        logger.info("Starting EffectiveStartDate < EffectiveEndDate validation...")

        if "EffectiveStartDate" in df.columns and "EffectiveEndDate" in df.columns:
            # Ensure proper datetime format
            df["EffectiveStartDate"] = pd.to_datetime(df["EffectiveStartDate"], errors="coerce")
            df["EffectiveEndDate"] = pd.to_datetime(df["EffectiveEndDate"], errors="coerce")

            # Find rows where start date is missing or not before end date
            invalid_mask = df["EffectiveEndDate"] <= df["EffectiveStartDate"]
            if invalid_mask.any():
                for idx in df[invalid_mask].index:
                    start = df.at[idx, "EffectiveStartDate"]
                    end = df.at[idx, "EffectiveEndDate"]
                    existing_reason = df.at[idx, "Reason for Failed"]
                    reason = f"EffectiveStartDate ({start.date()}) must be before EffectiveEndDate ({end.date()})"
                    df.at[idx, "Reason for Failed"] = f"{existing_reason}; {reason}" if existing_reason else reason

        logger.info("Completed EffectiveStartDate < EffectiveEndDate validation.")


        # 3. Perform Lookup Validations (after transformation and required field checks)
        logger.info("Starting lookup validations.")
        # `all_lookups` from payload is already in the correct format (attribute -> list of dicts)
        df = lookup_validations(df, all_lookups)
        logger.info("Finished lookup validations.")        

        
        # 4. Custom validation: First row for a Person Number must be 'HIRE' based on minimum Start Date
        logger.info("Starting custom validation: First row for Person Number must be 'HIRE'.")
        
        start_date_column = None
        if "EffectiveStartDate" in df.columns:
            start_date_column = "EffectiveStartDate"
        elif "DateStart" in df.columns:
            start_date_column = "DateStart"

        if "PersonNumber" in df.columns and "ActionCode" in df.columns and start_date_column:
            # Convert start_date_column to datetime, handling potential errors
            df[start_date_column] = pd.to_datetime(df[start_date_column], errors='coerce')

            # Drop rows where start_date_column could not be parsed, as they cannot be validated on date
            df_cleaned_dates = df.dropna(subset=[start_date_column]).copy()

            # Group by Person Number and find the row with the minimum DateStart
            idx = df_cleaned_dates.groupby('PersonNumber')[start_date_column].idxmin()
            first_rows_for_person = df_cleaned_dates.loc[idx]

            for _, row in first_rows_for_person.iterrows():
                person_number = row["PersonNumber"]
                action_code = str(row["ActionCode"]).strip().upper()

                if action_code not in hire_actions:
                    # Find all original indices for this person_number to mark all their rows as failed
                    original_indices_for_person = df[df["PersonNumber"] == person_number].index
                    for i in original_indices_for_person:
                        # Append the reason, ensuring it's not None
                        current_reason = df.loc[i, "Reason for Failed"]
                        if current_reason:
                            df.loc[i, "Reason for Failed"] = f"{current_reason}; First action for PersonNumber '{person_number}' (based on minimum start date) must be 'HIRE', but was '{action_code}'"
                        else:
                            df.loc[i, "Reason for Failed"] = f"First action for PersonNumber '{person_number}' (based on minimum start date) must be 'HIRE', but was '{action_code}'"
        else: 
            missing_cols = []
            if "PersonNumber" not in df.columns: missing_cols.append("PersonNumber")
            if "ActionCode" not in df.columns: missing_cols.append("ActionCode")
            if not start_date_column: missing_cols.append("EffectiveStartDate or DateStart") # Updated warning message
            if missing_cols:
                logger.warning(f"Skipping 'HIRE' validation due to missing column(s): {', '.join(missing_cols)}")
        logger.info("Finished custom validation: First row for Person Number must be 'HIRE'.")


        # Legal Employer Name has to be consistant 
        logger.info("Starting validation: LegalEmployerName consistency check with reset on HIRE/TERMINATION/GT.")
        required_cols = ["PersonNumber", "ActionCode", start_date_column, "LegalEmployerName"]
        if all(col in df.columns for col in required_cols):
            df[start_date_column] = pd.to_datetime(df[start_date_column], errors='coerce')
            df["ActionCode"] = df["ActionCode"].astype(str).str.strip().str.upper()
            df["LegalEmployerName"] = df["LegalEmployerName"].astype(str).str.strip()

            grouped = df.sort_values(by=["PersonNumber", start_date_column]).groupby("PersonNumber")

            for person, group in grouped:
                expected_legal = None
                for _, row in group.iterrows():
                    action = row["ActionCode"]
                    legal = row["LegalEmployerName"]
                    idx = row.name

                    if expected_legal is None:
                        expected_legal = legal
                    elif action in rehire_actions or action in gt_actions:
                        expected_legal = legal
                    else:
                        if legal != expected_legal:
                            existing_reason = df.at[idx, "Reason for Failed"]
                            reason = f"Inconsistent LegalEmployerName. Expected '{expected_legal}', but found '{legal}' and action code '{action}'."
                            df.at[idx, "Reason for Failed"] = f"{existing_reason}; {reason}" if existing_reason else reason

        else:
            logger.warning("Missing columns for LegalEmployerName consistency validation. Skipping.")

        logger.info("Completed validation: LegalEmployerName consistency check.")



        # --- GLOBAL TRANSFER validation: legal employer change + no employment gap ---
        logger.info("Starting GLOBAL TRANSFER validation (LegalEmployerName change + date continuity)...")

        required_cols = ["PersonNumber", "ActionCode", "EffectiveStartDate", "EffectiveEndDate", "LegalEmployerName"]
        if all(col in df.columns for col in required_cols):
            df["EffectiveStartDate"] = pd.to_datetime(df["EffectiveStartDate"], errors="coerce")
            df["EffectiveEndDate"] = pd.to_datetime(df["EffectiveEndDate"], errors="coerce")
            df["ActionCode"] = df["ActionCode"].astype(str).str.strip()
            df["LegalEmployerName"] = df["LegalEmployerName"].astype(str).str.strip()

            grouped = df.sort_values(by=["PersonNumber", "EffectiveStartDate"]).groupby("PersonNumber")

            for person_number, group in grouped:
                group = group.reset_index()

                for i in range(1, len(group)):
                    prev = group.loc[i - 1]
                    curr = group.loc[i]

                    if curr["ActionCode"].upper() in gt_actions:
                        legal_changed = curr["LegalEmployerName"] != prev["LegalEmployerName"]
                        expected_start = prev["EffectiveEndDate"] + pd.Timedelta(days=1)
                        actual_start = curr["EffectiveStartDate"]


                        reasons = []
                        if not legal_changed:
                            reasons.append("Legal Employer Name must change for change legal employer")
                        if pd.notna(expected_start) and pd.notna(actual_start) and actual_start != expected_start:
                            reasons.append(
                                f"EffectiveStartDate ({actual_start.date()}) must be the day after previous EffectiveEndDate ({prev['EffectiveEndDate'].date()})"
                            )

                        if reasons:
                            row_index = curr["index"]
                            existing_reason = df.at[row_index, "Reason for Failed"]
                            combined = "; ".join(reasons)
                            df.at[row_index, "Reason for Failed"] = f"{existing_reason}; {combined}" if existing_reason else combined
        else:
            missing = [c for c in required_cols if c not in df.columns]
            logger.warning(f"Skipping GLOBAL TRANSFER validation due to missing columns: {', '.join(missing)}")
        logger.info("Completed GLOBAL TRANSFER validation.")

        # --- WORKRELATIONSHIP specific rules ---
        if component_name.lower() == "workrelationship":
            logger.info("Applying WorkRelationship-specific rules.")
            df = apply_workrelationship_rules(
                df=df.copy(),
                hire_actions=hire_actions,
                rehire_actions=rehire_actions,
                gt_actions=gt_actions,
                term_actions=term_actions
            )
            logger.info("Completed WorkRelationship-specific rules.")


        
        # Filter passed and failed rows
        failed_df = df[df["Reason for Failed"] != ""].copy()
        passed_df = df[df["Reason for Failed"] == ""].copy()

        # Save results to temporary files
        output_dir = VALIDATION_RESULTS_DIR
        
        # Ensure the output directory exists
        output_dir.mkdir(parents=True, exist_ok=True)

        validation_results = {}

        # Step 1 — Get all PersonNumbers with at least one failed record
        if "PersonNumber" not in df.columns:
            logger.error("PersonNumber column missing when trying to cascade fail. Available columns: %s", df.columns.tolist())
        else:
            failed_persons = df.loc[df["Reason for Failed"] != "", "PersonNumber"].unique()

            # Step 2 — For all rows of those PersonNumbers, mark as failed (if not already)
            for person_number in failed_persons:
                person_rows = df[df["PersonNumber"] == person_number].index
                for idx in person_rows:
                    if df.at[idx, "Reason for Failed"] == "":
                        df.at[idx, "Reason for Failed"] = "Failed due to other row(s) for this PersonNumber failing validation."


        # Save failed_df to an Excel file if it's not empty
        if not failed_df.empty:
            failed_file_name = get_generic_filename("failed", component_name, "xlsx")
            failed_file_path = output_dir / failed_file_name
            try:
                failed_df.to_excel(failed_file_path, index=False, engine='openpyxl')
                logger.info(f"Failed records saved to {failed_file_path}")
                validation_results["failed_file_url"] = f"/excel_validation_results/{failed_file_name}" # URL for frontend
                validation_results["failed_count"] = len(failed_df)
            except Exception as e:
                logger.error(f"Error saving failed_df to Excel: {e}")
                validation_results["failed_file_error"] = f"Failed to save error file: {str(e)}"
        else:
            logger.info("No failed records found. Skipping failed file generation.")
            validation_results["failed_count"] = 0

        # --- inside validate_data ---
        saved_code_file = UPLOAD_DIR / "saved_code" / f"{customerName}_{instanceName}_{component_name}.py"

        if saved_code_file.exists():
            try:
                module = load_validation_module(saved_code_file)
                logger.info(f"Loaded custom validation module: {saved_code_file}")

                if hasattr(module, "validate_row"):
                    logger.info(f"Found 'validate_row' function in {saved_code_file}. Applying custom row-level validation.")
                    # Initialize column to track row-level validation failures
                    if "RowValidationFailed" not in df.columns:
                        df["RowValidationFailed"] = False

                    for idx, row in df.iterrows():
                        try:
                            logger.debug(f"Applying custom validation to row {idx}")
                            is_valid, reason = module.validate_row(row.to_dict(), idx)
                            # Normalize outputs
                            is_valid = bool(is_valid)
                            reason = str(reason).strip() if reason else ""
                            
                            if not is_valid:
                                logger.debug(f"Row {idx} failed custom validation: {reason}")
                                # Append to existing "Reason for Failed"
                                existing_reason = df.at[idx, "Reason for Failed"] or ""
                                combined_reason = f"{existing_reason}; {reason}" if existing_reason else reason
                                df.at[idx, "Reason for Failed"] = combined_reason
                                df.at[idx, "RowValidationFailed"] = True
                        except Exception as row_err:
                            # Catch row-level exceptions, log but continue
                            logger.error(f"Custom validation failed for row {idx}: {row_err}", exc_info=True)
                            existing_reason = df.at[idx, "Reason for Failed"] or ""
                            df.at[idx, "Reason for Failed"] = f"{existing_reason}; Custom validation error: {row_err}" if existing_reason else f"Custom validation error: {row_err}"
                            df.at[idx, "RowValidationFailed"] = True

                else:
                    logger.warning(f"No 'validate_row' function found in {saved_code_file}")

            except Exception as mod_err:
                logger.error(f"Failed to load custom validation module: {mod_err}", exc_info=True)
        else:
            logger.info(f"No custom validation file found at: {saved_code_file}")


        # Filter passed and failed rows
        failed_df = df[df["Reason for Failed"] != ""].copy()
        passed_df = df[df["Reason for Failed"] == ""].copy()

        # Save results to temporary files
        output_dir = VALIDATION_RESULTS_DIR

        # --- Filter columns for DAT file generation based on includeInDatFileGeneration flag ---
        # Remove 'Reason for Failed' column before any further processing if it's not meant for the DAT output.
        if 'Reason for Failed' in passed_df.columns:
            passed_df = passed_df.drop(columns=['Reason for Failed'])
            
        # --- Prepare iterator values for passed_df ---
        actioncode_col = None
        personnumber_col = None
        for col in passed_df.columns:
            if col.strip().lower() == "actioncode":
                actioncode_col = col
            if col.strip().lower() == "personnumber":
                personnumber_col = col
        iterator_list = []
        iterator_counter = 0
        prev_person_number = None
        combined_actions = hire_actions + gt_actions + rehire_actions
        logger.error(f"Combined Actions is : {combined_actions}")
        for _, row in passed_df.iterrows():
            current_person_number = str(row[personnumber_col]) if personnumber_col else None
            if personnumber_col and current_person_number != prev_person_number:
                iterator_counter = 0
                prev_person_number = current_person_number
            if actioncode_col and str(row[actioncode_col]).strip().upper() in combined_actions:
                iterator_counter += 1
                iterator_list.append(str(iterator_counter))
            else:
                iterator_list.append(iterator_counter)

        # Insert sourceKeys columns at the front, using iterator_list for {Iterator} (including as part of a string)
        source_keys = getattr(payload, 'sourceKeys', None)
        # Store names of source keys and their original order as they appear in the payload.sourceKeys dict
        source_key_names_in_order = [] 
        if source_keys and isinstance(source_keys, dict):
            # Create a temporary DataFrame for just the source keys, maintaining their intended order
            source_keys_data = {}
            for key, val in source_keys.items():
                source_key_names_in_order.append(key) # Keep track of the order of source keys
                col_values = []
                for idx, (_, row) in enumerate(passed_df.iterrows()): # Iterate over the passed_df rows to substitute values
                    def replacer(match):
                        if match.group(1) == 'Iterator':
                            return str(iterator_list[idx]) if idx < len(iterator_list) else ''
                        col_name = match.group(1)
                        # Ensure we are getting values from the original passed_df (before sourceKey insertion or initial drop of 'Reason for Failed')
                        # For this, it's safer to get from the 'df' DataFrame (initial loaded Excel data) if possible,
                        # or ensure 'passed_df' still contains all original columns before dropping 'Reason for Failed' and adding source keys.
                        # As it is, `passed_df` right before this block *does* contain original columns (minus 'Reason for Failed' if it was there).
                        return str(row.get(col_name, ''))
                    
                    substituted_value = re.sub(r'\{([^}]+)\}', replacer, str(val))
                    col_values.append(substituted_value)
                source_keys_data[key] = col_values
            
            # Create a DataFrame for source keys and concatenate it to the passed_df
            # This ensures source keys are at the beginning in their specified order
            source_keys_df = pd.DataFrame(source_keys_data, index=passed_df.index)
            # Use concat to place source_keys_df at the start
            passed_df = pd.concat([source_keys_df, passed_df], axis=1)
            logger.info(f"Added source keys to passed_df. New columns added: {source_key_names_in_order}")

        # Now, construct the final columns for DAT output based on datColumnOrder or original Excel order.
        explicitly_excluded_from_dat = {
            attr_data.Attributes for attr_data in attributes_to_validate if not attr_data.includeInDatFileGeneration
        }

        final_dat_columns_ordered = []

        # 1. Add source keys first
        for sk_name in source_key_names_in_order:
            if sk_name in passed_df.columns: # Ensure it actually exists in the combined DF
                final_dat_columns_ordered.append(sk_name)

        if dat_column_order:
            logger.info(f"Using provided datColumnOrder: {dat_column_order}")
            # Then add columns from datColumnOrder, respecting exclusions and ensuring they exist in passed_df
            for col_name in dat_column_order:
                # IMPORTANT: Only add if NOT already added as a source key AND NOT explicitly excluded
                if col_name not in source_key_names_in_order and col_name not in explicitly_excluded_from_dat:
                    if col_name in passed_df.columns:
                        final_dat_columns_ordered.append(col_name)
        else:
            logger.info("datColumnOrder not provided. Falling back to original Excel column order with source keys prepended.")
            # Then add original Excel columns
            for col_name in original_excel_columns:
                # IMPORTANT: Only add if NOT already added as a source key AND NOT explicitly excluded
                if col_name not in source_key_names_in_order and col_name not in explicitly_excluded_from_dat:
                    if col_name in passed_df.columns:
                        final_dat_columns_ordered.append(col_name)

        # Remove any duplicates that might arise from edge cases, while preserving order as much as possible
        seen = set()
        deduplicated_final_dat_columns_ordered = []
        for col in final_dat_columns_ordered:
            if col not in seen:
                deduplicated_final_dat_columns_ordered.append(col)
                seen.add(col)

        # Apply the final ordering and filtering to the DataFrame
        # Ensure that all columns in deduplicated_final_dat_columns_ordered are actually in passed_df
        existing_cols_in_final_order = [col for col in deduplicated_final_dat_columns_ordered if col in passed_df.columns]

        passed_df_final_output = passed_df[existing_cols_in_final_order]
        logger.info(f"Final columns for DAT file generation (ordered): {passed_df_final_output.columns.tolist()}")


        # --- GENERATE .DAT FILE FOR PASSED DATA ---
        passed_file_name = None
        failed_file_name = None
        special_counter = 0  # Ensure special_counter is always defined
        if not passed_df_final_output.empty: # Use the filtered DF for output
            
            passed_file_name = get_generic_filename(f"{component_name}_passed", "data", "dat")
            passed_file_path = output_dir / passed_file_name
            
            # Find ActionCode column again after potential column filtering
            actioncode_col = None
            for col in passed_df_final_output.columns:
                if col.strip().lower() == "actioncode":
                    actioncode_col = col
                    break
            
            # Calculate special_counter before writing .dat
            if actioncode_col:
                logger.info(f"Found ActionCode column: {actioncode_col}. Calculating special counter.")
                special_counter = sum(
                    str(row[actioncode_col]).strip().lower() in combined_actions
                    for _, row in passed_df_final_output.iterrows()
                )
                logger.info(f"Special counter calculated: {special_counter} for ActionCode values.")
            
            # Write .dat file (pipe-separated, include header, all filtered columns)
            with open(passed_file_path, "w", encoding="utf-8") as f:
                f.write("|".join([str(col) for col in passed_df_final_output.columns]) + "\n")
                for index, row in passed_df_final_output.iterrows(): # Use index for dtype lookup
                    formatted_values = []
                    for col_name, value in row.items():
                        # Check if the column's dtype is a datetime type
                        if pd.api.types.is_datetime64_any_dtype(passed_df_final_output[col_name]):
                            if pd.isna(value): # Handle NaT (Not a Time)
                                formatted_values.append('')
                            else:
                                formatted_values.append(value.strftime('%Y/%m/%d')) # Format as %Y/%m/%d
                        else:
                            formatted_values.append(str(value) if value is not None else '')
                    f.write("|".join(formatted_values) + "\n")
            logger.info(f"Passed validation data saved to: {passed_file_path}. Special count: {special_counter}")

        if not failed_df.empty:
            failed_file_name = get_generic_filename(f"{component_name}_failed", "data", "xlsx")
            failed_file_path = output_dir / failed_file_name
            tmp_path_failed = failed_file_path.with_suffix(".xlsx.tmp")
            failed_df.to_excel(tmp_path_failed, index=False, engine='openpyxl')
            os.replace(tmp_path_failed, failed_file_path)
            logger.info(f"Failed validation data saved to: {failed_file_path}")

        # Return results
        response_content = {
            "message": "Validation complete.",
            "status": "success",
            "passed_records_count": len(passed_df_final_output),
            "failed_records_count": len(failed_df),
            "passed_file_url": f"http://localhost:8000/validation_results/{passed_file_name}" if passed_file_name else None,
            "failed_file_url": f"http://localhost:8000/validation_results/{failed_file_name}" if failed_file_name else None,
            "errors": all_errors       
         }

        if len(failed_df) > 0:
            response_content["status"] = "failed"
            response_content["message"] = "Validation completed with errors."
        
        return JSONResponse(content=response_content)

    except FileNotFoundError as e:
        logger.error(f"File not found during validation: {e}")
        raise HTTPException(status_code=404, detail=f"Required file not found: {e}")
    except Exception as e:
        logger.exception("An error occurred during data validation.")
        raise HTTPException(status_code=500, detail=f"An unexpected error occurred during validation: {str(e)}")


VALIDATION_RESULTS_DIR = "validation_results"
os.makedirs(VALIDATION_RESULTS_DIR, exist_ok=True)
# Mount the directory for serving validation results. It's now ensured to exist at startup.
app.mount("/validation_results", StaticFiles(directory=VALIDATION_RESULTS_DIR), name="validation_results")

@app.post("/api/validate-personname")
async def validate_personname(
    user_id: str = Form(...),
    excel_file_id: str = Form(...),
    dat_file_id: str = Form(...),
    required_columns: str = Form(...),
    lookup_json: str = Form(None),
    mapping_json: str = Form(None)
):
    """
    Validate the uploaded Excel file for required columns, lookups, and mapping values in one go.
    Returns file paths for failed and passed validations.
    """
    try:
        # COMPLETED_FOLDER is now defined globally and created at startup
        # UPLOAD_DIR is now defined globally and created at startup
        excel_path = UPLOAD_DIR / excel_file_id
        dat_path = UPLOAD_DIR / dat_file_id

        if not excel_path.exists() or not dat_path.exists():
            logging.error(f"Excel or DAT file not found: {excel_path}, {dat_path}")
            raise HTTPException(status_code=404, detail="Excel or DAT file not found.")

        # Load Excel
        df = pd.read_excel(excel_path)
        df.columns = df.columns.str.strip().str.replace(" ", "")
        df = df.fillna("")
        required_cols = [col.strip().replace(" ", "") for col in required_columns.split(",") if col.strip()]
        errors = []

        # Ensure Reason for Failed column exists and is string type
        if "Reason for Failed" not in df.columns:
            df["Reason for Failed"] = ""
        else:
            df["Reason for Failed"] = df["Reason for Failed"].astype(str)

        # 1. Validate required columns
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            errors.append(f"Missing required columns: {', '.join(missing_cols)}")
            for col in missing_cols:
                df["Reason for Failed"] += f"; Missing required column: {col}"

        # 2. Validate required values (non-empty)
        for col in required_cols:
            if col in df.columns:
                empty_mask = df[col].astype(str).str.strip() == ""
                df.loc[empty_mask, "Reason for Failed"] += f"; {col} is required"

        # 3. Validate lookups (if provided)
        if lookup_json:
            try:
                lookup_data = json.loads(lookup_json)
                for lookup in lookup_data:
                    col = lookup.get("column")
                    valid_values = set(str(v).strip().lower() for v in lookup.get("valid_values", []))
                    if col and col in df.columns:
                        invalid_mask = ~df[col].astype(str).str.strip().str.lower().isin(valid_values)
                        df.loc[invalid_mask, "Reason for Failed"] += f"; Invalid lookup for {col}"
            except Exception as e:
                logging.error(f"Lookup validation error: {str(e)}")
                errors.append(f"Lookup validation error: {str(e)}")

        # 4. Validate mappings (if provided)
        if mapping_json:
            try:
                mapping_data = json.loads(mapping_json)
                for mapping in mapping_data:
                    col = mapping.get("column")
                    mapping_dict = {str(m["source"]).strip().lower(): m["target"] for m in mapping.get("mappings", [])}
                    if col and col in df.columns:
                        mapped = df[col].astype(str).str.strip().str.lower().map(mapping_dict)
                        invalid_mask = mapped.isnull()
                        df.loc[invalid_mask, "Reason for Failed"] += f"; Mapping failed for {col}"
                        # Optionally, update column with mapped values where valid
                        df.loc[~invalid_mask, col] = mapped[~invalid_mask]
            except Exception as e:
                logging.error(f"Mapping validation error: {str(e)}")
                errors.append(f"Mapping validation error: {str(e)}")

        
        # Clean up Reason for Failed column (remove leading/trailing semicolons and whitespace)
        df["Reason for Failed"] = df["Reason for Failed"].str.strip(" ;")

        # Split passed/failed
        failed_df = df[df["Reason for Failed"] != ""]
        passed_df = df[df["Reason for Failed"] == ""]

        # Save failed and passed files
        failed_file_path = None
        passed_file_path = None
        if not failed_df.empty:
            failed_file_path = COMPLETED_FOLDER / get_generic_filename(user_id, "validation_errors", "xlsx")
            # Use atomic write
            tmp_path = failed_file_path.with_suffix(".xlsx.tmp")
            failed_df.to_excel(tmp_path, index=False)
            os.replace(tmp_path, failed_file_path)
            logging.info(f"Failed rows saved at {failed_file_path}.")
        if not passed_df.empty:
            passed_file_path = COMPLETED_FOLDER / get_generic_filename(user_id, "validation_passed", "dat")
            # Write as .dat (pipe-separated)
            with open(passed_file_path, "w", encoding="utf-8") as f:
                f.write("|".join(passed_df.columns) + "\n")
                for _, row in passed_df.iterrows():
                    f.write("|".join([str(x) for x in row]) + "\n")
            logging.info(f"Passed rows saved at {passed_file_path}.")

        return JSONResponse(content={
            "status": "success" if failed_df.empty and not errors else "error",
            "failed_file_path": str(failed_file_path.relative_to(UPLOAD_DIR.parent)) if failed_file_path else None,
            "passed_file_path": str(passed_file_path.relative_to(UPLOAD_DIR.parent)) if passed_file_path else None,
            "error_messages": errors
        })
    except Exception as e:
        logging.error(f"Error in unified validation: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error in unified validation: {str(e)}")
    


# --- Define the base directory for uploads, matching the bulk upload script ---
UPLOAD_DIR = Path("uploads/Excel_Files")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True) # Ensure the base directory exists

# --- Define the request body model using Pydantic for automatic validation ---
class ValidationPayload(BaseModel):
    parent_name: str
    component_files: Dict[str, str]

# Helper function to clean dictionary for JSON serialization
def clean_dict_for_json(data):
    """
    Recursively cleans a dictionary or list for JSON serialization by converting
    NaN and NaT values to None.
    """
    if isinstance(data, dict):
        return {k: clean_dict_for_json(v) for k, v in data.items()}
    elif isinstance(data, list):
        return [clean_dict_for_json(elem) for elem in data]
    elif pd.isna(data): # Checks for both NaN and NaT from pandas
        return None
    elif isinstance(data, float) and math.isnan(data): # Checks for pure float NaNs
        return None
    elif isinstance(data, datetime): # Optional: Convert datetime objects to ISO format string
        return data.isoformat()
    else:
        return data


@app.post("/api/hdl/bulk/cross-file/personNumber/validate")
async def validate_person_numbers(
    parent_name: str = Body(...),
    component_files: Dict[str, str] = Body(...),
    all_mandatory_objects: List[str] = Body(..., embed=True),
    all_non_mandatory_objects: List[str] = Body(..., embed=True),
    export_as_excel: bool = Query(False)
):
    logging.info(f"Starting validation for parent: {parent_name}")

    all_person_numbers: Set[str] = set()
    common_person_numbers: Optional[Set[str]] = None
    all_rows_list: List[pd.DataFrame] = []
    person_numbers_per_component: Dict[str, Set[str]] = {}
    structured_failed_person_numbers: List[Dict] = []

    # Load mandatory components
    for component_name, file_name in component_files.items():
        file_path = UPLOAD_DIR / parent_name / file_name
        if not file_path.exists():
            logging.warning(f"File not found: {file_path}")
            continue

        try:
            df = pd.read_excel(file_path, header=1)
            if "PersonNumber" not in df.columns:
                logging.warning(f"Missing 'PersonNumber' in {file_path.name}, skipping.")
                continue

            df["PersonNumber"] = df["PersonNumber"].astype(str).str.strip().str.replace(r"\\.0$", "", regex=True)
            person_nums = set(df["PersonNumber"].dropna().unique())
            all_person_numbers.update(person_nums)

            if component_name in all_mandatory_objects:
                common_person_numbers = person_nums if common_person_numbers is None else common_person_numbers & person_nums

            person_numbers_per_component[component_name] = person_nums
            df["Component"] = component_name
            all_rows_list.append(df)

        except Exception as e:
            logging.error(f"Error reading file {file_name}: {e}", exc_info=True)
            continue

    if common_person_numbers is None:
        common_person_numbers = set()

    failed_person_numbers_set = all_person_numbers - common_person_numbers

    for p_num in failed_person_numbers_set:
        missing = [comp for comp in all_mandatory_objects if p_num not in person_numbers_per_component.get(comp, set())]
        structured_failed_person_numbers.append({
            "person_number": p_num,
            "missing_components": missing,
            "description": f"'{p_num}' missing in: {', '.join(missing)}"
        })

    # Load non-mandatory components directly from disk
    mandatory_person_numbers = set()
    for comp in all_mandatory_objects:
        mandatory_person_numbers.update(person_numbers_per_component.get(comp, set()))

    for non_mand_comp in all_non_mandatory_objects:
        file_name = f"{non_mand_comp}.xlsx"
        file_path = UPLOAD_DIR / parent_name / file_name

        if not file_path.exists():
            logging.info(f"Skipping non-mandatory file not found: {file_path}")
            continue

        try:
            df = pd.read_excel(file_path, header=1)
            if "PersonNumber" not in df.columns:
                logging.warning(f"'PersonNumber' column missing in {file_name}")
                continue

            df["PersonNumber"] = df["PersonNumber"].astype(str).str.strip().str.replace(r"\\.0$", "", regex=True)
            non_mand_pnums = set(df["PersonNumber"].dropna().unique())
            person_numbers_per_component[non_mand_comp] = non_mand_pnums

            for p_num in non_mand_pnums:
                if p_num not in mandatory_person_numbers:
                    structured_failed_person_numbers.append({
                        "person_number": p_num,
                        "missing_components": ["[Not Found in Mandatory]"],
                        "description": f"'{p_num}' is in non-mandatory '{non_mand_comp}' but not in any mandatory component"
                    })
                    failed_person_numbers_set.add(p_num)

            df["Component"] = non_mand_comp
            all_rows_list.append(df)

        except Exception as e:
            logging.error(f"Error reading non-mandatory file {file_name}: {e}", exc_info=True)
            continue

    structured_failed_person_numbers.sort(key=lambda x: x["person_number"])

    validation_failed_df = pd.DataFrame()
    if all_rows_list:
        full_df = pd.concat(all_rows_list, ignore_index=True)
        validation_failed_df = full_df[full_df["PersonNumber"].isin(failed_person_numbers_set)].copy()

        cols_to_check = [col for col in validation_failed_df.columns if col not in ["PersonNumber", "Component"]]
        mask = pd.Series(True, index=validation_failed_df.index)
        for col in cols_to_check:
            mask &= validation_failed_df["PersonNumber"] != validation_failed_df[col].astype(str)
        validation_failed_df = validation_failed_df[mask].reset_index(drop=True)

    exported_excel_filename = None
    export_status = None
    if export_as_excel and not validation_failed_df.empty:
        filename = f"validation_failed_{uuid.uuid4()}.xlsx"
        filepath = VALIDATION_RESULTS_DIR / filename
        validation_failed_df.to_excel(filepath, index=False)
        exported_excel_filename = filename
        export_status = "Excel file generated successfully."
    elif export_as_excel:
        export_status = "No failed records to export."

    response = {
        "validation_summary": {
            "total_unique_person_numbers": len(all_person_numbers),
            "passed_validation_count": len(common_person_numbers),
            "failed_validation_count": len(failed_person_numbers_set),
            "excel_export_status": export_status
        },
        "passed_person_numbers": sorted(list(common_person_numbers)),
        "failed_person_numbers": structured_failed_person_numbers,
        "validation_failed_details": validation_failed_df.to_dict(orient="records"),
        "exported_excel_filename": exported_excel_filename
    }

    return JSONResponse(content=clean_dict_for_json(response), status_code=200)


@app.post("/api/hdl/bulk/cross-file/personNumber/remove-failed-values")
async def remove_failed_person_numbers(
    parent_name: str = Body(..., description="The parent (L4) node name for the uploaded Excel."),
    component_files: Dict[str, str] = Body(..., description="Dictionary of component_name: file_path_in_uploads."),
    person_numbers_to_remove: List[str] = Body(..., description="List of person numbers to remove from the component files.")

):
    """
    Removes rows containing specified person numbers from the given component Excel files.
    This operation directly modifies the Excel files on disk.
    """
    logging.info(f"Received request to remove failed person numbers for parent: {parent_name}")
    logging.info(f"Component files to process: {component_files}")
    logging.info(f"Person numbers to remove: {person_numbers_to_remove}")

    removal_summary: Dict[str, Dict] = {}

    if not component_files:
        logging.warning("No component files provided for removal operation.")
        return JSONResponse(
            content={"message": "No component files provided for removal.", "removal_summary": {}},
            status_code=200
        )

    for component_name, file_name_only in component_files.items():
        file_path = UPLOAD_DIR / parent_name / file_name_only
        logging.info(f"Attempting to process file for removal: {file_path} (Component: {component_name})")

        if not file_path.exists():
            logging.warning(f"File not found for component {component_name}: {file_path}. Skipping removal for this file.")
            removal_summary[component_name] = {"status": "skipped", "reason": "File not found", "rows_removed": 0}
            continue

        try:
            df = pd.read_excel(file_path, header=1)
            logging.info(f"Successfully read file for removal: {file_path}")

            if 'PersonNumber' not in df.columns:
                logging.warning(f"No 'PersonNumber' column found in {file_path}. Skipping removal for this file.")
                removal_summary[component_name] = {"status": "skipped", "reason": "No 'PersonNumber' column", "rows_removed": 0}
                continue

            # Convert PersonNumber column to string for consistent comparison
            df['PersonNumber'] = df['PersonNumber'].astype(str)

            # Get the initial number of rows
            initial_rows = len(df)

            # Filter out rows where 'PersonNumber' is in the 'person_numbers_to_remove' list
            # The ~ operator negates the boolean Series, keeping rows NOT in the set
            df_cleaned = df[~df['PersonNumber'].isin(person_numbers_to_remove)].reset_index(drop=True)

            # Calculate the number of rows removed
            rows_removed = initial_rows - len(df_cleaned)

            if rows_removed > 0:
                # Save the modified DataFrame back to the original file
                df_cleaned.to_excel(file_path, index=False)
                logging.info(f"Removed {rows_removed} rows from {file_path}. File updated successfully.")
                removal_summary[component_name] = {"status": "success", "rows_removed": rows_removed, "reason": None}
            else:
                logging.info(f"No rows removed from {file_path} as no matching person numbers were found.")
                removal_summary[component_name] = {"status": "no_change", "rows_removed": 0, "reason": "No matching person numbers found"}

        except Exception as e:
            logging.error(f"Error processing file {file_path} for removal: {e}", exc_info=True)
            removal_summary[component_name] = {"status": "failed", "reason": str(e), "rows_removed": 0}

    cleaned_summary = clean_dict_for_json(removal_summary)
    return JSONResponse(
        content={
            "message": "Removal operation completed.",
            "removal_summary": cleaned_summary
        },
        status_code=200
    )



@app.get("/api/hdl/download-excel/{filename}")
async def download_excel_file(filename: str):
    """
    Allows downloading a previously generated Excel file by its filename.
    """
    file_path = VALIDATION_RESULTS_DIR / filename
    if not file_path.is_file():
        logging.error(f"File not found for download: {file_path}")
        raise HTTPException(status_code=404, detail="File not found.")

    try:
        return FileResponse(
            path=file_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename,
        )
    except Exception as e:
        logging.error(f"Error serving file {filename}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to serve file: {e}")



@app.post("/api/hdl/bulk/cross-file/legalEmployer/validate")
async def validate_workterms_file(
    file: UploadFile = File(...),
    hire_action_codes: Optional[str] = Form(None),
    termination_action_codes: Optional[str] = Form(None),
    allowed_le_change_action_codes: Optional[str] = Form(None)
):
    """
    Uploads an Excel file (e.g., WorkTerms.xlsx) and validates consistency
    of 'Legal Employer Name' across all records for a 'Person Number'.
    
    Rules:
    1. A 'Legal Employer Name' change for a 'Person Number' is only valid if it's
       preceded by a defined termination action and followed by a defined hire/rehire action,
       OR if the change itself is triggered by a specified 'allowed_le_change_action_code'
       (e.g., GLOBAL_TRANSFER).
    2. If an active employment period exists, the Legal Employer Name must remain consistent,
       unless an explicitly allowed LE change action occurs.
    3. Termination records should ideally have a preceding hire.
    """
    logger.info(f"Received file: {file.filename}")
    logger.info(f"Received Hire Action Codes from Form: '{hire_action_codes}'")
    logger.info(f"Received Termination Action Codes from Form: '{termination_action_codes}'")
    logger.info(f"Received Allowed LE Change Action Codes from Form: '{allowed_le_change_action_codes}'")


    if not file.filename.lower().endswith(('.xls', '.xlsx', '.csv')):
        logger.warning(f"Invalid file type uploaded: {file.filename}")
        raise HTTPException(status_code=400, detail="Invalid file type. Only Excel (.xls, .xlsx) or CSV files are allowed.")

    try:
        file_content = await file.read()
        
        if file.filename.lower().endswith(('.xls', '.xlsx')):
            df = pd.read_excel(BytesIO(file_content))
            logger.info("File read as Excel.")
        elif file.filename.lower().endswith('.csv'):
            df = pd.read_csv(BytesIO(file_content))
            logger.info("File read as CSV.")
        else:
            logger.error(f"Unsupported file type after initial check: {file.filename}")
            raise HTTPException(status_code=400, detail="Unsupported file type.")

        # Normalize column names to upper case and strip spaces for robust handling
        df.columns = df.columns.str.strip().str.replace(' ', '').str.upper()

        required_columns = ['PERSONNUMBER', 'ACTIONCODE', 'LEGALEMPLOYERNAME', 'EFFECTIVESTARTDATE']
        for col in required_columns:
            if col not in df.columns:
                logger.error(f"Required column '{col}' is missing in the file: {file.filename}. Available columns: {df.columns.tolist()}")
                raise HTTPException(status_code=400, detail=f"Required column '{col}' is missing in the file.")
        
        # Drop rows where critical columns are missing
        df.dropna(subset=required_columns, inplace=True)
        if df.empty:
            logger.warning(f"File {file.filename} is empty or contains no valid data after filtering for required columns.")
            raise HTTPException(status_code=400, detail="File contains no valid data after filtering for required columns or all rows have missing critical data.")

        logger.info(f"Columns in file (after normalization): {df.columns.tolist()}")

        # Ensure relevant columns are correctly typed and cleaned
        df['ACTIONCODE'] = df['ACTIONCODE'].astype(str).str.strip().str.upper()
        df['LEGALEMPLOYERNAME'] = df['LEGALEMPLOYERNAME'].astype(str).str.strip()
        df['PERSONNUMBER'] = df['PERSONNUMBER'].astype(str).str.strip()
        df['EFFECTIVESTARTDATE'] = pd.to_datetime(df['EFFECTIVESTARTDATE'])

        hire_actions_list = []
        if hire_action_codes:
            hire_actions_list = [code.strip().upper() for code in hire_action_codes.split(',') if code.strip()]
        logger.info(f"Effective Hire Action Codes for validation: {hire_actions_list}")
        
        termination_actions_list = [] 
        if termination_action_codes:
            termination_actions_list = [code.strip().upper() for code in termination_action_codes.split(',') if code.strip()]
        logger.info(f"Effective Termination Action Codes for validation: {termination_actions_list}")
        allowed_le_change_actions_list = []
        if allowed_le_change_action_codes:
            allowed_le_change_actions_list = [code.strip().upper() for code in allowed_le_change_action_codes.split(',') if code.strip()]
        logger.info(f"Effective Allowed LE Change Action Codes: {allowed_le_change_actions_list}")

        inconsistent_records = []

        for person_number, group in df.groupby('PERSONNUMBER'):
            logger.debug(f"Processing PersonNumber: {person_number}")
            
            # Sort by EffectiveStartDate to ensure chronological processing
            group = group.sort_values(by='EFFECTIVESTARTDATE').reset_index(drop=True)

            current_le = None
            last_action_was_termination = False # Flag to indicate if the immediate previous action was a termination

            for i, row in group.iterrows():
                action_code = row['ACTIONCODE']
                legal_employer_name = row['LEGALEMPLOYERNAME']
                effective_start_date = row['EFFECTIVESTARTDATE']
                
                # Special handling for null/empty LegalEmployerName
                if pd.isna(legal_employer_name) or legal_employer_name == '':
                    inconsistent_records.append({
                        'PersonNumber': person_number,
                        'EffectiveStartDate': effective_start_date.strftime('%Y-%m-%d'),
                        'ActionCode': action_code,
                        'LegalEmployerName': legal_employer_name,
                        'Scenario': 'Missing Legal Employer Name',
                        'Status': 'Legal Employer Name is missing or null for this record.'
                    })
                    logger.warning(f"Person {person_number} (Effectivity {effective_start_date}): Missing Legal Employer Name.")
                    continue # Skip further checks for this row as LE is invalid

                logger.debug(f"Person {person_number} (Effective {effective_start_date}) - Action: {action_code}, LE: '{legal_employer_name}', Current LE state: '{current_le}', Last Action was Term: {last_action_was_termination}")


                if action_code in hire_actions_list:
                    # Case 1: Hire/Rehire action
                    if current_le is not None:
                        if not last_action_was_termination:
                            pass 
                        if current_le != legal_employer_name:
                            if not last_action_was_termination:
                                inconsistent_records.append({
                                    'PersonNumber': person_number,
                                    'EffectiveStartDate': effective_start_date.strftime('%Y-%m-%d'),
                                    'ActionCode': action_code,
                                    'LegalEmployerName': legal_employer_name,
                                    'PreviousLegalEmployer': current_le,
                                    'Scenario': 'Legal Employer Change with HIRE/REHIRE without prior Termination',
                                    'Status': f"Legal Employer changed to '{legal_employer_name}' with '{action_code}' but previous employment period was not terminated. Previous LE was '{current_le}'."
                                })
                                logger.info(f"Inconsistency for {person_number} (Effectivity {effective_start_date}): LE changed with '{action_code}' but no prior termination.")
                            else:
                                logger.debug(f"Person {person_number} (Effectivity {effective_start_date}): Valid REHIRE with new LE '{legal_employer_name}' after termination.")
                    
                    current_le = legal_employer_name 
                    last_action_was_termination = False 

                elif action_code in termination_actions_list:
                    if current_le is None:
                        inconsistent_records.append({
                            'PersonNumber': person_number,
                            'EffectiveStartDate': effective_start_date.strftime('%Y-%m-%d'),
                            'ActionCode': action_code,
                            'LegalEmployerName': legal_employer_name,
                            'Scenario': 'Termination without preceding active employment',
                            'Status': 'Termination record found without an active prior hire record for this person.'
                        })
                        logger.info(f"Inconsistency for {person_number} (Effectivity {effective_start_date}): Termination without prior active employment.")
                    if current_le is not None and current_le != legal_employer_name:
                        inconsistent_records.append({
                            'PersonNumber': person_number,
                            'EffectiveStartDate': effective_start_date.strftime('%Y-%m-%d'),
                            'ActionCode': action_code,
                            'LegalEmployerName': legal_employer_name,
                            'PreviousLegalEmployer': current_le,
                            'Scenario': 'Legal Employer Name Mismatch at Termination',
                            'Status': f"Legal Employer '{legal_employer_name}' at termination does not match last active LE '{current_le}'."
                        })
                        logger.info(f"Inconsistency for {person_number} (Effectivity {effective_start_date}): LE mismatch at termination.")

                    current_le = None 
                    last_action_was_termination = True 
                    logger.debug(f"Person {person_number} (Effectivity {effective_start_date}): TERMINATION '{action_code}'. Active period ended.")

                else: 
                    if current_le is None:
                        # Action without an active employment period (e.g., after a termination but before a rehire)
                        # This could be legitimate if it's a "post-employment" record, or an error.
                        # It depends on how your data is structured. For now, flag if it's not a hire/term.
                        inconsistent_records.append({
                            'PersonNumber': person_number,
                            'EffectiveStartDate': effective_start_date.strftime('%Y-%m-%d'),
                            'ActionCode': action_code,
                            'LegalEmployerName': legal_employer_name,
                            'Scenario': 'Action without active employment period',
                            'Status': f"Record found with non-hire/termination action '{action_code}' and LE '{legal_employer_name}' outside an active employment period."
                        })
                        logger.info(f"Inconsistency for {person_number} (Effectivity {effective_start_date}): Action '{action_code}' found without active employment.")
                    else:
                        # An active employment period exists. Check if LE changed.
                        if current_le != legal_employer_name:
                            # LE changed, but it's not a hire/rehire or termination.
                            # Is this an allowed LE change action (e.g., GLOBAL_TRANSFER)?
                            if action_code in allowed_le_change_actions_list:
                                logger.info(f"Person {person_number} (Effectivity {effective_start_date}): Valid LE change from '{current_le}' to '{legal_employer_name}' via allowed action '{action_code}'.")
                                current_le = legal_employer_name # Update current LE as this is an allowed change
                            else:
                                # This is an inconsistency: LE changed mid-employment without a proper trigger.
                                inconsistent_records.append({
                                    'PersonNumber': person_number,
                                    'EffectiveStartDate': effective_start_date.strftime('%Y-%m-%d'),
                                    'ActionCode': action_code,
                                    'LegalEmployerName': legal_employer_name,
                                    'PreviousLegalEmployer': current_le,
                                    'Scenario': 'Legal Employer Name changed mid-employment without proper action',
                                    'Status': f"Legal Employer Name changed from '{current_le}' to '{legal_employer_name}' with action '{action_code}'. This action is not a hire, termination, or an explicitly allowed LE change action."
                                })
                                logger.info(f"Inconsistency found for {person_number} (Effectivity {effective_start_date}): LE changed from '{current_le}' to '{legal_employer_name}' with invalid action '{action_code}'.")
                                current_le = legal_employer_name # Still update for subsequent checks

                    last_action_was_termination = False # Reset flag

        # --- Post-processing checks for hire/termination alignment (your original scenario 1 & 2) ---
        # These checks might duplicate some of what the chronological loop found, but can catch
        # broader discrepancies in the dataset where a hire-termination pair exists.
        
        # We need a unique list of inconsistencies, converting dicts to hashable tuples
        unique_inconsistent_records_final = []
        seen_inconsistencies = set()
        
        # Add the inconsistencies found by the chronological scan
        for rec in inconsistent_records:
            rec_tuple = tuple(sorted(rec.items()))
            if rec_tuple not in seen_inconsistencies:
                unique_inconsistent_records_final.append(rec)
                seen_inconsistencies.add(rec_tuple)

        # Now, perform the original first-hire vs first-termination check
        for person_number, group in df.groupby('PERSONNUMBER'):
            hire_rows_sorted = group[group['ACTIONCODE'].isin(hire_actions_list)].sort_values(by='EFFECTIVESTARTDATE')
            termination_rows_sorted = group[group['ACTIONCODE'].isin(termination_actions_list)].sort_values(by='EFFECTIVESTARTDATE')

            if not hire_rows_sorted.empty and not termination_rows_sorted.empty:
                first_hire_le = hire_rows_sorted['LEGALEMPLOYERNAME'].iloc[0]
                first_termination_le = termination_rows_sorted['LEGALEMPLOYERNAME'].iloc[0]
                first_hire_date = hire_rows_sorted['EFFECTIVESTARTDATE'].iloc[0]
                first_termination_date = termination_rows_sorted['EFFECTIVESTARTDATE'].iloc[0]

                # Only check if first hire occurs BEFORE first termination
                if first_hire_date < first_termination_date:
                    if first_hire_le != first_termination_le:
                        rec_data = {
                            'PersonNumber': person_number,
                            'Scenario': 'First Hire vs First Termination LE Mismatch',
                            'FirstHireLegalEmployer': first_hire_le,
                            'FirstTerminationLegalEmployer': first_termination_le,
                            'Status': 'Legal Employer of initial hire does not match Legal Employer of first termination.'
                        }
                        rec_tuple = tuple(sorted(rec_data.items()))
                        if rec_tuple not in seen_inconsistencies:
                            unique_inconsistent_records_final.append(rec_data)
                            seen_inconsistencies.add(rec_tuple)
                            logger.info(f"Inconsistency found for {person_number} (First Hire vs First Termination LE): '{first_hire_le}' != '{first_termination_le}'")
            elif not termination_rows_sorted.empty and hire_rows_sorted.empty:
                rec_data = {
                    'PersonNumber': person_number,
                    'Scenario': 'Termination without any Hire',
                    'Status': f"One or more {', '.join(termination_actions_list)} record(s) found but no {', '.join(hire_actions_list)} record(s) exist for this person."
                }
                rec_tuple = tuple(sorted(rec_data.items()))
                if rec_tuple not in seen_inconsistencies:
                    unique_inconsistent_records_final.append(rec_data)
                    seen_inconsistencies.add(rec_tuple)
                    logger.info(f"Inconsistency for {person_number} (Termination without Hire).")
        
        # Final response
        if unique_inconsistent_records_final:
            logger.info(f"Validation complete for {file.filename}. Inconsistencies found for {len(unique_inconsistent_records_final)} unique records.")
            return JSONResponse(
                content={
                    "message": "Validation complete. Inconsistencies found.",
                    "inconsistent_records": unique_inconsistent_records_final
                },
                status_code=200
            )
        else:
            logger.info(f"Validation complete for {file.filename}. All 'Legal Employer Name' records are consistent as per rules.")
            return JSONResponse(
                content={
                    "message": "Validation complete. All 'Legal Employer Name' records are consistent as per rules."
                },
                status_code=200
            )

    except KeyError as ke:
        logger.error(f"KeyError: Required column missing - {str(ke)} in file {file.filename}", exc_info=True)
        raise HTTPException(status_code=400, detail=f"Required column missing in the file: {str(ke)}. Please ensure 'PersonNumber', 'ActionCode', 'LegalEmployerName', and 'EffectiveStartDate' columns exist (case-insensitive).")
    except pd.errors.EmptyDataError:
        logger.error(f"Pandas EmptyDataError: The file {file.filename} is empty or malformed.")
        raise HTTPException(status_code=400, detail="The uploaded file is empty or contains no data.")
    except pd.errors.ParserError:
        logger.error(f"Pandas ParserError: Could not parse the file {file.filename}. It might be malformed or not in the expected format.", exc_info=True)
        raise HTTPException(status_code=400, detail="Could not parse the file. Please ensure it is a valid Excel or CSV file.")
    except Exception as e:
        logger.critical(f"An unhandled error occurred during file processing or validation for {file.filename}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"An internal server error occurred during file processing or validation: {str(e)}")


BASE_DIR = Path(__file__).resolve().parent
VALIDATION_RESULTS_DIR = BASE_DIR / "validation_results"
BUNDLE_DEPOT_ZONE = BASE_DIR / "BUNDLE_DEPOT_ZONE"

# Create the funky folder if it doesn't exist
BUNDLE_DEPOT_ZONE.mkdir(parents=True, exist_ok=True)


class FileBundleRequest(BaseModel):
    """
    Pydantic model for the request body, expecting a list of filenames to bundle.
    """
    files: List[str]

@app.post("/api/hdl/download-bundle")
async def bundle_dat_files_in_order(payload: FileBundleRequest):
    """
    Bundles .dat files by dynamically parsing the 'W{integer}' value from the
    'SourceSystemId' column of *each data record*.
    All records identified as 'W1' or 'ungrouped' will be combined into a single 'W1' bundle.
    Other unique W{integer} (e.g., W2, W3) will result in separate bundled .dat files.
    Each bundle will include the original METADATA header from each component file
    before its corresponding row data, with a line break between components.
    The bundle filename will include the W-group, a list of original component names (without _passed_data_timestamp), and a timestamp.
    Returns a list of URLs for the generated bundle files.
    """
    logger.info(f"Received bundle request for files: {payload.files}")
    try:
        if not payload.files:
            logger.warning("No filenames provided in the payload.")
            raise HTTPException(status_code=400, detail="No filenames provided.")

        # Initialize bundle_results here
        bundle_results = []
        
        # Recursively find all .dat files in VALIDATION_RESULTS_DIR
        all_dat_paths = {f.name: f for f in VALIDATION_RESULTS_DIR.rglob("*.dat")}
        logger.info(f"Found {len(all_dat_paths)} .dat files in VALIDATION_RESULTS_DIR.")

        # Dictionary to hold individual data records grouped by 'W{integer}'
        # Key: "W1", "W2", etc. Value: List of tuples (data_line_string, original_filename_string)
        grouped_records: Dict[str, List[Tuple[str, str]]] = defaultdict(list)
        
        # Store headers for each file, keyed by filename
        all_file_headers: Dict[str, str] = {}

        # To store the set of original filenames that contribute to each group
        group_source_components: Dict[str, Set[str]] = defaultdict(set)

        # Regex to capture the "W{integer}" pattern from the content.
        content_pattern = re.compile(r"(W\d+)", re.IGNORECASE)
        logger.debug(f"Using content regex pattern: {content_pattern.pattern}")

        # Regex to remove '_passed_data_YYYYMMDD_HHMMSS' from filenames
        filename_cleanup_pattern = re.compile(r'_passed_data_\d{8}_\d{6}')

        for fname in payload.files:
            logger.debug(f"Processing file: {fname}")
            dat_path = all_dat_paths.get(fname)
            if not dat_path:
                logger.warning(f"File '{fname}' not found in validation results directory. Skipping.")
                continue

            try:
                with open(dat_path, "r", encoding="utf-8") as f:
                    lines = f.readlines()
                    
                    if not lines:
                        logger.warning(f"File '{fname}' is empty. Skipping.")
                        continue

                    # Log raw lines to debug what's actually being read
                    logger.debug(f"File '{fname}': Raw lines[0]: '{lines[0].strip()}'")
                    if len(lines) > 1:
                        logger.debug(f"File '{fname}': Raw lines[1]: '{lines[1].strip()}'")

                    # The actual header line is the one starting with "METADATA|"
                    current_file_metadata_line = lines[0].strip()
                    
                    if not current_file_metadata_line.startswith("METADATA|"):
                        logger.error(f"File '{fname}': First line does not start with 'METADATA|'. Expected header format not found. Treating all lines as ungrouped data.")
                        # If the first line isn't the expected METADATA header,
                        # we can't reliably find SourceSystemId. Group all its data as ungrouped.
                        # These will now go into the 'W1' bundle as per the new logic.
                        for data_line in lines: # All lines are potentially data if header is missing
                            grouped_records["W1"].append((data_line.strip(), fname)) # Group into W1
                            group_source_components["W1"].add(fname)
                        continue # Skip to next file

                    # Store this file's header for later use in bundling
                    all_file_headers[fname] = current_file_metadata_line
                    logger.debug(f"Stored header for '{fname}': '{current_file_metadata_line}'")
                    
                    # Determine SourceSystemId column index for *this specific file's header*
                    # Use the METADATA line itself to get the headers
                    current_headers = [h.strip() for h in current_file_metadata_line.split('|')]
                    current_source_system_id_col_index = -1
                    for i, header in enumerate(current_headers):
                        if "SourceSystemId" in header: # Checks for "SourceSystemId" or "PersonId(SourceSystemId)"
                            current_source_system_id_col_index = i
                            break
                    
                    if current_source_system_id_col_index == -1:
                        logger.warning(f"File '{fname}': 'SourceSystemId' column not found in its METADATA/HEADER line. All data records from this file will be grouped as 'W1'.")
                        # If SourceSystemId is not found in this file's header, all its data lines go to 'W1'
                        for line_num, data_line in enumerate(lines[1:], start=2): # Iterate from the first data line
                            grouped_records["W1"].append((data_line.strip(), fname))
                            group_source_components["W1"].add(fname)
                        continue # Move to next file

                    # Process data lines (starting from the second line, as first is METADATA/HEADER)
                    for line_num, data_line in enumerate(lines[1:], start=2):
                        data_line_stripped = data_line.strip()
                        if not data_line_stripped: # Skip empty data lines
                            continue

                        data_fields = data_line_stripped.split('|')
                        
                        if len(data_fields) > current_source_system_id_col_index:
                            source_system_id_value = data_fields[current_source_system_id_col_index]
                            logger.debug(f"File '{fname}', Line {line_num}: Extracted 'SourceSystemId' value: '{source_system_id_value}'")
                            match = content_pattern.search(source_system_id_value)
                            if match:
                                group_key = match.group(1) # Extract the captured "W{integer}"
                                # If the group key is W1, or if it was originally 'ungrouped', map it to 'W1'
                                if group_key.upper() == "W1":
                                    grouped_records["W1"].append((data_line_stripped, fname))
                                    group_source_components["W1"].add(fname)
                                    logger.debug(f"File '{fname}', Line {line_num}: Found '{group_key}'. Appended to group 'W1'.")
                                else:
                                    grouped_records[group_key].append((data_line_stripped, fname))
                                    group_source_components[group_key].add(fname)
                                    logger.debug(f"File '{fname}', Line {line_num}: Found '{group_key}'. Appended to group '{group_key}'.")
                            else:
                                # If no W{integer} pattern found, append to 'W1' group
                                grouped_records["W1"].append((data_line_stripped, fname))
                                group_source_components["W1"].add(fname)
                                logger.debug(f"File '{fname}', Line {line_num}: No 'W{{integer}}' pattern found in 'SourceSystemId' ('{source_system_id_value}'). Appended to 'W1'.")
                        else:
                            # If SourceSystemId column value is not accessible, append to 'W1' group
                            grouped_records["W1"].append((data_line_stripped, fname))
                            group_source_components["W1"].add(fname)
                            logger.warning(f"File '{fname}', Line {line_num}: Data line has fewer columns ({len(data_fields)}) than expected for 'SourceSystemId' index ({current_source_system_id_col_index}). Appended to 'W1'.")

            except Exception as file_process_error:
                logger.error(f"Error processing file '{fname}': {file_process_error}", exc_info=True)
                # If an error occurs during processing a file, its records might be lost or misgrouped.
                # For robustness, we could try to append remaining lines to 'W1' group.
                # For now, just log and skip this file's further processing.


        # Final check if any records were grouped at all
        if not grouped_records:
            logger.warning("No data records were successfully grouped from any of the provided files.")
            raise HTTPException(
                status_code=404,
                detail="No data records found or processed from the provided files for bundling."
            )
        
        # Ensure that all necessary headers were captured from the input files.
        # If any file was processed and had data, its header should be in all_file_headers.
        if not all_file_headers:
            logger.critical("No valid METADATA/Header lines could be extracted from any input file. Cannot form valid bundles.")
            raise HTTPException(
                status_code=500,
                detail="Failed to extract essential METADATA/Header information from input files. Cannot proceed with bundling."
            )

        # Log the final state of grouped_records and group_source_components before bundle generation
        logger.info(f"Final grouped records before bundle generation (showing first 50 chars of each record and source file):")
        for k, v in grouped_records.items():
            logger.info(f"   Group '{k}' has {len(v)} records. Example: {[f'{rec[:50]}... (from {src})' for rec, src in v[:min(3, len(v))]]}")
        logger.info(f"Final group source components: {json.dumps({k: list(v) for k, v in group_source_components.items()}, indent=2)}")


        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        logger.info(f"Starting bundle generation for {len(grouped_records)} groups.")

        # Generate bundles for each group
        for group_key, records_with_source_list in grouped_records.items():
            if not records_with_source_list:
                logger.debug(f"Skipping empty group: '{group_key}'.")
                continue # Skip empty groups (e.g., if 'ungrouped' is empty)

            # Build the content for the bundle, adding specific headers and line breaks
            bundle_content_lines = []
            previous_source_fname = None
            
            # Sort records by original filename to ensure components are grouped together within the bundle
            records_with_source_list.sort(key=lambda x: x[1])

            for record_line, source_fname in records_with_source_list:
                # If the source file changes, add a line break and the new component's header
                if previous_source_fname is None or source_fname != previous_source_fname:
                    if previous_source_fname is not None: # Don't add a blank line before the very first header in the bundle
                        bundle_content_lines.append("") # Add a blank line for separation
                        logger.debug(f"Added line break in '{group_key}' bundle: source changed from '{previous_source_fname}' to '{source_fname}'")
                    
                    # Retrieve the header for the current component (source_fname)
                    header_for_current_component = all_file_headers.get(source_fname)
                    if header_for_current_component:
                        bundle_content_lines.append(header_for_current_component)
                        logger.debug(f"Added header for component '{source_fname}' to '{group_key}' bundle.")
                    else:
                        logger.warning(f"Header for source file '{source_fname}' not found in all_file_headers. Skipping header for this block in '{group_key}' bundle.")

                bundle_content_lines.append(record_line)
                previous_source_fname = source_fname
            
            # Construct the full content for the bundle
            combined_content = "\n".join(bundle_content_lines)
            
            # Prepend the required command to the combined content
            final_bundle_content = "SET PURGE_FUTURE_CHANGES N\n\n" + combined_content

            # Construct filename for the bundle based on the group key and component list
            # Get sorted unique component names for this group
            component_names_for_bundle = sorted(list(group_source_components[group_key]))
            # Remove the '_passed_data_timestamp' pattern and '.dat' extension for cleaner filename
            clean_component_names = []
            for name in component_names_for_bundle:
                cleaned_name = filename_cleanup_pattern.sub('', name) # Remove _passed_data_timestamp
                cleaned_name = cleaned_name.replace('.dat', '') # Remove .dat extension
                clean_component_names.append(cleaned_name)
            
            components_string = "_".join(clean_component_names)
            
            # Ensure filename doesn't get too long (optional, but good practice)
            max_filename_len = 200 # Max length for the component string part
            if len(components_string) > max_filename_len:
                components_string = components_string[:max_filename_len] + "_TRUNCATED"
                logger.warning(f"Component string for filename truncated for group '{group_key}'.")

            bundle_filename = f"{group_key}_{components_string}_{timestamp}.dat"
            
            # Define the full path where the bundled file will be saved
            save_path = BUNDLE_DEPOT_ZONE / bundle_filename

            try:
                with open(save_path, "w", encoding="utf-8") as f:
                    f.write(final_bundle_content) # Write the content with the prepended command
                
                logger.info(f"Successfully generated bundle for group '{group_key}': {save_path}")
                
                # Add the bundle information to the results list
                bundle_results.append({
                    "group": group_key,
                    "filename": bundle_filename,
                    "url": f"/bundle_depot_zone/{bundle_filename}"
                })
            except Exception as bundle_write_error:
                logger.error(f"Error writing bundle file for group '{group_key}' at '{save_path}': {bundle_write_error}", exc_info=True)
                bundle_results.append({
                    "group": group_key,
                    "filename": bundle_filename,
                    "url": None,
                    "error": f"Failed to save bundle: {bundle_write_error}"
                })


        # Sort bundle results for consistent output (e.g., W1, W2, then other Ws)
        def sort_key(item):
            group = item['group']
            # 'W1' should come first as it now includes 'ungrouped'
            if group == 'W1':
                return (0, 0) # Highest priority for W1
            match = re.match(r'W(\d+)', group)
            if match:
                try:
                    return (1, int(match.group(1))) # Sorts other W numbers numerically
                except ValueError:
                    return (2, group) # Fallback for non-numeric W groups
            return (2, group) # Fallback for other unexpected group names (shouldn't happen with this logic)

        bundle_results.sort(key=sort_key)
        logger.info(f"Bundling process completed. Generated {len(bundle_results)} bundles.")
        return JSONResponse(content={"bundles": bundle_results}, status_code=200)

    except HTTPException as e:
        logger.error(f"HTTPException during bundling: {e.detail}", exc_info=True)
        raise e
    except Exception as e:
        logger.critical(f"A critical unexpected error occurred during .dat file bundling: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Internal server error during bundling: {str(e)}")


# Mount the BUNDLE_DEPOT_ZONE for serving generated bundles
app.mount("/bundle_depot_zone", StaticFiles(directory=BUNDLE_DEPOT_ZONE), name="bundle_depot_zone")

@app.post("/api/hdl/zip-dat")
async def zip_dat_file_by_name(
    fileName: str = Form(...),
    componentName: str = Form(...),
    group: str= Form(...)
):
    """
    ✅ Searches recursively for `fileName` inside bundle folders,
    renames it as `<componentName>.dat`,
    zips it as `<componentName>.zip`,
    and returns the zip file to the frontend.
    """
    try:
        if not fileName.endswith(".dat"):
            raise HTTPException(status_code=400, detail="File must be a .dat file")

        logger.info(f"📂 Looking for {fileName} under {BUNDLE_DEPOT_ZONE}...")
        matches = list(BUNDLE_DEPOT_ZONE.rglob(fileName))

        if not matches:
            logger.warning(f"❌ {fileName} not found anywhere inside bundle folders.")
            raise HTTPException(status_code=404, detail="Original .dat file not found.")
        
        new_fileName = fileName.replace(".dat", f".zip")
        original_path = matches[0]
        new_dat_path = BUNDLE_DEPOT_ZONE / f"{componentName}.dat"
        new_zip_path = BUNDLE_DEPOT_ZONE / f"{new_fileName}"

        logger.info(f"🔁 Copying {original_path.name} → {new_dat_path.name}")
        shutil.copy(original_path, new_dat_path)

        logger.info(f"🗜️ Zipping {new_dat_path.name} → {new_zip_path.name}")
        with zipfile.ZipFile(new_zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
            zipf.write(new_dat_path, arcname=new_dat_path.name)

        # Optional: delete renamed .dat after zip to avoid clutter
        # os.remove(new_dat_path)

        logger.info(f"✅ Successfully zipped {componentName}.zip — sending response.")

        return FileResponse(
            new_zip_path,
            media_type="application/zip",
            filename=new_zip_path.name
        )

    except Exception as e:
        logger.error(f"🔥 Zipping failed: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Zip failed: {str(e)}")

@app.post("/api/hdl/zip-to-base64-by-name")
async def zip_to_base64_from_name(
    fileName: str = Form(...)
):
    zip_path = BUNDLE_DEPOT_ZONE / fileName

    if not zip_path.exists():
        raise HTTPException(status_code=404, detail="ZIP file not found.")

    try:
        with open(zip_path, "rb") as f:
            file_bytes = f.read()

        encoded = base64.b64encode(file_bytes).decode("utf-8")

        return {
            "fileName": fileName,
            "content": encoded
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Encoding failed: {e}")
    

    
class HDLUploadRequest(BaseModel):
    file_name: str


class HDLTriggerRequest(BaseModel):
    customerName: str
    instanceName: str
    contentId: str

    # Optional Oracle fields
    deleteSourceFileFlag: Optional[str] = None
    dataSetName: Optional[str] = None
    loadConcurrentThreads: Optional[int] = None
    importMaxErrorPercentage: Optional[int] = None
    verificationKey: Optional[str] = None
    loadMaxErrorPercentage: Optional[int] = None
    fileEncryption: Optional[str] = None
    fileAction: Optional[str] = None
    importConcurrentThreads: Optional[int] = None

    class Config:
        extra = "ignore" 


def load_oracle_credentials(customer_name: str, instance_name: str):
    load_dotenv(dotenv_path="./.env")  # make sure the path is right

    base_key = f"{customer_name}_{instance_name}".replace(" ", "").upper()
    
    oracle_env = os.getenv(f"{base_key}_ORACLE_URL")
    username = os.getenv(f"{base_key}_ORACLE_USERNAME")
    password = os.getenv(f"{base_key}_ORACLE_PASSWORD")

    if not all([oracle_env, username, password]):
        raise ValueError(f"Missing Oracle credentials for {base_key}")

    return oracle_env, username, password

class OracleUploadRequest(BaseModel):
    content: str
    fileName: str
    contentId: None
    fileEncryption: str
    customerName: str
    instanceName: str

@app.post("/api/hdl/upload-to-oracle")
async def upload_zip_to_oracle(req: OracleUploadRequest):
    try:
        customerName = req.customerName
        instanceName = req.instanceName
        oracle_env, username, password = load_oracle_credentials(customerName, instanceName)

        url = f"{oracle_env}/hcmRestApi/resources/11.13.18.05/dataLoadDataSets/action/uploadFile"

        # ✅ Build payload strictly per schema
        payload = {
            "fileName": req.fileName,
            "contentId": req.contentId,
            "content": req.content,  # base64-encoded zip
            "fileEncryption": req.fileEncryption or "NONE"
        }

        headers = {"Content-Type": "application/vnd.oracle.adf.action+json"}

        logger.info("Uploading HDL file to Oracle")
        logger.debug("Final Oracle payload: %s", json.dumps(payload, indent=2))

        res = requests.post(url, json=payload, auth=(username, password), headers=headers)
        res.raise_for_status()
        response = {
            "status_code": res.status_code,
            "response_text": res.json(),
            "env_creds": {
                "oracle_env": oracle_env,
                "username": username,
                "password": password,
            }
        }
        return response

    except Exception as e:
        logger.exception("HDL upload failed")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/hdl/trigger-oracle-job")
async def trigger_hdl_job(req: HDLTriggerRequest):
    try:
        # ✅ Extract values properly from req, not the class
        customerName = req.customerName
        instanceName = req.instanceName
        oracle_env, username, password = load_oracle_credentials(customerName, instanceName)

        url = f"{oracle_env}/hcmRestApi/resources/11.13.18.05/dataLoadDataSets/action/createFileDataSet"

        # ✅ Build payload strictly according to schema
        payload = {
            "deleteSourceFileFlag": req.deleteSourceFileFlag,
            "dataSetName": req.dataSetName,
            "contentId": req.contentId,
            "loadConcurrentThreads": req.loadConcurrentThreads,
            "importMaxErrorPercentage": req.importMaxErrorPercentage,
            "verificationKey": req.verificationKey,
            "loadMaxErrorPercentage": req.loadMaxErrorPercentage,
            "fileEncryption": req.fileEncryption or "NONE",
            "fileAction": req.fileAction or "IMPORT_AND_LOAD",
            "importConcurrentThreads": req.importConcurrentThreads,
        }

        # 🔒 Remove keys with None (Oracle rejects unknown/empty props)
        payload = {k: v for k, v in payload.items() if v is not None}

        headers = {
            "Content-Type": "application/vnd.oracle.adf.action+json",
            "Accept": "application/json"
        }

        logger.info("Triggering HDL job | DataSet=%s | ContentId=%s", 
                    payload.get("dataSetName"), payload.get("contentId"))
        logger.debug("Final Oracle payload: %s", json.dumps(payload, indent=2))

        res = requests.post(url, auth=(username, password), headers=headers, json=payload)

        if not res.ok:
            logger.warning("Oracle HDL job trigger failed | Status=%s | Response=%s",
                           res.status_code, res.text)
            raise HTTPException(status_code=res.status_code,
                                detail=f"Trigger HDL job failed: {res.text}")

        data = res.json()
        request_id = (
            data.get("RequestId")
            or data.get("requestId")
            or data.get("result", {}).get("RequestId")
            or "UNKNOWN"
        )

        return {
            "message": "Trigger job request submitted successfully.",
            "RequestId": request_id,
        }

    except Exception as e:
        logger.exception("HDL job trigger failed")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/hdl/status/{customerName}/{instanceName}/{request_id}")
def get_status_by_request_id(customerName: str, instanceName: str, request_id: str):
    """
    Retrieves the status of an Oracle HDL data load request by its request ID.

    Args:
        customerName (str): Customer name to resolve Oracle credentials.
        instanceName (str): Instance name (e.g., Prod, Test).
        request_id (str): The ID of the HDL data load request.

    Returns:
        dict: A dictionary containing the full Oracle response.

    Raises:
        HTTPException: If there's an error during the API call or environment variable retrieval.
    """
    try:
        # Resolve environment + credentials from helper
        oracle_env, username, password = load_oracle_credentials(customerName, instanceName)

        if not oracle_env or not username or not password:
            logging.error(f"❌ Missing Oracle credentials for {customerName}/{instanceName}")
            raise HTTPException(
                status_code=500,
                detail=f"Server configuration error: Oracle credentials missing for {customerName}/{instanceName}."
            )

        # Ensure oracle_env does not end with a slash
        oracle_env = oracle_env.rstrip('/')

        # Build Oracle status check URL
        url = f"{oracle_env}/hcmRestApi/resources/11.13.18.05/dataLoadDataSets/{request_id}"
        logging.info(f"🔍 Fetching HDL status from Oracle for RequestId={request_id}, URL={url}")

        headers = {
            "Accept": "application/json"
        }

        # Make the GET request to Oracle
        res = requests.get(url, auth=(username, password), headers=headers)

        if not res.ok:
            logging.error(f"❌ Oracle HDL Status Failed: HTTP {res.status_code} - {res.text}")
            raise HTTPException(status_code=res.status_code, detail=res.text)

        data = res.json()
        logging.info(f"✅ Successfully retrieved HDL status for RequestId={request_id}")

        return {
            "requestId": request_id,
            "oracle_response": data
        }

    except requests.exceptions.ConnectionError as ce:
        logging.error(f"🌐 Connection Error to Oracle: {ce}")
        raise HTTPException(status_code=503, detail=f"Failed to connect to Oracle environment: {ce}")
    except requests.exceptions.Timeout as te:
        logging.error(f"⏳ Timeout while connecting to Oracle: {te}")
        raise HTTPException(status_code=504, detail=f"Request to Oracle timed out: {te}")
    except requests.exceptions.RequestException as req_e:
        logging.error(f"⚠️ Unexpected request error: {req_e}")
        raise HTTPException(status_code=500, detail=f"Unexpected Oracle request error: {req_e}")
    except HTTPException:
        raise
    except Exception as e:
        logging.critical(f"🔥 Unhandled error checking HDL status: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to check HDL status due to internal error: {e}")




@app.get("/api/hdl/status/byContentId/{content_id}/{oracle_env}/{username}/{password}")
def get_status_by_content_id(content_id: str):
    try:
        oracle_env = os.getenv("ORACLE_ENV")
        username = os.getenv("ORACLE_USERNAME")
        password = os.getenv("ORACLE_PASSWORD")

        url = f"{oracle_env}/hcmRestApi/resources/11.13.18.05/dataLoadSubmissions?q=ContentId={content_id}"
        headers = { "Accept": "application/json" }

        res = requests.get(url, auth=(username, password), headers=headers)

        # ✅ If Oracle says 404 – don't crash, return gracefully
        if res.status_code == 404:
            return {
                "status": "NOT_FOUND",
                "requestId": "UNKNOWN"
            }

        if not res.ok:
            logger.info("❌ Oracle HDL Status Failed:", res.status_code, res.text)
            raise HTTPException(status_code=500, detail="Failed to fetch Oracle job status.")

        data = res.json()
        if not data.get("items"):
            return {
                "status": "NOT_FOUND",
                "requestId": "UNKNOWN"
            }

        job = data["items"][0]

        return {
            "status": job.get("Status") or "UNKNOWN",
            "requestId": job.get("RequestId") or "UNKNOWN",
            "submissionRef": job.get("SubmissionReference") or "UNKNOWN"
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/hdl/errors/{customerName}/{instanceName}/{request_id}")
async def get_oracle_errors(customerName: str, instanceName: str, request_id: str):
    """
    Retrieves the error messages for a given Oracle HDL data load request.

    Args:
        customerName (str): Customer name to resolve Oracle credentials.
        instanceName (str): Instance name (e.g., Prod, Test).
        request_id (str): The ID of the HDL data load request.

    Returns:
        dict: Oracle error messages response.
    """
    try:
        # Resolve environment + credentials
        oracle_env, username, password = load_oracle_credentials(customerName, instanceName)

        if not oracle_env or not username or not password:
            logging.error(f"❌ Missing Oracle credentials for {customerName}/{instanceName}")
            raise HTTPException(
                status_code=500,
                detail=f"Server configuration error: Oracle credentials missing for {customerName}/{instanceName}."
            )

        # Ensure no trailing slash in URL
        oracle_env = oracle_env.rstrip('/')

        # Oracle error messages API endpoint
        url = (
            f"{oracle_env}/hcmRestApi/resources/11.13.18.05/dataLoadDataSets/{request_id}/child/messages"
            "?totalResults=true"
            "&orderBy=DatFileName,FileLine"
            "&fields=DatFileName,BusinessObjectDiscriminator,OriginatingProcessCode,FileLine,"
            "ConcatenatedUserKey,SourceSystemOwner,SourceSystemId,SourceReference001,"
            "MessageTypeCode,MessageText,MessageUserDetails"
            "&onlyData=true"
        )
        logging.info(f"🔍 Fetching Oracle HDL error messages for RequestId={request_id}, URL={url}")

        headers = {"Accept": "application/json"}
        res = requests.get(url, auth=(username, password), headers=headers)

        if not res.ok:
            logging.error(f"❌ Oracle HDL Errors Fetch Failed: HTTP {res.status_code} - {res.text}")
            raise HTTPException(status_code=res.status_code, detail=res.text)

        data = res.json()
        logging.info(f"✅ Successfully retrieved HDL error messages for RequestId={request_id}")

        return {
            "requestId": request_id,
            "oracle_response": data
        }

    except requests.exceptions.ConnectionError as ce:
        logging.error(f"🌐 Connection Error to Oracle: {ce}")
        raise HTTPException(status_code=503, detail=f"Failed to connect to Oracle environment: {ce}")
    except requests.exceptions.Timeout as te:
        logging.error(f"⏳ Timeout while fetching Oracle errors: {te}")
        raise HTTPException(status_code=504, detail=f"Request to Oracle timed out: {te}")
    except requests.exceptions.RequestException as req_e:
        logging.error(f"⚠️ Unexpected Oracle request error: {req_e}")
        raise HTTPException(status_code=500, detail=f"Unexpected Oracle request error: {req_e}")
    except HTTPException:
        raise
    except Exception as e:
        logging.critical(f"🔥 Unhandled error fetching Oracle HDL errors: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Internal server error: {e}")


class InstanceModel(BaseModel):
    instanceName: str
    oracleUrl: str
    oracleUsername: str
    oraclePassword: str

class CustomerModel(BaseModel):
    customerName: str
    instances: List[InstanceModel]

DATA_EXCEL_FILE_PATH = "Required_files/Hiearchy_data.xlsx"

@app.post("/api/customers")
def sync_hierarchy_with_customers(customers: List[CustomerModel]):
    """
    Synchronizes customer and instance hierarchy with the main Excel file.
    For new Level-1 (customerName) and Level-2 (instanceName) pairs,
    it duplicates Level-1 entries from Hiearchy_data.xlsx as Level-3 entries,
    and subsequent levels (Level-2 to Level-8) from Hiearchy_data.xlsx
    are mapped to Level-4 to Level-10 respectively in the main file.
    """
    try:
        logger.info("Starting hierarchy synchronization process.")

        # --- File Existence Checks ---
        # Explicitly convert to Path object before calling .exists() to ensure type correctness
        if not Path(EXCEL_FILE_PATH).exists():
            logger.error(f"Main hierarchy file not found: {EXCEL_FILE_PATH}")
            raise HTTPException(status_code=404, detail=f"Main hierarchy file not found at {EXCEL_FILE_PATH}")
        if not Path(DATA_EXCEL_FILE_PATH).exists():
            logger.error(f"Hierarchy data source file not found: {DATA_EXCEL_FILE_PATH}")
            raise HTTPException(status_code=404, detail=f"Hierarchy data source file not found at {DATA_EXCEL_FILE_PATH}")

        # --- Load DataFrames ---
        df = pd.read_excel(EXCEL_FILE_PATH)
        dataf = pd.read_excel(DATA_EXCEL_FILE_PATH)
        logger.info("Successfully loaded main hierarchy and data source files.")

        added_pairs = []
        skipped_pairs = []
        all_new_rows_to_add = [] # This list will collect all new rows before concatenating them to df

        # Normalize existing Level-1 and Level-2 pairs from the main hierarchy DataFrame
        # Using a set for efficient lookup
        existing_hierarchy_pairs = set(
            (str(row.get("Level-1", "")).strip(), str(row.get("Level-2", "")).strip())
            for _, row in df.iterrows()
        )
        logger.info(f"Loaded {len(existing_hierarchy_pairs)} existing Level-1/Level-2 pairs.")

        # The full dataf DataFrame will be used for duplication, not just unique Level-1s
        # This ensures all levels from dataf are considered for mapping
        logger.info(f"Using full data source for duplication, containing {len(dataf)} rows.")

        # --- Process Incoming Customer Data ---
        for customer in customers:
            for instance in customer.instances:
                current_l1 = customer.customerName.strip()
                current_l2 = instance.instanceName.strip()
                current_pair = (current_l1, current_l2)

                # Check if the Level-1/Level-2 pair already exists in the main hierarchy
                # Note: This check is for the specific Level-1/Level-2 pair.
                # If a pair exists, no new blank row or duplicated hierarchy will be added for it.
                if current_pair in existing_hierarchy_pairs:
                    skipped_pairs.append(current_pair)
                    logger.debug(f"Skipping existing pair: {current_pair}")
                    continue  # Move to the next instance

                # If it's a new pair, add it to our tracking and prepare for insertion
                existing_hierarchy_pairs.add(current_pair) # Add to set to prevent duplicate processing in same run
                added_pairs.append(current_pair)
                logger.info(f"Found new pair to add: {current_pair}")

                # --- Step 1: Add the initial blank row for the new Level-1/Level-2 pair ---
                # This row represents the new customer/instance entry itself.
                blank_row_data = {col: "" for col in df.columns}
                blank_row_data["Level-1"] = current_l1
                blank_row_data["Level-2"] = current_l2
                all_new_rows_to_add.append(blank_row_data)
                logger.debug(f"Prepared initial blank row for {current_pair}.")

                # --- Step 2: Duplicate full hierarchy data from dataf for this new pair ---
                # Iterate through each row of the hierarchy data source (dataf)
                for _, data_row in dataf.iterrows():
                    # Create a new row, ensuring it has all columns present in the main df
                    duplicated_row_data = {col: "" for col in df.columns}

                    # Populate Template Name and File Name from the source hierarchy data (dataf)
                    duplicated_row_data["Template Name"] = data_row.get("Template Name", "")
                    duplicated_row_data["File Name"] = data_row.get("File Name", "")

                    # Set Level-1 and Level-2 to the newly added customer/instance pair
                    duplicated_row_data["Level-1"] = current_l1
                    duplicated_row_data["Level-2"] = current_l2

                    # Map Level-1 from dataf to Level-3 in the main df, and so on.
                    # Hiearchy_data.xlsx has levels up to Level-8.
                    # Main HDL_BO_Hierarchy_All_Objects_Charlie.xlsx has levels up to Level-10.
                    # So, Level-X from dataf maps to Level-(X+2) in the main df.
                    for i in range(1, 9): # Iterate for Level-1 to Level-8 from dataf
                        source_level_col = f"Level-{i}"
                        target_level_col = f"Level-{i+2}" # Map to Level-3, Level-4, ..., Level-10

                        if source_level_col in data_row and target_level_col in df.columns:
                            duplicated_row_data[target_level_col] = data_row.get(source_level_col, "")

                    # Populate Mandatory_Objects from the source hierarchy data (dataf) if it exists
                    if "Mandatory_Objects" in data_row and "Mandatory_Objects" in df.columns:
                        duplicated_row_data["Mandatory_Objects"] = data_row.get("Mandatory_Objects", False) # Default to False

                    all_new_rows_to_add.append(duplicated_row_data)
                logger.debug(f"Prepared duplicated full hierarchy rows for {current_pair}.")

        # --- Append All New Rows and Save ---
        if all_new_rows_to_add:
            # Concatenate all collected new rows to the main DataFrame
            df = pd.concat([df, pd.DataFrame(all_new_rows_to_add)], ignore_index=True)
            logger.info(f"Appended {len(all_new_rows_to_add)} new rows to the main hierarchy DataFrame.")
        else:
            logger.info("No new rows to add to the main hierarchy DataFrame.")

        # Save the updated DataFrame back to the Excel file
        # Use index=False to prevent writing the DataFrame index as a column in the Excel file
        df.to_excel(EXCEL_FILE_PATH, index=False)
        logger.info(f"Successfully saved updated hierarchy to {EXCEL_FILE_PATH}.")
        # Build env-friendly payload
        env_payload = []
        for cust in customers:
            env_payload.append({
                "customerName": str(cust.customerName),
                "instances": [
                    {
                        "instanceName": str(inst.instanceName),
                        "oracleUrl": str(inst.oracleUrl),
                        "oracleUsername": str(inst.oracleUsername),
                        "oraclePassword": str(inst.oraclePassword),
                    }
                    for inst in cust.instances
                ]
            })

        merge_env_files(customers=env_payload)
        load_dotenv(".env", override=True)

        # --- Return Response ---
        return {
            "message": "Customer hierarchy sync complete.",
            "added": [f"{l1} - {l2}" for l1, l2 in added_pairs],
            "skipped": [f"{l1} - {l2}" for l1, l2 in skipped_pairs],
            "total_added": len(added_pairs),
            "total_skipped": len(skipped_pairs),
            "output_file": str(EXCEL_FILE_PATH)
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"An unexpected error occurred during hierarchy sync: {e}")
        raise HTTPException(status_code=500, detail=f"An unexpected error occurred during hierarchy sync: {e}")


class CustomerUpdateModel(BaseModel):
    old_customerName: str
    old_instanceName: str
    new_customerName: str
    new_instanceName: str

@app.put("/api/customers")
def update_customer_instance(update: CustomerUpdateModel):
    try:
        if not EXCEL_FILE_PATH.exists():
            raise HTTPException(status_code=404, detail="Main Excel file not found.")

        df = pd.read_excel(EXCEL_FILE_PATH)
        if df.empty:
            raise HTTPException(status_code=400, detail="Excel file is empty.")

        mask = (
            df["Level-1"].astype(str).str.strip() == update.old_customerName.strip()
        ) & (
            df["Level-2"].astype(str).str.strip() == update.old_instanceName.strip()
        )

        if not mask.any():
            raise HTTPException(status_code=404, detail="Customer/Instance pair not found.")

        df.loc[mask, "Level-1"] = update.new_customerName.strip()
        df.loc[mask, "Level-2"] = update.new_instanceName.strip()

        df.to_excel(EXCEL_FILE_PATH, index=False)
        logger.info(f"Updated ({update.old_customerName}, {update.old_instanceName}) to ({update.new_customerName}, {update.new_instanceName})")

        return {
            "message": "Customer and instance updated successfully.",
            "updated_rows": int(mask.sum())
        }

    except Exception as e:
        logger.error(f"Update failed: {e}")
        raise HTTPException(status_code=500, detail=f"Update failed: {str(e)}")


class CustomerDeleteModel(BaseModel):
    customerName: str
    instanceName: str

@app.delete("/api/customers")
def delete_customer_instance(delete_req: CustomerDeleteModel):
    try:
        if not EXCEL_FILE_PATH.exists():
            raise HTTPException(status_code=404, detail="Main Excel file not found.")

        df = pd.read_excel(EXCEL_FILE_PATH)
        if df.empty:
            raise HTTPException(status_code=400, detail="Excel file is empty.")

        original_count = len(df)
        mask = ~(
            (df["Level-1"].astype(str).str.strip() == delete_req.customerName.strip()) &
            (df["Level-2"].astype(str).str.strip() == delete_req.instanceName.strip())
        )

        new_df = df[mask]
        removed_count = original_count - len(new_df)

        if removed_count == 0:
            raise HTTPException(status_code=404, detail="No matching customer/instance pair found to delete.")

        new_df.to_excel(EXCEL_FILE_PATH, index=False)
        logger.info(f"Deleted {removed_count} rows for customer {delete_req.customerName} and instance {delete_req.instanceName}")

        return {
            "message": "Customer/Instance pair deleted successfully.",
            "deleted_rows": removed_count
        }

    except Exception as e:
        logger.error(f"Delete failed: {e}")
        raise HTTPException(status_code=500, detail=f"Delete failed: {str(e)}")







#Env population
# -------------------- 🚀 Helper Function ---------------------
def format_env_filename(customer: str, instance: str):
    # Remove spaces and special chars
    safe_customer = "".join(e for e in customer if e.isalnum())
    safe_instance = "".join(e for e in instance if e.isalnum())
    return f"{safe_customer}_{safe_instance}.env"

def write_env_file(file_path: Path, env_data: dict):
    with file_path.open("w") as f:
        for key, val in env_data.items():
            f.write(f'{key}="{val}"\n')  # Safely quote values

# -------------------- 🎯 API Endpoint ---------------------
def sanitize(text: str) -> str:
    return "".join(e for e in text if e.isalnum()).upper()

def reload_env():
    load_dotenv(dotenv_path=".env", override=True)

def write_env_file_for_customers(customers: List[CustomerModel], env_path=".env"):
    env_lines = []
    
    for customer in customers:
        customer_key = customer.customerName.replace(" ", "").upper()
        for instance in customer.instances:
            instance_key = instance.instanceName.replace(" ", "").upper()
            prefix = f"{customer_key}_{instance_key}"
            # Remove the last "/" if present in URL
            instance.oracleUrl = instance.oracleUrl.rstrip("/") if instance.oracleUrl else ""
            # Only write if all fields are present
            if instance.oracleUrl and instance.oracleUsername and instance.oraclePassword:
                env_lines.extend([
                    f"{prefix}_ORACLE_URL={instance.oracleUrl.strip()}",
                    f"{prefix}_ORACLE_USERNAME={instance.oracleUsername.strip()}",
                    f"{prefix}_ORACLE_PASSWORD={instance.oraclePassword.strip()}",
                ])
            else:
                print(f"⚠️ Missing credentials for {prefix}. Skipping...")

    # Write to .env file
    env_file = Path(env_path).resolve()
    with env_file.open("w", encoding="utf-8") as f:
        f.write("\n".join(env_lines))

    print(f"✅ .env saved at: {env_file}")
    load_dotenv(dotenv_path=env_file, override=True)

    # Print sample env check
    if env_lines:
        sample_key = env_lines[0].split("=")[0]
        print(f"🔍 Sample env check: {sample_key} = {os.getenv(sample_key)}")


@app.post("/api/save-env")
def save_envs(customers: List[CustomerModel]):
    try:
        env_lines = []

        for customer in customers:
            customer_key = customer.customerName.replace(" ", "").upper()
            for instance in customer.instances:
                instance_key = instance.instanceName.replace(" ", "").upper()
                prefix = f"{customer_key}_{instance_key}"

                # Skip if any critical field is missing
                if not instance.oracleUrl or not instance.oracleUsername or not instance.oraclePassword:
                    continue
                # Remove trailing slash from URL if present
                instance.oracleUrl = instance.oracleUrl.rstrip("/")
                env_lines.append(f"{prefix}_ORACLE_URL=\"{instance.oracleUrl}\"")
                env_lines.append(f"{prefix}_ORACLE_USERNAME=\"{instance.oracleUsername}\"")
                env_lines.append(f"{prefix}_ORACLE_PASSWORD=\"{instance.oraclePassword}\"")

        if not env_lines:
            raise HTTPException(status_code=400, detail="No valid instance data to write to .env")

        env_path = Path(".env")
        with env_path.open("w", encoding="utf-8") as f:
            f.write("\n".join(env_lines) + "\n")

        load_dotenv(dotenv_path=env_path, override=True)

        return {
            "message": "✅ Environment variables saved",
            "lines_written": len(env_lines),
            "env_path": str(env_path.resolve())
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"❌ Failed to save .env file: {str(e)}")

@app.get("/api/customers")
def get_customers_from_root_env():
    try:
        env_path = Path(".env")
        if not env_path.exists():
            raise HTTPException(status_code=404, detail=".env file not found at root")

        env_data = dotenv_values(env_path)
        print("🌱 Loaded .env data:", env_data)

        customer_map = {}

        for full_key, value in env_data.items():
            print(f"🔍 Parsing key: {full_key} = {value}")
            if "_ORACLE_" not in full_key:
                continue

            try:
                prefix, field = full_key.split("_ORACLE_", 1)
                field = field.lower()

                parts = prefix.split("_")
                if len(parts) < 2:
                    print(f"⚠️ Skipping malformed key: {full_key}")
                    continue

                customer = "_".join(parts[:-1])
                instance = parts[-1]

                if customer not in customer_map:
                    customer_map[customer] = {}

                if instance not in customer_map[customer]:
                    customer_map[customer][instance] = {
                        "instanceName": instance
                    }

                customer_map[customer][instance][f"oracle{field.capitalize()}"] = value

            except Exception as inner_e:
                print(f"🚨 Error parsing key: {full_key} — {inner_e}")
                continue

        print("✅ Final parsed customer map:", customer_map)

        response_data = []
        for customer, instances in customer_map.items():
            instance_list = list(instances.values())
            response_data.append({
                "customerName": customer,
                "instances": instance_list
            })

        return response_data

    except Exception as e:
        print("❌ Fatal error in get_customers_from_root_env:", e)
        raise HTTPException(status_code=500, detail="Could not load customers from .env")

def read_excel_sheet_with_dynamic_header(file_stream: io.BytesIO, sheet_name: str, header_row_index: int = 0) -> pd.DataFrame:
    """
    Reads a specific sheet from an Excel file stream, inferring the header from a specific row index (0-based).
    Skips rows above the header and uses the specified row as column names.
    It also cleans column names by stripping whitespace and replacing special characters.

    Args:
        file_stream (io.BytesIO): The byte stream of the Excel file.
        sheet_name (str): The name of the sheet to read.
        header_row_index (int): The 0-based index of the row to be used as the header.

    Returns:
        pandas.DataFrame: The loaded and cleaned DataFrame, or an empty DataFrame if an error occurs.
    """
    try:
        logging.info(f"Attempting to read sheet '{sheet_name}' from file stream with header at row {header_row_index}.")
        # Ensure the stream is at the beginning before reading each sheet
        file_stream.seek(0)
        df = pd.read_excel(file_stream, sheet_name=sheet_name, skiprows=header_row_index, header=0, engine='openpyxl')
        
        # Drop any completely empty columns that might result from uneven header rows
        original_cols = df.columns.tolist()
        df = df.dropna(axis=1, how='all')
        if len(df.columns) < len(original_cols):
            logging.warning(f"Dropped {len(original_cols) - len(df.columns)} empty columns from sheet '{sheet_name}'.")

        # Clean column names: strip whitespace, replace special characters
        # This regex replaces non-alphanumeric characters (except underscore) and then spaces with underscores.
        df.columns = df.columns.str.strip().str.replace(r'[^a-zA-Z0-9_]', '', regex=True).str.replace(' ', '_', regex=False)
        logging.info(f"Successfully loaded sheet '{sheet_name}'. Columns: {df.columns.tolist()}")
        return df
    except KeyError:
        logging.error(f"Error: Sheet '{sheet_name}' not found in the uploaded Excel file. Please check the sheet name.")
        return pd.DataFrame()
    except Exception as e:
        logging.error(f"An unexpected error occurred while reading sheet '{sheet_name}' from file stream: {e}", exc_info=True)
        return pd.DataFrame()

def format_date(date_series: pd.Series) -> pd.Series:
    """
    Converts a pandas Series to datetime objects and formats them as 'YYYY-MM-DD' strings.
    Invalid dates are coerced to NaT (Not a Time) and then replaced with empty strings.

    Args:
        date_series (pandas.Series): The series containing date values.

    Returns:
        pandas.Series: The series with dates formatted as 'YYYY-MM-DD' strings, or empty strings.
    """
    if date_series.empty:
        return pd.Series([], dtype='object')
    return pd.to_datetime(date_series, errors='coerce').dt.strftime('%Y-%m-%d').replace({np.nan: ''})

def validate_columns(df: pd.DataFrame, required_columns: list, df_name: str) -> bool:
    """
    Checks if all required columns are present in the DataFrame.
    Logs a warning if any required column is missing.

    Args:
        df (pandas.DataFrame): The DataFrame to validate.
        required_columns (list): A list of column names that must be present.
        df_name (str): The name of the DataFrame for logging purposes.

    Returns:
        bool: True if all required columns are present, False otherwise.
    """
    if df.empty:
        logging.warning(f"DataFrame '{df_name}' is empty. Skipping column validation.")
        return False
    
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        logging.warning(f"Missing required columns in '{df_name}': {', '.join(missing_columns)}. This might affect conversion accuracy.")
        return False
    return True

# --- Conversion Functions ---

def convert_worker_data(person_df: pd.DataFrame, work_relationship_df: pd.DataFrame) -> pd.DataFrame:
    required_person_cols = ['PersonNumber', 'EarliestHireDate', 'EffectiveEndDate', 'DateOfBirth', 'Country']
    required_work_rel_cols = ['PersonNumber', 'ActionCode', 'ReasonCode', 'DateStart']

    if not validate_columns(person_df, required_person_cols, 'Person') or \
       not validate_columns(work_relationship_df, required_work_rel_cols, 'WorkRelationship'):
        logging.warning("Skipping Worker data conversion due to missing critical source columns.")
        return pd.DataFrame(columns=['Employee Number', 'Start Date', 'End Date', 'Hire Date', 'Action Code', 'Reason Code', 'Birth Date', 'Country of Birth', 'Region of Birth', 'Town of Birth', 'Correspondence Language', 'Blood Type'])

    person_df = person_df.rename(columns={'PersonNumber': 'Employee Number'})
    work_relationship_df = work_relationship_df.rename(columns={'PersonNumber': 'Employee Number'})

    hire_work_rel = work_relationship_df[work_relationship_df['ActionCode'] == 'HIRE'].copy()
    if not hire_work_rel.empty:
        hire_work_rel['DateStart'] = pd.to_datetime(hire_work_rel['DateStart'], errors='coerce')
        hire_work_rel = hire_work_rel.sort_values(by=['Employee Number', 'DateStart']).drop_duplicates(subset=['Employee Number'], keep='first')
    else:
        logging.warning("No 'HIRE' records found in WorkRelationship data. Worker 'Action Code' and 'Reason Code' might be incomplete.")

    worker_df = pd.merge(person_df, hire_work_rel[['Employee Number', 'ActionCode', 'ReasonCode']], on='Employee Number', how='left')

    worker_df_converted = pd.DataFrame()
    worker_df_converted['Employee Number'] = worker_df['Employee Number']
    worker_df_converted['Start Date'] = format_date(worker_df['EarliestHireDate'])
    worker_df_converted['End Date'] = format_date(worker_df['EffectiveEndDate'])
    worker_df_converted['Hire Date'] = format_date(worker_df['EarliestHireDate'])
    worker_df_converted['Action Code'] = worker_df['ActionCode'].fillna('HIRE')
    worker_df_converted['Reason Code'] = worker_df['ReasonCode'].fillna('')
    worker_df_converted['Birth Date'] = format_date(worker_df['DateOfBirth'])
    
    worker_df_converted['Country of Birth'] = ''
    worker_df_converted['Region of Birth'] = ''
    worker_df_converted['Town of Birth'] = ''
    worker_df_converted['Correspondence Language'] = ''
    worker_df_converted['Blood Type'] = ''

    return worker_df_converted

def convert_person_name_data(person_df: pd.DataFrame) -> pd.DataFrame:
    required_cols = ['PersonNumber', 'EarliestHireDate', 'LastName', 'FirstName', 'NameSuffix', 'MiddleNames', 'KnownAs', 'PreviousLastName', 'Title', 'Country']
    if not validate_columns(person_df, required_cols, 'Person'):
        logging.warning("Skipping PersonName data conversion due to missing critical source columns.")
        return pd.DataFrame(columns=['Employee Number', 'Start Date', 'Last Name', 'First Name', 'Suffix', 'Middle Names', 'Known As', 'Previous Last Name', 'Title', 'Country Code', 'Name Type'])

    person_name_df = person_df.copy()
    person_name_df_converted = pd.DataFrame()

    person_name_df_converted['Employee Number'] = person_name_df['PersonNumber']
    person_name_df_converted['Start Date'] = format_date(person_name_df['EarliestHireDate'])
    person_name_df_converted['Last Name'] = person_name_df['LastName']
    person_name_df_converted['First Name'] = person_name_df['FirstName']
    person_name_df_converted['Suffix'] = person_name_df['NameSuffix'].fillna('')
    person_name_df_converted['Middle Names'] = person_name_df['MiddleNames'].fillna('')
    person_name_df_converted['Known As'] = person_name_df['KnownAs'].fillna('')
    person_name_df_converted['Previous Last Name'] = person_name_df['PreviousLastName'].fillna('')
    person_name_df_converted['Title'] = person_name_df['Title'].fillna('')
    person_name_df_converted['Country Code'] = person_name_df['Country'].fillna('')
    person_name_df_converted['Name Type'] = 'GLOBAL'

    return person_name_df_converted

def convert_work_relationship_data(work_relationship_df: pd.DataFrame, assignment_df: pd.DataFrame) -> pd.DataFrame:
    required_work_rel_cols = ['PersonNumber', 'DateStart', 'ActualTerminationDate', 'ActionCode', 'ReasonCode', 'LegalEmployerName', 'WorkerType', 'PrimaryFlag', 'RehireRecommendationFlag']
    required_assignment_cols = ['PersonNumber', 'EffectiveStartDate', 'EffectiveEndDate', 'BusinessUnitShortCode', 'JobCode', 'GradeCode', 'LocationCode', 'AssignmentStatusType', 'AssignmentType', 'HourlySalariedCode', 'ManagerFlag', 'WorkingatHome', 'WorkerPeriodType', 'NormalHours', 'Frequency', 'PrimaryAssignmentFlag', 'DepartmentName', 'UserPersonType']

    if not validate_columns(work_relationship_df, required_work_rel_cols, 'WorkRelationship') or \
       not validate_columns(assignment_df, required_assignment_cols, 'Assignment'):
        logging.warning("Skipping WorkRelationship data conversion due to missing critical source columns.")
        return pd.DataFrame(columns=['Employee Number', 'Start Date', 'End Date', 'Effective Sequence', 'Latest Change Flag', 'Action Code', 'Reason Code', 'Legal Employer Name', 'Business Unit Short Code', 'Department Name', 'Job Code', 'Grade Code', 'Location Code', 'Assignment Status Type Code', 'Assignment Type', 'Hourly Salaried Code', 'Manager Flag', 'Work At Home Flag', 'Permanent Temporary', 'Full Part Time', 'Normal Hours', 'Frequency', 'Primary Assignment Flag', 'Worker Type', 'Person Type Code', 'System Person Type', 'Start Date', 'Actual Termination Date', 'Rehire Recommendation Flag', 'Global Transfer Flag', 'PrimaryFlag'])

    work_rel_df = work_relationship_df.rename(columns={'PersonNumber': 'Employee Number'})
    asg_df = assignment_df.rename(columns={'PersonNumber': 'Employee Number'})

    work_rel_df['DateStart'] = pd.to_datetime(work_rel_df['DateStart'], errors='coerce')
    work_rel_df['ActualTerminationDate'] = pd.to_datetime(work_rel_df['ActualTerminationDate'], errors='coerce')
    asg_df['EffectiveStartDate'] = pd.to_datetime(asg_df['EffectiveStartDate'], errors='coerce')
    asg_df['EffectiveEndDate'] = pd.to_datetime(asg_df['EffectiveEndDate'], errors='coerce')

    work_rel_df = work_rel_df.sort_values(by=['Employee Number', 'DateStart'])
    asg_df_sorted = asg_df.sort_values(by=['Employee Number', 'EffectiveStartDate'])

    merged_df = pd.merge_asof(
        work_rel_df,
        asg_df_sorted,
        left_on='DateStart',
        right_on='EffectiveStartDate',
        by='Employee Number',
        direction='backward',
        suffixes=('_x', '_y')
    )

    merged_df['Effective Sequence'] = merged_df.groupby('Employee Number').cumcount() + 1
    merged_df['Latest Change Flag'] = merged_df.groupby('Employee Number')['DateStart'].transform(lambda x: x == x.max()).map({True: 'Y', False: 'N'})

    work_rel_converted = pd.DataFrame()
    work_rel_converted['Employee Number'] = merged_df['Employee Number']
    work_rel_converted['Start Date'] = format_date(merged_df['DateStart'])
    work_rel_converted['End Date'] = format_date(merged_df['ActualTerminationDate'])
    work_rel_converted['Effective Sequence'] = merged_df['Effective Sequence']
    work_rel_converted['Latest Change Flag'] = merged_df['Latest Change Flag']
    work_rel_converted['Action Code'] = merged_df['ActionCode_x'].fillna('')
    work_rel_converted['Reason Code'] = merged_df['ReasonCode_x'].fillna('')
    work_rel_converted['Legal Employer Name'] = merged_df['LegalEmployerName_x'].fillna('')
    work_rel_converted['Business Unit Short Code'] = merged_df['BusinessUnitShortCode'].fillna('')
    work_rel_converted['Department Name'] = merged_df['DepartmentName'].fillna('') 
    work_rel_converted['Job Code'] = merged_df['JobCode'].fillna('')
    work_rel_converted['Grade Code'] = merged_df['GradeCode'].fillna('')
    work_rel_converted['Location Code'] = merged_df['LocationCode'].fillna('')
    work_rel_converted['Assignment Status Type Code'] = merged_df['AssignmentStatusType'].fillna('')
    work_rel_converted['Assignment Type'] = merged_df['AssignmentType'].fillna('')
    work_rel_converted['Hourly Salaried Code'] = merged_df['HourlySalariedCode'].fillna('')
    work_rel_converted['Manager Flag'] = merged_df['ManagerFlag'].fillna('')
    work_rel_converted['Work At Home Flag'] = merged_df['WorkingatHome'].fillna('')
    work_rel_converted['Permanent Temporary'] = merged_df['WorkerPeriodType'].fillna('') # Mapping WorkerPeriodType to Permanent Temporary
    work_rel_converted['Full Part Time'] = '' # No direct mapping in snippet, leaving empty
    work_rel_converted['Normal Hours'] = merged_df['NormalHours'].fillna('')
    work_rel_converted['Frequency'] = merged_df['Frequency'].fillna('')
    work_rel_converted['Primary Assignment Flag'] = merged_df['PrimaryAssignmentFlag_y'].fillna('')
    work_rel_converted['Worker Type'] = merged_df['WorkerType_x'].fillna('')
    work_rel_converted['Person Type Code'] = merged_df.get('UserPersonType', '').fillna('')
    work_rel_converted['System Person Type'] = merged_df.get('SystemPersonType', '').fillna('')
    work_rel_converted['Start Date'] = format_date(merged_df['DateStart'])
    work_rel_converted['Actual Termination Date'] = format_date(merged_df['ActualTerminationDate'])
    work_rel_converted['Rehire Recommendation Flag'] = merged_df['RehireRecommendationFlag'].fillna('')
    work_rel_converted['Global Transfer Flag'] = ''
    work_rel_converted['PrimaryFlag'] = merged_df['PrimaryFlag_x'].fillna('')

    return work_rel_converted

def convert_work_terms_data(work_relationship_df: pd.DataFrame, assignment_df: pd.DataFrame) -> pd.DataFrame:
    # WorkTerms has the same columns as WorkRelationship in the target format
    return convert_work_relationship_data(work_relationship_df, assignment_df)

def convert_assignment_data(assignment_df: pd.DataFrame) -> pd.DataFrame:
    required_cols = ['PersonNumber', 'EffectiveStartDate', 'EffectiveEndDate', 'ActionCode', 'ReasonCode', 'LegalEmployerName', 'BusinessUnitShortCode', 'DepartmentName', 'JobCode', 'GradeCode', 'LocationCode', 'AssignmentStatusType', 'AssignmentType', 'HourlySalariedCode', 'ManagerFlag', 'WorkingatHome', 'WorkerPeriodType', 'NormalHours', 'Frequency', 'PrimaryAssignmentFlag', 'UserPersonType']
    if not validate_columns(assignment_df, required_cols, 'Assignment'):
        logging.warning("Skipping Assignment data conversion due to missing critical source columns.")
        return pd.DataFrame(columns=['Employee Number', 'Start Date', 'End Date', 'Effective Sequence', 'Latest Change Flag', 'Action Code', 'Reason Code', 'Legal Employer Name', 'Business Unit Short Code', 'Department Name', 'Job Code', 'Grade Code', 'Location Code', 'Assignment Status Type Code', 'Assignment Type', 'Hourly Salaried Code', 'Manager Flag', 'Work At Home Flag', 'Permanent Temporary', 'Full Part Time', 'Normal Hours', 'Frequency', 'Primary Assignment Flag', 'Worker Type', 'Person Type Code', 'System Person Type'])

    asg_df = assignment_df.rename(columns={'PersonNumber': 'Employee Number'})

    asg_df['EffectiveStartDate'] = pd.to_datetime(asg_df['EffectiveStartDate'], errors='coerce')
    asg_df['EffectiveEndDate'] = pd.to_datetime(asg_df['EffectiveEndDate'], errors='coerce')
    
    asg_df = asg_df.sort_values(by=['Employee Number', 'EffectiveStartDate']).reset_index(drop=True)

    asg_df['Effective Sequence'] = asg_df.groupby('Employee Number').cumcount() + 1
    asg_df['Latest Change Flag'] = asg_df.groupby('Employee Number')['EffectiveStartDate'].transform(lambda x: x == x.max()).map({True: 'Y', False: 'N'})

    asg_converted = pd.DataFrame()
    asg_converted['Employee Number'] = asg_df['Employee Number']
    asg_converted['Start Date'] = format_date(asg_df['EffectiveStartDate'])
    asg_converted['End Date'] = format_date(asg_df['EffectiveEndDate'])
    asg_converted['Effective Sequence'] = asg_df['Effective Sequence']
    asg_converted['Latest Change Flag'] = asg_df['Latest Change Flag']
    asg_converted['Action Code'] = asg_df['ActionCode'].fillna('')
    asg_converted['Reason Code'] = asg_df['ReasonCode'].fillna('')
    asg_converted['Legal Employer Name'] = asg_df['LegalEmployerName'].fillna('')
    asg_converted['Business Unit Short Code'] = asg_df['BusinessUnitShortCode'].fillna('')
    asg_converted['Department Name'] = asg_df['DepartmentName'].fillna('')
    asg_converted['Job Code'] = asg_df['JobCode'].fillna('')
    asg_converted['Grade Code'] = asg_df['GradeCode'].fillna('')
    asg_converted['Location Code'] = asg_df['LocationCode'].fillna('')
    asg_converted['Assignment Status Type Code'] = asg_df['AssignmentStatusType'].fillna('')
    asg_converted['Assignment Type'] = asg_df['AssignmentType'].fillna('')
    asg_converted['Hourly Salaried Code'] = asg_df['HourlySalariedCode'].fillna('')
    asg_converted['Manager Flag'] = asg_df['ManagerFlag'].fillna('')
    asg_converted['Work At Home Flag'] = asg_df['WorkingatHome'].fillna('')
    asg_converted['Permanent Temporary'] = asg_df['WorkerPeriodType'].fillna('')
    asg_converted['Full Part Time'] = '' # No direct mapping in snippet, leaving empty
    asg_converted['Normal Hours'] = asg_df['NormalHours'].fillna('')
    asg_converted['Frequency'] = asg_df['Frequency'].fillna('')
    asg_converted['Primary Assignment Flag'] = asg_df['PrimaryAssignmentFlag'].fillna('')
    asg_converted['Worker Type'] = asg_df['WorkerType'].fillna('') # Corrected to 'WorkerType' from HDL header
    asg_converted['Person Type Code'] = asg_df.get('UserPersonType', '').fillna('')
    asg_converted['System Person Type'] = asg_df.get('SystemPersonType', '').fillna('')

    return asg_converted

def convert_contract_data(assignment_df: pd.DataFrame) -> pd.DataFrame:
    required_cols = ['PersonNumber', 'EffectiveStartDate', 'ActionCode']
    if not validate_columns(assignment_df, required_cols, 'Assignment'):
        logging.warning("Skipping Contract data conversion due to missing critical source columns.")
        return pd.DataFrame(columns=['Employee Number', 'Start Date', 'Contract Type', 'Duration', 'Duration Units', 'Action Code'])

    contract_df = assignment_df.rename(columns={'PersonNumber': 'Employee Number'})
    
    contract_df['EffectiveStartDate'] = pd.to_datetime(contract_df['EffectiveStartDate'], errors='coerce')
    contract_df = contract_df.sort_values(by=['Employee Number', 'EffectiveStartDate']).reset_index(drop=True)

    contract_converted = pd.DataFrame()
    contract_converted['Employee Number'] = contract_df['Employee Number']
    contract_converted['Start Date'] = format_date(contract_df['EffectiveStartDate'])
    contract_converted['Contract Type'] = ''
    contract_converted['Duration'] = ''
    contract_converted['Duration Units'] = ''
    contract_converted['Action Code'] = contract_df['ActionCode'].fillna('')

    return contract_converted

def convert_national_id_data(person_df: pd.DataFrame, nat_id_multi_df: pd.DataFrame) -> pd.DataFrame:
    nat_id_converted = pd.DataFrame()

    multi_required_cols = ['PersonNumber', 'LegislationCode', 'NationalIdentifierType', 'NationalIdentifierNumber']
    person_required_cols = ['PersonNumber', 'Country', 'NationalIdentifierType', 'NationalIdentifierNumber'] # Updated based on Person HDL Header

    if not nat_id_multi_df.empty and validate_columns(nat_id_multi_df, multi_required_cols, 'Nat. ID Multi'):
        nat_id_df = nat_id_multi_df.copy()
        nat_id_converted['Employee Number'] = nat_id_df['PersonNumber'].fillna('')
        nat_id_converted['Legislation Code'] = nat_id_df['LegislationCode'].fillna('')
        nat_id_converted['National Identifier Type'] = nat_id_df['NationalIdentifierType'].fillna('')
        nat_id_converted['National Identifier Number'] = nat_id_df['NationalIdentifierNumber'].fillna('')
    elif not person_df.empty and validate_columns(person_df, person_required_cols, 'Person'):
        nat_id_df = person_df.rename(columns={'PersonNumber': 'Employee Number'})
        nat_id_converted['Employee Number'] = nat_id_df['Employee Number']
        nat_id_converted['Legislation Code'] = nat_id_df['Country'].fillna('')
        nat_id_converted['National Identifier Type'] = nat_id_df['NationalIdentifierType'].fillna('')
        nat_id_converted['National Identifier Number'] = nat_id_df['NationalIdentifierNumber'].fillna('')
    else:
        logging.warning("Skipping National ID data conversion due to missing critical source columns in both 'Nat. ID Multi' and 'Person' dataframes.")
        return pd.DataFrame(columns=['Employee Number', 'Legislation Code', 'National Identifier Type', 'National Identifier Number'])

    return nat_id_converted

def convert_person_religion_data(person_df: pd.DataFrame) -> pd.DataFrame:
    required_cols = ['PersonNumber', 'Country']
    if not validate_columns(person_df, required_cols, 'Person'):
        logging.warning("Skipping PersonReligion data creation due to missing critical source columns.")
        return pd.DataFrame(columns=['Employee Number', 'Legislation Code', 'Religion', 'Primary Flag'])

    religion_df = person_df.rename(columns={'PersonNumber': 'Employee Number'})
    religion_converted = pd.DataFrame()
    religion_converted['Employee Number'] = religion_df['Employee Number']
    religion_converted['Legislation Code'] = religion_df['Country'].fillna('US')
    religion_converted['Religion'] = 'Christianity'
    religion_converted['Primary Flag'] = 'Y'

    return religion_converted

def convert_person_address_data(address_df: pd.DataFrame) -> pd.DataFrame:
    required_cols = ['PersonNumber', 'EffectiveStartDate', 'AddressType', 'Country', 'AddressLine1', 'AddressLine2', 'TownOrCity', 'Region2', 'PostalCode', 'Region1']
    if not validate_columns(address_df, required_cols, 'Address'):
        logging.warning("Skipping PersonAddress data conversion due to missing critical source columns.")
        return pd.DataFrame(columns=['Employee Number', 'Start Date', 'Address Type', 'Country', 'Address Line 1', 'Address Line 2', 'Town Or City', 'State', 'Zip Code', 'County'])

    address_df = address_df.rename(columns={'PersonNumber': 'Employee Number'})
    address_converted = pd.DataFrame()

    address_converted['Employee Number'] = address_df['Employee Number']
    address_converted['Start Date'] = format_date(address_df['EffectiveStartDate'])
    address_converted['Address Type'] = address_df['AddressType'].fillna('')
    address_converted['Country'] = address_df['Country'].fillna('')
    address_converted['Address Line 1'] = address_df['AddressLine1'].fillna('')
    address_converted['Address Line 2'] = address_df['AddressLine2'].fillna('')
    address_converted['Town Or City'] = address_df['TownOrCity'].fillna('')
    address_converted['State'] = address_df['Region2'].fillna('')
    address_converted['Zip Code'] = address_df['PostalCode'].fillna('')
    address_converted['County'] = address_df['Region1'].fillna('')

    return address_converted

def convert_person_citizenship_data(citizenship_df: pd.DataFrame) -> pd.DataFrame:
    required_cols = ['PersonNumber', 'DateFrom', 'LegislationCode', 'CitizenshipStatus']
    if not validate_columns(citizenship_df, required_cols, 'Citizenship'):
        logging.warning("Skipping PersonCitizenship data conversion due to missing critical source columns.")
        return pd.DataFrame(columns=['Employee Number', 'Start Date', 'Country Code', 'Status'])

    citizenship_df = citizenship_df.rename(columns={'PersonNumber': 'Employee Number'})
    citizenship_converted = pd.DataFrame()

    citizenship_converted['Employee Number'] = citizenship_df['Employee Number']
    citizenship_converted['Start Date'] = format_date(citizenship_df['DateFrom'])
    citizenship_converted['Country Code'] = citizenship_df['LegislationCode'].fillna('')
    citizenship_converted['Status'] = citizenship_df['CitizenshipStatus'].fillna('')

    return citizenship_converted

def convert_person_email_data(email_df: pd.DataFrame) -> pd.DataFrame:
    required_cols = ['PersonNumber', 'DateFrom', 'EmailAddress', 'PrimaryFlag', 'EmailType']
    if not validate_columns(email_df, required_cols, 'Email'):
        logging.warning("Skipping PersonEmail data conversion due to missing critical source columns.")
        return pd.DataFrame(columns=['Employee Number', 'Start date', 'Email Address', 'Primary Flag', 'Email Type'])

    email_df = email_df.rename(columns={'PersonNumber': 'Employee Number'})
    email_converted = pd.DataFrame()

    email_converted['Employee Number'] = email_df['Employee Number']
    email_converted['Start date'] = format_date(email_df['DateFrom'])
    email_converted['Email Address'] = email_df['EmailAddress'].fillna('')
    email_converted['Primary Flag'] = email_df['PrimaryFlag'].fillna('')
    email_converted['Email Type'] = email_df['EmailType'].fillna('')

    return email_converted

def convert_person_phone_data(phone_df: pd.DataFrame) -> pd.DataFrame:
    """
    Converts data to the 'PersonPhone.csv' format.
    """
    required_cols = ['PersonNumber', 'DateFrom', 'PhoneType', 'LegislationCode', 'AreaCode', 'PhoneNumber', 'PrimaryFlag', 'DateTo']
    if not validate_columns(phone_df, required_cols, 'Phone'):
        logging.warning("Skipping PersonPhone data conversion due to missing critical source columns.")
        return pd.DataFrame(columns=['Employee Number', 'Start Date', 'Phone Type', 'Legislation Code', 'Area Code', 'Phone Number', 'Primary Flag', 'End Date'])

    phone_df = phone_df.rename(columns={'PersonNumber': 'Employee Number'})
    phone_converted = pd.DataFrame()

    phone_converted['Employee Number'] = phone_df['Employee Number']
    phone_converted['Start Date'] = format_date(phone_df['DateFrom'])
    phone_converted['Phone Type'] = phone_df['PhoneType'].fillna('')
    phone_converted['Legislation Code'] = phone_df['LegislationCode'].fillna('')
    phone_converted['Area Code'] = phone_df['AreaCode'].fillna('')
    phone_converted['Phone Number'] = phone_df['PhoneNumber'].fillna('')
    phone_converted['Primary Flag'] = phone_df['PrimaryFlag'].fillna('')
    phone_converted['End Date'] = format_date(phone_df['DateTo'])

    return phone_converted

def convert_person_ethnicity_data(person_df: pd.DataFrame, multi_diversity_df: pd.DataFrame) -> pd.DataFrame:
    ethnicity_converted = pd.DataFrame()

    multi_diversity_required_cols = ['EMPLID', 'LEGISLATIONCODE', 'ETHNICITY', 'PRIMARY_FLAG2']
    person_required_cols = ['PersonNumber', 'Country', 'PER_ETHNICITY', 'PrimaryFlag1'] # Updated to PrimaryFlag1

    if not multi_diversity_df.empty and validate_columns(multi_diversity_df, multi_diversity_required_cols, 'MultiDiversity'):
        ethnicity_df = multi_diversity_df.copy()
        ethnicity_converted['PersonNumber'] = ethnicity_df['EMPLID'].fillna('')
        ethnicity_converted['LegislationCode'] = ethnicity_df['LEGISLATIONCODE'].fillna('')
        ethnicity_converted['Ethnicity'] = ethnicity_df['ETHNICITY'].fillna('')
        ethnicity_converted['PrimaryFlag'] = ethnicity_df['PRIMARY_FLAG2'].fillna('')
    elif not person_df.empty and validate_columns(person_df, person_required_cols, 'Person'):
        ethnicity_df = person_df.rename(columns={'PersonNumber': 'PersonNumber'})
        ethnicity_converted['PersonNumber'] = ethnicity_df['PersonNumber']
        ethnicity_converted['LegislationCode'] = ethnicity_df['Country'].fillna('')
        ethnicity_converted['Ethnicity'] = ethnicity_df['PER_ETHNICITY'].fillna('')
        ethnicity_converted['PrimaryFlag'] = person_df.get('PrimaryFlag1', 'Y').fillna('Y') # Use .get for robustness, default to 'Y'
    else:
        logging.warning("Skipping PersonEthnicity data conversion due to missing critical source columns in both 'MultiDiversity' and 'Person' dataframes.")
        return pd.DataFrame(columns=['PersonNumber', 'LegislationCode', 'Ethnicity', 'PrimaryFlag'])

    return ethnicity_converted

# --- FastAPI Endpoint ---

@app.post("/convert-excel/", summary="Convert Excel Employee Data", response_description="Converted Excel file")
async def convert_excel_endpoint(excel_file: UploadFile = File(..., description="The Excel file (EmployeeWithHistory-Template.xlsx) to convert.")):
    """
    Receives an Excel file, processes its sheets, converts the data format,
    and returns a new Excel file with the converted data.
    """
    if not excel_file.filename.endswith('.xlsx'):
        logging.error(f"Invalid file type uploaded: {excel_file.filename}")
        raise HTTPException(status_code=400, detail="Invalid file type. Only .xlsx files are allowed.")

    try:
        # Read the incoming Excel file content into a BytesIO object
        file_content = await excel_file.read()
        file_stream = io.BytesIO(file_content)

        # Configuration for source Excel sheets (header rows are fixed based on template)
        sheets_config = {
            'Person': {'sheet_name': 'Person', 'header_row_index': 6}, 
            'WorkRelationship': {'sheet_name': 'WorkRelationship', 'header_row_index': 7}, 
            'Assignment': {'sheet_name': 'Assignment', 'header_row_index': 6}, 
            'MultiDiversity': {'sheet_name': 'MultiDiversity', 'header_row_index': 0}, 
            'Address': {'sheet_name': 'Address', 'header_row_index': 6}, 
            'Nat. ID Multi': {'sheet_name': 'Nat. ID Multi', 'header_row_index': 5}, 
            'Phone': {'sheet_name': 'Phone', 'header_row_index': 5}, 
            'Email': {'sheet_name': 'Email', 'header_row_index': 5}, 
            'Citizenship': {'sheet_name': 'Citizenship', 'header_row_index': 4}, 
        }

        source_dfs = {}
        for name, sheet_config in sheets_config.items():
            source_dfs[name] = read_excel_sheet_with_dynamic_header(
                file_stream,
                sheet_config['sheet_name'],
                sheet_config['header_row_index']
            )

        output_dfs = {}
        logging.info("Performing data transformations via API.")

        output_dfs['Worker'] = convert_worker_data(
            source_dfs.get('Person', pd.DataFrame()),
            source_dfs.get('WorkRelationship', pd.DataFrame())
        )

        output_dfs['PersonName'] = convert_person_name_data(
            source_dfs.get('Person', pd.DataFrame())
        )

        output_dfs['WorkRelationship'] = convert_work_relationship_data(
            source_dfs.get('WorkRelationship', pd.DataFrame()),
            source_dfs.get('Assignment', pd.DataFrame())
        )

        output_dfs['WorkTerms'] = convert_work_terms_data(
            source_dfs.get('WorkRelationship', pd.DataFrame()),
            source_dfs.get('Assignment', pd.DataFrame())
        )

        output_dfs['Assignment'] = convert_assignment_data(
            source_dfs.get('Assignment', pd.DataFrame())
        )

        output_dfs['Contract'] = convert_contract_data(
            source_dfs.get('Assignment', pd.DataFrame())
        )

        output_dfs['PersonNationalIdentifier'] = convert_national_id_data(
            source_dfs.get('Person', pd.DataFrame()),
            source_dfs.get('Nat. ID Multi', pd.DataFrame())
        )

        output_dfs['PersonReligion'] = convert_person_religion_data(
            source_dfs.get('Person', pd.DataFrame())
        )

        output_dfs['PersonAddress'] = convert_person_address_data(
            source_dfs.get('Address', pd.DataFrame())
        )

        output_dfs['PersonCitizenship'] = convert_person_citizenship_data(
            source_dfs.get('Citizenship', pd.DataFrame())
        )

        output_dfs['PersonEmail'] = convert_person_email_data(
            source_dfs.get('Email', pd.DataFrame())
        )
        
        output_dfs['PersonPhone'] = convert_person_phone_data(
            source_dfs.get('Phone', pd.DataFrame())
        )

        output_dfs['PersonEthnicity'] = convert_person_ethnicity_data(
            source_dfs.get('Person', pd.DataFrame()),
            source_dfs.get('MultiDiversity', pd.DataFrame())
        )

        # Write converted dataframes to a BytesIO object
        output_stream = io.BytesIO()
        with pd.ExcelWriter(output_stream, engine='openpyxl') as writer:
            for sheet_name, df in output_dfs.items():
                if not df.empty:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    logging.info(f"Prepared sheet '{sheet_name}' for output.")
                else:
                    logging.warning(f"Sheet '{sheet_name}' is empty, skipping writing to output Excel.")
        output_stream.seek(0) # Rewind to the beginning of the stream

        logging.info("Excel conversion successful. Sending file back.")
        return StreamingResponse(
            output_stream,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition": "attachment; filename=Converted_Employee_Data.xlsx"}
        )

    except Exception as e:
        logging.error(f"An unexpected error occurred during Excel processing: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Internal server error during processing: {e}")




#------------- Setup data storage ------------------#
class HDLSetupPayload(BaseModel):
    customerName: str
    instanceName: str
    hireActions: List[str]
    rehireActions: List[str]
    termActions: List[str]
    globalTransferActions: List[str]
    statusTypes: List[str]  # Must be length 6
    assignmentStatusRules: List[Dict[str, str]]

@app.post("/api/hdl/save-setup")
def save_hdl_setup(data: HDLSetupPayload):
    try:
        # Construct directory and filename
        setup_dir = Path("User/setup_files")
        setup_dir.mkdir(parents=True, exist_ok=True)

        # Normalize file name
        filename = f"{data.customerName.replace(' ', '_')}_{data.instanceName.replace(' ', '_')}_setup.json"
        filepath = setup_dir / filename

        # Save JSON
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data.dict(), f, indent=4)

        return {
            "message": "✅ Setup saved successfully",
            "file": str(filepath.resolve())
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error saving setup: {e}")
    

@app.get("/api/hdl/get-setup/{customer_name}/{instance_name}")
def get_hdl_setup(
    customer_name: str,
    instance_name: str 
):
    try:
        setup_dir = Path("User/setup_files")
        filename = f"{customer_name.replace(' ', '_')}_{instance_name.replace(' ', '_')}_setup.json"
        filepath = setup_dir / filename

        if not filepath.exists():
            raise HTTPException(status_code=404, detail="Setup file not found.")

        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)

        return JSONResponse(content=data)

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching setup: {e}")


class AttributeMappingPayload(BaseModel):
    customerName: str
    instanceName: str
    componentName: str
    mappedAttributes: Dict[str, str]

@app.post("/api/hdl/save-attribute-mapping")
def save_hdl_attribute_mapping(data: AttributeMappingPayload):
    try:
        # 🔐 Normalize folder and file path
        safe_customer = data.customerName.replace(" ", "_")
        safe_instance = data.instanceName.replace(" ", "_")
        safe_component = data.componentName.replace(" ", "_")

        mapping_dir = Path(f"User/attribute_mappings")
        mapping_dir.mkdir(parents=True, exist_ok=True)
        filename = f"{safe_customer.replace(' ', '_')}_{safe_instance.replace(' ', '_')}_{safe_component.replace(' ', '_')}_attributes.json"
        file_path = mapping_dir / filename

        # 💾 Save the mapping to JSON
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(data.mappedAttributes, f, indent=4)

        return {
            "message": "✅ Attribute mapping saved successfully.",
            "file": str(file_path.resolve())
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error saving attribute mapping: {e}")


@app.get("/api/hdl/get-attribute-mapping/{customer_name}/{instance_name}/{component_name}")
def get_hdl_attribute_mapping(customer_name: str, instance_name: str, component_name: str):
    try:
        safe_customer = customer_name.replace(" ", "_")
        safe_instance = instance_name.replace(" ", "_")
        safe_component = component_name.replace(" ", "_")
        filename = f"{safe_customer.replace(' ', '_')}_{safe_instance.replace(' ', '_')}_{safe_component.replace(' ', '_')}_attributes.json"
        file_path = Path(f"User/attribute_mappings/{filename}")

        if not file_path.exists():
            return {
                "success": 200,
                "create" : "create New"
            }

        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        return {
            "customerName": customer_name,
            "instanceName": instance_name,
            "componentName": component_name,
            "mappedAttributes": data
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error reading mapping: {e}")



class LookupDataAPIOracle(BaseModel):
    customerName: str
    instanceName: str

@app.post("/api/hdl/oracle_fetch/lookupdataload")
async def LookupDataLoading(req: LookupDataAPIOracle):
    customerName = req.customerName
    instanceName = req.instanceName

    oracle_env, username, password = [x.strip() for x in load_oracle_credentials(customerName, instanceName)]
    logger.warning(f"oracle credentials are, {oracle_env}, {username}, {password}")
    
    # Define SOAP save zone path
    soap_save_zone = Path(f"{customerName}/{instanceName}/soap_temp_storage")
    soap_save_zone.mkdir(parents=True, exist_ok=True)  # create dirs if they don't exist
    LookupData_Directory = Path(f"Required_files/{customerName}_{instanceName}_LookupData.xlsx")
    LookupData_Directory.parent.mkdir(parents=True, exist_ok=True)

    # SOAP service URL
    SOAP_URL = f"{oracle_env}/xmlpserver/services/ExternalReportWSSService?wsdl"

    headers = {
        "Content-Type": "application/soap+xml; charset=utf-8",
        "SOAPAction": "",
    }

    # Read from soap_request.xml
    with open("soap_request_Lookup_Data.xml", "r", encoding="utf-8") as file:
        soap_body = file.read().strip()

    # Save a copy of the SOAP request XML to the soap_save_zone
    try:
        saved_file_path = soap_save_zone / "soap_request_Lookup_Data_saved.xml"
        with open(saved_file_path, "w", encoding="utf-8") as f:
            f.write(soap_body)
        logger.info(f"SOAP request saved at: {saved_file_path}")
    except Exception as save_err:
        logger.error(f"Failed to save SOAP XML: {save_err}")

    # Now make the actual SOAP call
    try:
        async with httpx.AsyncClient(timeout=90) as client:
            #log everything
            logger.info(f"Making SOAP request to {SOAP_URL} with user {username}")
            response = await client.post(SOAP_URL, data=soap_body, headers=headers, auth=(username, password))
            response.raise_for_status()
            logger.info(f"SOAP request successful with status code {response.status_code}")
            # log all the response code as well as content
            logger.warning(f"Response content: {response.text[:500]}...")  # Log first 500 chars for brevity
            saved_file_path = soap_save_zone / "soap_response_saved.xml"
            with open(saved_file_path, "w", encoding="utf-8") as f:
                f.write(response.text)
            logger.info(f"SOAP request saved at: {saved_file_path}")

            excel_path = parse_soap_response_to_excel(saved_file_path, LookupData_Directory)
            return {
                "status": "success",
                "results": "Lookup data loaded successfully"
            }
    except httpx.HTTPError as e:
        logger.error(f"SOAP call failed: {str(e)}")
        if e.response:
            logger.error(f"Response content: {e.response.text}")
        raise HTTPException(status_code=500, detail=f"SOAP call failed: {str(e)}")

class mandatoryFieldsReqOracle(BaseModel):
    customerName: str
    instanceName: str

@app.post("/api/hdl/oracle_fetch/mandatoryFields")
async def MandatoryfieldsLoading(req: mandatoryFieldsReqOracle):
    customerName = req.customerName
    instanceName = req.instanceName

    oracle_env, username, password = [x.strip() for x in load_oracle_credentials(customerName, instanceName)]
    logger.warning(f"oracle credentials are, {oracle_env}, {username}, {password}")
    
    # Define SOAP save zone path
    soap_save_zone = Path(f"{customerName}/{instanceName}/soap_temp_storage")
    soap_save_zone.mkdir(parents=True, exist_ok=True)  # create dirs if they don't exist
    LookupData_Directory = Path(f"Required_files/{customerName}_{instanceName}_MandatoryFields.xlsx")
    LookupData_Directory.parent.mkdir(parents=True, exist_ok=True)

    # SOAP service URL
    SOAP_URL = f"{oracle_env}/xmlpserver/services/ExternalReportWSSService?wsdl"

    headers = {
        "Content-Type": "application/soap+xml; charset=utf-8",
        "SOAPAction": "",
    }

    # Read from soap_request.xml
    with open("soap_request_Mandatory_Fields.xml", "r", encoding="utf-8") as file:
        soap_body = file.read().strip()

    # Save a copy of the SOAP request XML to the soap_save_zone
    try:
        saved_file_path = soap_save_zone / "soap_request_Mandatory_Fields_saved.xml"
        with open(saved_file_path, "w", encoding="utf-8") as f:
            f.write(soap_body)
        logger.info(f"SOAP request saved at: {saved_file_path}")
    except Exception as save_err:
        logger.error(f"Failed to save SOAP XML: {save_err}")

    # Now make the actual SOAP call
    try:
        async with httpx.AsyncClient(timeout=90) as client:
            #log everything
            logger.info(f"Making SOAP request to {SOAP_URL} with user {username}")
            response = await client.post(SOAP_URL, data=soap_body, headers=headers, auth=(username, password))
            response.raise_for_status()
            logger.info(f"SOAP request successful with status code {response.status_code}")
            # log all the response code as well as content
            logger.warning(f"Response content: {response.text[:500]}...")  # Log first 500 chars for brevity
            saved_file_path = soap_save_zone / "soap_response_saved.xml"
            with open(saved_file_path, "w", encoding="utf-8") as f:
                f.write(response.text)
            logger.info(f"SOAP request saved at: {saved_file_path}")

            excel_path = parse_soap_response_to_excel(saved_file_path, LookupData_Directory)
            return {
                "status": "success",
                "results": "Mandatory Fields loaded successfully"
            }
    except httpx.HTTPError as e:
        logger.error(f"SOAP call failed: {str(e)}")
        if e.response:
            logger.error(f"Response content: {e.response.text}")
        raise HTTPException(status_code=500, detail=f"SOAP call failed: {str(e)}")




def parse_soap_response_to_excel(xml_path: str, output_excel_path: str = "output.xlsx", customerName: str = "", instanceName: str = "") -> str:
    """
    Parses SOAP response XML and saves embedded base64 Excel data as a file.
    
    Args:
        xml_path (str): Path to the response XML file.
        output_excel_path (str): Path to save the decoded Excel file.
    
    Returns:
        str: Path of the saved Excel file.
    """
    ns = {
        'env': 'http://www.w3.org/2003/05/soap-envelope',
        'ns2': 'http://xmlns.oracle.com/oxp/service/PublicReportService'
    }

    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()

        # Navigate to the reportBytes element
        report_bytes_elem = root.find('.//ns2:reportBytes', ns)
        if report_bytes_elem is None or not report_bytes_elem.text:
            raise ValueError("reportBytes not found or empty in the response.")

        # Decode base64 content
        decoded_excel = base64.b64decode(report_bytes_elem.text)

        # Save as Excel file
        output_path = Path(output_excel_path)
        output_path.write_bytes(decoded_excel)

        return str(output_path.resolve())

    except Exception as e:
        raise RuntimeError(f"Failed to extract Excel from SOAP XML: {e}")


@app.get("/api/lookupdata/available")
def get_available_lookupdata_files(customerName: str = "", instanceName: str = ""):
    try:
        lookupdata_dir = Path("Required_files")
        pattern = f"{customerName}_{instanceName}_LookupData.xlsx" if customerName and instanceName else "*_LookupData.xlsx"
        files = list(lookupdata_dir.glob(pattern))
        file_list = [f.name for f in files]
        return {
            "available": len(file_list) > 0,  # ✅ return availability flag
            "files": file_list
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to list files: {str(e)}")

@app.get("/api/mandatoryfields/available")
def get_available_mandatoryfields_files(customerName: str = "", instanceName: str = ""):
    try:
        mandatoryfields_dir = Path("Required_files")
        pattern = f"{customerName}_{instanceName}_MandatoryFields.xlsx" if customerName and instanceName else "*_MandatoryFields.xlsx"
        files = list(mandatoryfields_dir.glob(pattern))
        file_list = [f.name for f in files]
        return {
            "available": len(file_list) > 0,  # ✅ return availability flag
            "files": file_list
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to list files: {str(e)}")
        


ENV_DIR = Path("Required_files/env_store")
ENV_DIR.mkdir(parents=True, exist_ok=True)


# ✅ Get customers from env files
@app.get("/api/env/customers")
def get_all_env_customers():
    try:
        all_customers = []
        for env_file in ENV_DIR.glob("*.env"):
            env_data = dotenv_values(env_file)
            grouped = {}

            for key, value in env_data.items():
                if not "_" in key:
                    continue
                parts = key.split("_")
                if len(parts) < 3:
                    continue
                customer = parts[0]
                instance = parts[1]
                field = "_".join(parts[2:])

                if customer not in grouped:
                    grouped[customer] = {}
                if instance not in grouped[customer]:
                    grouped[customer][instance] = {"instanceName": instance}

                grouped[customer][instance][field.lower()] = value

            for customer, instances in grouped.items():
                instance_list = list(instances.values())
                all_customers.append({
                    "customerName": customer,
                    "assigned_instances": instance_list
                })

        return {"data": all_customers}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to load customers: {str(e)}")


# ✅ Merge all into .env.generated
@app.post("/api/save-env")
def merge_env_files(customers: List[Dict]):
    try:
        merged_lines = []
        added, skipped = 0, 0

        for customer in customers:
            customer_name = customer["customerName"].strip().upper().replace(" ", "_")
            for instance in customer["instances"]:
                inst_name = instance["instanceName"].strip().upper().replace(" ", "_")
                prefix = f"{customer_name}_{inst_name}"

                merged_lines.append(f"{prefix}_URL={instance['oracleUrl'].strip()}".strip())
                merged_lines.append(f"{prefix}_USERNAME={instance['oracleUsername'].strip()}".strip())
                merged_lines.append(f"{prefix}_PASSWORD={instance['oraclePassword'].strip()}".strip())

                added += 1

        logger.info(f"merged_lines to env are {merged_lines}")

        with open(".env", "w") as f:
            f.write("\n".join(merged_lines))

        return {"message": "All envs merged", "total_added": added, "total_skipped": skipped}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Merge failed: {str(e)}")



ENV_FILE = Path(".env")
# ✅ Delete customer
@app.delete("/api/customers/{customer_name}")
@app.delete("/api/customers/{customer_name}/{instance_name}")
def delete_customer(customer_name: str, instance_name: str = None):
    try:
        
        if not ENV_FILE.exists():
            raise HTTPException(status_code=404, detail="Env file not found")

        lines = ENV_FILE.read_text().splitlines()

        if instance_name:
            # Delete only that specific instance
            key_prefix = f"{customer_name.upper()}_{instance_name.upper()}_"
        else:
            # Delete all instances of that customer
            key_prefix = f"{customer_name.upper()}_"

        # Filter out matching lines
        new_lines = [line for line in lines if not line.startswith(key_prefix)]

        if len(lines) == len(new_lines):
            if instance_name:
                raise HTTPException(
                    status_code=404,
                    detail=f"Customer instance '{customer_name}_{instance_name}' not found"
                )
            else:
                raise HTTPException(
                    status_code=404,
                    detail=f"Customer '{customer_name}' not found"
                )

        # Write back updated file
        ENV_FILE.write_text("\n".join(new_lines) + "\n")

        if instance_name:
            return {"message": f"Customer instance '{customer_name}_{instance_name}' deleted successfully."}
        else:
            return {"message": f"All instances of customer '{customer_name}' deleted successfully."}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Delete failed: {str(e)}")


# ✅ Delete a specific instance from a customer
@app.delete("/api/customers/{customer_name}/instances/{instance_name}")
def delete_instance(customer_name: str, instance_name: str):
    try:
        customer_name = customer_name.upper().replace(" ", "_")
        instance_name = instance_name.upper().replace(" ", "_")
        env_file = ENV_DIR / f"{customer_name}.env"

        if not env_file.exists():
            raise HTTPException(status_code=404, detail="Customer not found")

        lines_to_keep = []
        with open(env_file, "r") as f:
            for line in f:
                if not line.startswith(f"{customer_name}_{instance_name}_"):
                    lines_to_keep.append(line.strip())

        with open(env_file, "w") as f:
            f.write("\n".join(lines_to_keep))

        return {"message": f"Instance '{instance_name}' deleted from customer '{customer_name}'."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Instance delete failed: {str(e)}")

@app.delete("/api/customers")
def delete_all_customers():
    try:
        deleted_files = []
        for env_file in ENV_DIR.glob("*.env"):
            env_file.unlink()
            deleted_files.append(env_file.name)

        return {"message": "All customer .env files deleted.", "deleted": deleted_files}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to delete all customers: {str(e)}")


@app.post("/api/admin/reset-system")
async def reset_system(
    reset_type: str = Query(..., description="Type of reset: 'soft' (keeps configs) or 'hard' (complete reset)"),
    admin_token: str = Query(..., description="Admin authentication token")
):
    """
    Reset the system for deployment/production. Cleans up temporary files and data.
    Use with caution as this will delete user data and temporary files.
    """
    # Simple admin token check (you should use a more secure method in production)
    expected_token = os.getenv("ADMIN_RESET_TOKEN", "reset123")
    if admin_token != expected_token:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid admin token"
        )

    if reset_type not in ["soft", "hard"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Reset type must be 'soft' or 'hard'"
        )

    try:
        reset_log = []
        
        # Define directories to clean
        directories_to_clean = [
            UPLOAD_DIR,
            VALIDATION_RESULTS_DIR,
            COMPLETED_FOLDER,
            BUNDLE_DEPOT_ZONE,
            Path("uploads/user")  # User saved component data
        ]

        # Files to keep (configuration files)
        files_to_keep = [
            EXCEL_FILE_PATH,
            TRANSFORMATION_ATTRIBUTES_FILE_PATH,
            USER_EXCEL_FILE_PATH,
            Path("Required_files/Mandatory Fields.xlsx"),
            Path("Required_files/Available_NLP.xlsx"),
            Path(".env")  # Keep .env file structure but clear sensitive data in hard reset
        ]

        if reset_type == "soft":
            # Soft reset - keep configuration files, remove temporary data
            reset_log.append("🔄 Performing SOFT reset...")
            
            for directory in directories_to_clean:
                if directory.exists():
                    # Remove contents but keep directory structure
                    for item in directory.iterdir():
                        if item.is_file():
                            item.unlink()
                            reset_log.append(f"🗑️ Deleted file: {item}")
                        elif item.is_dir():
                            shutil.rmtree(item)
                            reset_log.append(f"🗑️ Deleted directory: {item}")
                    reset_log.append(f"✅ Cleaned directory: {directory}")

            # Clear in-memory data
            global pass_df, fail_df, USER_DB
            pass_df = pd.DataFrame()
            fail_df = pd.DataFrame()
            USER_DB = load_user_data(USER_EXCEL_FILE_PATH)  # Reload user data
            
            reset_log.append("✅ Cleared in-memory data structures")

        elif reset_type == "hard":
            # Hard reset - remove everything except essential configuration
            reset_log.append("🔥 Performing HARD reset...")
            combos = extract_customer_instance_names_from_env(env_path=Path(".env"))
            customerName = combos[0][0] if combos else ""
            instanceName = combos[0][1] if combos else ""
            reset_log.append(f"Identified customer: {customerName}, instance: {instanceName}")
            #Remove all user-specific LookupData files
            lookupdata_pattern = f"{customerName}_{instanceName}_LookupData.xlsx"
            mandatory_pattern = f"{customerName}_{instanceName}_MandatoryFields.xlsx"
            try: 
                for file in Path("Required_files").glob(lookupdata_pattern):
                    file.unlink()
                    reset_log.append(f"🗑️ Deleted LookupData file: {file}")
            except Exception as e:
                reset_log.append(f"⚠️ Failed to delete LookupData files: {e}")
            
            try:
                for file in Path("Required_files").glob(mandatory_pattern):
                    file.unlink()
                    reset_log.append(f"🗑️ Deleted MandatoryFields file: {file}")
            except Exception as e:
                reset_log.append(f"⚠️ Failed to delete MandatoryFields files: {e}")
            

            # Clear environment variables from memory
            reset_log.append("🧹 Clearing environment variables from memory...")
            env_vars_to_preserve = {
                'ADMIN_RESET_TOKEN': os.getenv('ADMIN_RESET_TOKEN'),
                'PYTHONPATH': os.getenv('PYTHONPATH'),
                'PATH': os.getenv('PATH')
            }
            
            # Clear all environment variables in current process
            os.environ.clear()
            
            # Restore essential variables
            for key, value in env_vars_to_preserve.items():
                if value:
                    os.environ[key] = value
            
            reset_log.append("✅ Cleared environment variables from memory")

            # Reset .env file to default template (keep structure but clear sensitive data)
            env_file = Path(".env")
            if env_file.exists():
                # Create a minimal .env template
                minimal_env_content = """"""
                env_file.write_text(minimal_env_content)
                reset_log.append("🔧 Reset .env file to minimal template")

            # Clean directories
            for directory in directories_to_clean:
                if directory.exists():
                    shutil.rmtree(directory)
                    directory.mkdir(parents=True, exist_ok=True)
                    reset_log.append(f"🔥 Recreated directory: {directory}")

            # Recreate essential directories
            essential_dirs = [
                UPLOAD_DIR,
                VALIDATION_RESULTS_DIR,
                COMPLETED_FOLDER,
                BUNDLE_DEPOT_ZONE,
                DAT_FILES_DIR,
                Path("uploads/user")
            ]
            
            for dir_path in essential_dirs:
                dir_path.mkdir(parents=True, exist_ok=True)
                reset_log.append(f"📁 Ensured directory exists: {dir_path}")

            # Clear all in-memory data
            USER_DB = {}  # Clear user database
            
            # Reload essential configuration files if they exist
            if USER_EXCEL_FILE_PATH.exists():
                USER_DB = load_user_data(USER_EXCEL_FILE_PATH)
            
            reset_log.append("✅ Cleared all in-memory data structures")

        # Log system information
        reset_log.append(f"📊 System reset completed at: {datetime.now().isoformat()}")
        reset_log.append(f"🔧 Reset type: {reset_type}")
        reset_log.append(f"🐍 Python version: {sys.version}")
        
        # Log directory sizes for monitoring
        for dir_path in [UPLOAD_DIR, VALIDATION_RESULTS_DIR, BUNDLE_DEPOT_ZONE]:
            if dir_path.exists():
                size = sum(f.stat().st_size for f in dir_path.rglob('*') if f.is_file())
                reset_log.append(f"💾 {dir_path.name} size: {size / (1024*1024):.2f} MB")

        return {
            "status": "success",
            "message": f"System reset completed successfully ({reset_type} reset)",
            "reset_type": reset_type,
            "timestamp": datetime.now().isoformat(),
            "details": reset_log
        }

    except Exception as e:
        logger.error(f"System reset failed: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"System reset failed: {str(e)}"
        )


@app.get("/api/admin/system-status")
async def get_system_status(admin_token: str = Query(...)):
    """
    Get current system status and disk usage information.
    """
    expected_token = os.getenv("ADMIN_RESET_TOKEN", "reset123")
    if admin_token != expected_token:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid admin token"
        )

    try:
        status_info = {
            "timestamp": datetime.now().isoformat(),
            "system_directories": {},
            "file_counts": {},
            "total_disk_usage_mb": 0,
            "important_files_exist": {}
        }

        # Check important directories
        important_dirs = {
            "uploads": UPLOAD_DIR,
            "validation_results": VALIDATION_RESULTS_DIR,
            "bundle_depot": BUNDLE_DEPOT_ZONE,
            "completed_files": COMPLETED_FOLDER,
            "dat_files": DAT_FILES_DIR
        }

        for name, dir_path in important_dirs.items():
            if dir_path.exists():
                files = list(dir_path.rglob("*"))
                file_count = len([f for f in files if f.is_file()])
                total_size = sum(f.stat().st_size for f in files if f.is_file())
                
                status_info["system_directories"][name] = {
                    "exists": True,
                    "file_count": file_count,
                    "size_mb": total_size / (1024 * 1024),
                    "path": str(dir_path)
                }
                status_info["total_disk_usage_mb"] += total_size / (1024 * 1024)
            else:
                status_info["system_directories"][name] = {
                    "exists": False,
                    "file_count": 0,
                    "size_mb": 0,
                    "path": str(dir_path)
                }

        # Check important configuration files
        important_files = {
            "hierarchy_excel": EXCEL_FILE_PATH,
            "transformation_attributes": TRANSFORMATION_ATTRIBUTES_FILE_PATH,
            "user_database": USER_EXCEL_FILE_PATH,
            "mandatory_fields": Path("Required_files/Mandatory Fields.xlsx"),
            "nlp_rules": Path("Required_files/Available_NLP.xlsx"),
            "environment_file": Path(".env")
        }

        for name, file_path in important_files.items():
            status_info["important_files_exist"][name] = {
                "exists": file_path.exists(),
                "path": str(file_path),
                "size_mb": file_path.stat().st_size / (1024 * 1024) if file_path.exists() else 0
            }

        # Add memory usage info
        import psutil
        process = psutil.Process()
        memory_info = process.memory_info()
        status_info["memory_usage_mb"] = memory_info.rss / (1024 * 1024)
        
        # Add system info
        status_info["python_version"] = sys.version
        status_info["platform"] = sys.platform

        return status_info

    except Exception as e:
        logger.error(f"Failed to get system status: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to get system status: {str(e)}"
        )
    
#------------- HDL Job Management ------------------#
DATA_FILE = Path("hdl_jobs.json")


class HDLJob(BaseModel):
    id: Optional[int] = None
    component: str
    fileName: str
    timeCreated: Optional[datetime] = None
    contentId: Optional[str] = None
    requestId: Optional[str] = None
    status: str = "-"
    oracleJobSummary: Optional[str] = None


# -----------------------------
# Persistence Layer
# -----------------------------
def load_data():
    """Safely load jobs data from JSON file"""
    if not DATA_FILE.exists() or DATA_FILE.stat().st_size == 0:
        return {}
    with open(DATA_FILE, "r") as f:
        try:
            return json.load(f)
        except json.JSONDecodeError:
            return {}


def save_data(data):
    """Persist jobs data to JSON file"""
    with open(DATA_FILE, "w") as f:
        json.dump(data, f, indent=2, default=str)


# -----------------------------
# API Endpoints
# -----------------------------
@app.get("/api/hdl/fetchdata/{customer}/{instance}", response_model=List[HDLJob])
async def fetch_data(customer: str, instance: str):
    """Fetch jobs for given customer + instance"""
    data = load_data()
    jobs = data.get(customer, {}).get(instance, [])
    return [HDLJob(**j) for j in jobs]


@app.post("/api/hdl/getdata/{customer}/{instance}")
async def add_data(customer: str, instance: str, job: HDLJob):
    """Add new job for customer+instance"""
    data = load_data()

    if customer not in data:
        data[customer] = {}
    if instance not in data[customer]:
        data[customer][instance] = []

    jobs = data[customer][instance]

    # Auto-assign ID if not provided
    if not job.id:
        max_id = max((j.get("id", 0) for j in jobs), default=0)
        job.id = max_id + 1

    # Add creation timestamp if missing
    if not job.timeCreated:
        job.timeCreated = datetime.utcnow()

    jobs.append(job.dict())
    save_data(data)
    return {"message": "Job added", "job": job}


@app.put("/api/hdl/updatedata/{customer}/{instance}/{job_id}")
async def update_data(customer: str, instance: str, job_id: int, job_update: HDLJob):
    """Update existing job by ID"""
    data = load_data()

    if customer not in data or instance not in data[customer]:
        raise HTTPException(status_code=404, detail="Customer/Instance not found")

    jobs = data[customer][instance]
    for i, j in enumerate(jobs):
        if j["id"] == job_id:
            # Preserve ID if missing from update payload
            if job_update.id is None:
                job_update.id = job_id
            jobs[i] = job_update.dict()
            save_data(data)
            return {"message": "Job updated", "job": job_update}

    raise HTTPException(status_code=404, detail="Job not found")